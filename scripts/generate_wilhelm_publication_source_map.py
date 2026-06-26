import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "地图_中国故事集_出版地和故事来源地.xlsx"
if not SOURCE.exists():
    SOURCE = ROOT / "地图_中国故事集_出版地和故事来源地.xlsx"
OUTPUT = ROOT / "frontend" / "src" / "data" / "wilhelmPublicationSourceMap.json"

PROVINCE_COORDS = {
    "北京": [116.4, 39.9],
    "天津": [117.2, 39.12],
    "河北": [114.5, 38.04],
    "山西": [112.55, 37.87],
    "内蒙古": [111.67, 40.82],
    "辽宁": [123.43, 41.8],
    "吉林": [125.32, 43.9],
    "黑龙江": [126.64, 45.76],
    "上海": [121.47, 31.23],
    "江苏": [118.8, 32.1],
    "浙江": [120.2, 30.3],
    "安徽": [117.27, 31.86],
    "福建": [119.3, 26.08],
    "江西": [115.86, 28.68],
    "山东": [117.0, 36.7],
    "河南": [113.62, 34.75],
    "湖北": [114.3, 30.6],
    "湖南": [112.98, 28.2],
    "广东": [113.27, 23.13],
    "广西": [108.32, 22.82],
    "海南": [110.35, 20.02],
    "重庆": [106.55, 29.56],
    "四川": [104.06, 30.67],
    "贵州": [106.63, 26.65],
    "云南": [102.71, 25.04],
    "西藏": [91.13, 29.65],
    "陕西": [108.94, 34.34],
    "甘肃": [103.82, 36.06],
    "青海": [101.78, 36.62],
    "宁夏": [106.27, 38.47],
    "新疆": [87.62, 43.82],
    "台湾": [121.0, 23.7],
    "香港": [114.17, 22.32],
    "澳门": [113.55, 22.2],
}

CITY_COORDS = {
    "Berlin": [13.405, 52.52],
    "Jena": [11.59, 50.93],
    "München": [11.58, 48.14],
    "Munich": [11.58, 48.14],
    "Leipzig": [12.37, 51.34],
    "Frankfurt am Main": [8.68, 50.11],
    "Frankfurt": [8.68, 50.11],
    "Stuttgart": [9.18, 48.78],
    "Basel": [7.59, 47.56],
    "Sankt Augustin": [7.19, 50.78],
    "Esslingen": [9.31, 48.74],
    "Norderstedt": [9.98, 53.71],
    "Bickenbach": [8.62, 49.76],
    "Hamburg": [9.99, 53.55],
    "Köln": [6.96, 50.94],
    "Cologne": [6.96, 50.94],
    "Düsseldorf": [6.77, 51.23],
    "Freiburg": [7.85, 47.99],
    "Eisenach": [10.32, 50.98],
    "Kassel": [9.49, 51.31],
    "Zürich": [8.54, 47.38],
    "Prag": [14.42, 50.08],
    "Wien": [16.37, 48.21],
    "Bayreuth": [11.58, 49.95],
    "Meerbusch": [6.69, 51.25],
    "Augsburg": [10.9, 48.37],
    "Bielefeld": [8.53, 52.02],
    "Schiedlberg": [14.27, 48.1],
    "Kreuzlingen": [9.18, 47.65],
}

COUNTRY_BY_CITY = {
    "Zürich": "Switzerland",
    "Basel": "Switzerland",
    "Kreuzlingen": "Switzerland",
    "Wien": "Austria",
    "Prag": "Czech Republic",
}


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def base_province(value):
    text = clean(value)
    if "哈尔滨" in text:
        return "黑龙江"
    for name in [
        "内蒙古",
        "黑龙江",
        "广西",
        "宁夏",
        "新疆",
        "西藏",
        "香港",
        "澳门",
        "台湾",
        "北京",
        "天津",
        "上海",
        "重庆",
        "河北",
        "山西",
        "辽宁",
        "吉林",
        "江苏",
        "浙江",
        "安徽",
        "福建",
        "江西",
        "山东",
        "河南",
        "湖北",
        "湖南",
        "广东",
        "海南",
        "四川",
        "贵州",
        "云南",
        "陕西",
        "甘肃",
        "青海",
    ]:
        if name in text:
            return name
    return re.sub(r"(省|市|自治区|特别行政区|壮族|回族|维吾尔|北部|中部|东部|地区)", "", text).strip()


def row_center(row, province):
    values = []
    for key in ["min_lon", "max_lon", "min_lat", "max_lat"]:
        try:
            values.append(float(row.get(key)))
        except Exception:
            values.append(None)
    if all(value is not None for value in values):
        min_lon, max_lon, min_lat, max_lat = values
        return [(min_lon + max_lon) / 2, (min_lat + max_lat) / 2]
    return PROVINCE_COORDS.get(province)


def city_candidates(value):
    raw = clean(value).replace("u.a.", "").strip()
    parts = [clean(part) for part in re.split(r"\s*/\s*|;|,", raw) if clean(part)]
    return [part for part in parts if part in CITY_COORDS]


def read_sheet(workbook, sheet_name):
    worksheet = workbook[sheet_name]
    headers = [clean(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        yield {headers[index]: row[index] for index in range(len(headers))}


def main():
    workbook = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)

    regions = {}
    for row in read_sheet(workbook, "Sheet1"):
        region = clean(row.get("source region"))
        province = base_province(row.get("province"))
        center = row_center(row, province)
        if region and province and center:
            regions.setdefault(region, []).append({
                "province": province,
                "sourceProvince": clean(row.get("province")),
                "coords": center,
            })

    for alias, target in [
        ("主要源于山东省", "山东"),
        ("牧牛人的故事以及来自斯基德格朗和丁日的传说", "西藏"),
    ]:
        if target in regions and alias not in regions:
            regions[alias] = regions[target]

    records = []
    flows = []
    skipped = []

    for row_index, row in enumerate(read_sheet(workbook, "工作表1"), start=2):
        source_region = clean(row.get("source region（如何呈现？）"))
        if not source_region:
            continue
        cities = city_candidates(row.get("city"))
        mapped_provinces = regions.get(source_region, [])
        if not cities or not mapped_provinces:
            skipped.append({
                "row": row_index,
                "title": clean(row.get("title")),
                "city": clean(row.get("city")),
                "sourceRegion": source_region,
                "reason": "missing city coords or region mapping",
            })
            continue

        for city in cities:
            country = COUNTRY_BY_CITY.get(city) or clean(row.get("country")) or "Germany"
            year_match = re.search(r"\d{4}", str(row.get("year") or ""))
            year = int(year_match.group(0)) if year_match else 0
            title_cn = clean(row.get("title(Chinese)"))
            title = clean(row.get("title"))
            for mapped in mapped_provinces:
                record_id = f"excel-map-{row_index}-{city}-{mapped['province']}"
                record = {
                    "id": record_id,
                    "source": "地图_中国故事集_出版地和故事来源地.xlsx#工作表1+Sheet1",
                    "title": title_cn or title,
                    "foreignTitle": title,
                    "year": year,
                    "yearText": str(year) if year else "未标年",
                    "city": city,
                    "country": country,
                    "publisher": clean(row.get("publisher")),
                    "province": mapped["province"],
                    "sourceRegion": source_region,
                    "sourceProvince": mapped["sourceProvince"],
                    "from": mapped["coords"],
                    "to": CITY_COORDS[city],
                    "language": "德语",
                }
                records.append(record)
                flows.append({
                    "id": record_id,
                    "title": record["title"],
                    "sectionId": "stories",
                    "resourceType": "中国故事集出版地和故事来源地",
                    "language": "德语",
                    "year": year,
                    "from": mapped["coords"],
                    "to": CITY_COORDS[city],
                    "fromLabel": mapped["province"],
                    "toLabel": f"{city} · {country}",
                    "province": mapped["province"],
                    "sourceRegion": source_region,
                    "sourceProvince": mapped["sourceProvince"],
                    "country": country,
                    "weight": 1,
                })

    OUTPUT.write_text(
        json.dumps(
            {
                "sourceWorkbook": SOURCE.name,
                "records": records,
                "flows": flows,
                "skipped": skipped,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"records": len(records), "flows": len(flows), "skipped": skipped}, ensure_ascii=False))


if __name__ == "__main__":
    main()
