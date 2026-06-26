from __future__ import annotations

import collections
import json
import math
import re
from itertools import combinations
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
PUBLIC_ASSETS = ROOT / "frontend" / "public" / "assets"
DIST_ASSETS = ROOT / "frontend" / "dist" / "assets"


def asset_path(filename: str) -> Path:
    data_path = DATA_DIR / filename
    if data_path.exists():
        return data_path
    public_path = PUBLIC_ASSETS / filename
    if public_path.exists():
        return public_path
    legacy_names = {
        "中国故事集总表_知识库(1).xlsx": "中国故事集总表_知识库.xlsx",
    }
    if filename in legacy_names:
        data_alias = DATA_DIR / legacy_names[filename]
        if data_alias.exists():
            return data_alias
        public_alias = PUBLIC_ASSETS / legacy_names[filename]
        if public_alias.exists():
            return public_alias
    return DIST_ASSETS / filename
OUT = ROOT / "frontend" / "src" / "data" / "storyCollections.json"

COLLECTION_FILE = asset_path("中国故事集总表_知识库(1).xlsx")
CHILD_FILE = asset_path("中国故事集_子故事（3533篇）.xlsx")
PREFACE_FILE = asset_path("中国故事集_序跋.xlsx")
SOURCE_MAP_FILE = asset_path("地图_中国故事集_故事来源地.xlsx")
WILHELM_TEXT_FILE = asset_path("中国民间童话.xlsx")
WILHELM_MAP_FILE = asset_path("地图_中国民间童话.xlsx")

COUNTRY_COORDS = {
    "Germany": [10.45, 51.16],
    "France": [2.35, 48.86],
    "United Kingdom": [-0.13, 51.51],
    "UK": [-0.13, 51.51],
    "United States": [-74.01, 40.71],
    "USA": [-74.01, 40.71],
    "Switzerland": [8.54, 47.37],
    "Austria": [16.37, 48.2],
    "Czechia": [14.43, 50.08],
    "China": [116.4, 39.9],
}

CITY_COORDS = {
    "Berlin": [13.405, 52.52],
    "Jena": [11.59, 50.93],
    "Munich": [11.58, 48.14],
    "München": [11.58, 48.14],
    "Leipzig": [12.37, 51.34],
    "Frankfurt": [8.68, 50.11],
    "Hamburg": [9.99, 53.55],
    "Stuttgart": [9.18, 48.78],
    "Basel": [7.59, 47.56],
    "Zürich": [8.54, 47.38],
    "Prag": [14.42, 50.08],
    "Wien": [16.37, 48.21],
    "Peking": [116.4, 39.9],
    "Beijing": [116.4, 39.9],
    "Shanghai": [121.47, 31.23],
}

CITY_ZH = {
    "Berlin": "柏林",
    "Jena": "耶拿",
    "Munich": "慕尼黑",
    "München": "慕尼黑",
    "Leipzig": "莱比锡",
    "Frankfurt": "法兰克福",
    "Hamburg": "汉堡",
    "Stuttgart": "斯图加特",
    "Basel": "巴塞尔",
    "Zürich": "苏黎世",
    "Prag": "布拉格",
    "Wien": "维也纳",
    "Peking": "北京",
    "Beijing": "北京",
    "Shanghai": "上海",
}

PROVINCE_COORDS = {
    "北京": [116.4, 39.9],
    "天津": [117.2, 39.12],
    "河北": [114.5, 38.04],
    "山西": [112.55, 37.87],
    "内蒙古": [111.67, 40.82],
    "辽宁": [123.43, 41.8],
    "吉林": [125.32, 43.9],
    "黑龙江": [126.64, 45.76],
    "上海": [121.47, 31.23],
    "江苏": [118.8, 32.1],
    "浙江": [120.2, 30.3],
    "安徽": [117.27, 31.86],
    "福建": [119.3, 26.08],
    "江西": [115.86, 28.68],
    "山东": [117.0, 36.7],
    "河南": [113.62, 34.75],
    "湖北": [114.3, 30.6],
    "湖南": [112.98, 28.2],
    "广东": [113.27, 23.13],
    "广西": [108.32, 22.82],
    "海南": [110.35, 20.02],
    "重庆": [106.55, 29.56],
    "四川": [104.06, 30.67],
    "贵州": [106.63, 26.65],
    "云南": [102.71, 25.04],
    "西藏": [91.13, 29.65],
    "陕西": [108.94, 34.34],
    "甘肃": [103.82, 36.06],
    "青海": [101.78, 36.62],
    "宁夏": [106.27, 38.47],
    "新疆": [87.62, 43.82],
    "台湾": [121.0, 23.7],
    "香港": [114.17, 22.32],
    "澳门": [113.55, 22.2],
}

THEME_LEXICON = {
    "动物": ["龙", "虎", "蛇", "狐", "狐狸", "牛", "马", "鸟", "鱼", "龟", "狗", "猫", "鼠", "猴", "兔", "鹰", "燕", "蚊"],
    "气象": ["雨", "雷", "风", "云", "月", "太阳", "星", "天", "雪", "雾", "霜"],
    "神怪": ["神", "仙", "鬼", "妖", "怪", "魔", "龙王", "女娲", "嫦娥"],
    "人物身份": ["皇帝", "国王", "公主", "农夫", "和尚", "道士", "猎人", "妻子", "母亲", "父亲", "兄弟"],
    "伦理": ["孝", "报恩", "贪婪", "善良", "惩罚", "复仇", "智慧", "忠诚"],
    "空间": ["山", "海", "河", "井", "桥", "村", "城", "宫", "天宫", "田"],
}

STOP_WORDS = {
    "中国", "故事", "民间", "童话", "序言", "前言", "后记", "文本", "作者", "版本", "一个", "一些", "我们", "他们", "人们",
    "chinesische", "märchen", "sagen", "und", "der", "die", "das", "aus", "von", "mit",
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def year(value) -> int | None:
    match = re.search(r"\d{4}", clean(value))
    return int(match.group()) if match else None


def number(value) -> int:
    match = re.search(r"\d+", clean(value))
    return int(match.group()) if match else 0


def strip_book_marks(value: str) -> str:
    return clean(value).strip("《》")


def norm(value: str) -> str:
    value = clean(value).lower()
    return re.sub(r"[《》（）()\s:：,，.;；·\"“”‘’\-]+", "", value)


def rows_from_xlsx(path: Path, sheet_name: str | None = None) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    raw_headers = next(sheet.iter_rows(values_only=True))
    headers = [clean(item) or f"字段{index + 1}" for index, item in enumerate(raw_headers)]
    rows = []
    for values in sheet.iter_rows(values_only=True, min_row=2):
        row = {headers[index]: clean(values[index]) if index < len(values) else "" for index in range(len(headers))}
        if any(row.values()):
            rows.append(row)
    return rows


def read_prefaces() -> list[dict]:
    rows = rows_from_xlsx(PREFACE_FILE)
    prefaces = []
    for index, row in enumerate(rows, start=1):
        title = row.get("故事集标题", "")
        prefaces.append(
            {
                "id": f"preface-{index:03d}",
                "collectionTitle": title,
                "year": year(row.get("年份")) or 0,
                "yearText": row.get("年份", ""),
                "intro": row.get("故事集介绍", ""),
                "author": row.get("序跋作者", ""),
                "type": row.get("序跋类型", ""),
                "text": row.get("序跋文本", ""),
            }
        )
    return prefaces


def read_children() -> list[dict]:
    rows = rows_from_xlsx(CHILD_FILE)
    children = []
    for index, row in enumerate(rows, start=1):
        book = row.get("故事集标题", "")
        canonical = strip_book_marks(row.get("规范故事名", ""))
        child_title = row.get("子故事标题", "")
        if not (book or canonical or child_title):
            continue
        children.append(
            {
                "id": f"child-{index:04d}",
                "sheet": "子故事",
                "ethnicity": row.get("故事民族来源", ""),
                "storyType": "",
                "creator": "",
                "canonicalName": canonical,
                "variantName": child_title,
                "year": year(row.get("年份")) or 0,
                "yearText": row.get("年份", ""),
                "translator": "",
                "reference": "",
                "nationality": "",
                "language": "德语",
                "translationMode": "",
                "carrier": "图书",
                "bookName": book,
                "subtitle": "",
                "journalIssue": "",
                "editor": "",
                "country": "德国",
                "place": "",
                "publisher": "",
                "version": "",
                "versionNote": "",
                "notes": row.get("故事民族来源", ""),
                "url": "",
            }
        )
    return children


def source_region_map() -> dict[str, str]:
    result = {}
    for row in rows_from_xlsx(SOURCE_MAP_FILE):
        title = row.get("title", "")
        region = row.get("source region", "")
        if title and region:
            result[norm(title)] = region
    return result


def province_from_region(region: str) -> str:
    text = clean(region)
    for province in PROVINCE_COORDS:
        if province in text:
            return province
    if "港澳台" in text:
        return "香港"
    return "北京"


def read_collections(children: list[dict], prefaces: list[dict]) -> list[dict]:
    child_ids_by_book: dict[str, list[str]] = collections.defaultdict(list)
    for child in children:
        child_ids_by_book[norm(child["bookName"])].append(child["id"])

    preface_by_title = {}
    for preface in prefaces:
        preface_by_title.setdefault(norm(preface["collectionTitle"]), preface)

    source_by_title = source_region_map()
    items = []
    for index, row in enumerate(rows_from_xlsx(COLLECTION_FILE), start=1):
        title = row.get("故事集标题", "")
        chinese = strip_book_marks(row.get("故事集标题（中文）", ""))
        key = norm(title)
        preface = preface_by_title.get(key, {})
        region = source_by_title.get(key, "")
        city = row.get("城市", "")
        publisher = row.get("出版社", "")
        publisher_text = f"{city}: {publisher}" if city and publisher else publisher or city
        items.append(
            {
                "id": f"story-collection-{index:03d}",
                "name": title,
                "chineseTitle": chinese,
                "foreignTitle": title,
                "year": year(row.get("年份")) or 0,
                "yearText": row.get("年份", ""),
                "editor": row.get("译者/编者", ""),
                "editorRole": row.get("译者/编者身份", ""),
                "prefaceAuthor": row.get("序跋作者", "") or preface.get("author", ""),
                "prefaceText": preface.get("text", ""),
                "prefaceType": preface.get("type", ""),
                "prefaceIntro": preface.get("intro", ""),
                "country": row.get("国家", ""),
                "publisher": publisher_text,
                "city": city,
                "declaredChildCount": number(row.get("子故事数量")),
                "sourceRegion": region,
                "sourceProvince": province_from_region(region),
                "matchedChildIds": child_ids_by_book.get(key, []),
            }
        )
    return items


def country_coord(country: str) -> list[float]:
    return COUNTRY_COORDS.get(country, COUNTRY_COORDS["Germany"])


def build_flows(collection_items: list[dict]) -> list[dict]:
    flows = []
    for item in collection_items:
        province = item.get("sourceProvince") or "北京"
        city = item.get("city") or ""
        country = item.get("country") or "Germany"
        to = CITY_COORDS.get(city, country_coord(country))
        flows.append(
            {
                "id": item["id"],
                "title": item["name"],
                "sectionId": "stories",
                "resourceType": "故事集",
                "language": "德语",
                "year": item["year"],
                "from": PROVINCE_COORDS.get(province, PROVINCE_COORDS["北京"]),
                "to": to,
                "fromLabel": item.get("sourceRegion") or province,
                "toLabel": f"{CITY_ZH.get(city, city) or country} · {country}",
                "province": province,
                "country": country,
                "weight": max(0.65, min(3.2, (len(item["matchedChildIds"]) or item["declaredChildCount"] or 1) / 24)),
            }
        )
    return flows


def read_wilhelm_stories() -> list[dict]:
    rows = rows_from_xlsx(WILHELM_TEXT_FILE)
    stories = []
    for index, row in enumerate(rows, start=1):
        title = row.get("单篇译文故事名", "")
        text = row.get("译文内容", "")
        if not (title or text):
            continue
        if title.strip().lower() in {"序言", "前言", "preface", "vorwort", "einleitung"}:
            continue
        stories.append(
            {
                "id": f"wilhelm-story-{index:03d}",
                "title": title,
                "text": text,
                "source": row.get("故事来源", ""),
                "category": row.get("卫礼贤的分类", "") or "未分类",
                "year": 1914,
            }
        )
    return stories


def read_wilhelm_editions() -> list[dict]:
    rows = rows_from_xlsx(WILHELM_MAP_FILE)
    editions = []
    for index, row in enumerate(rows, start=1):
        city = row.get("city", "")
        country = row.get("country", "Germany")
        title = row.get("titel", "Chinesische Volksmärchen")
        editions.append(
            {
                "id": f"wilhelm-edition-{index:03d}",
                "source": "地图_中国民间童话.xlsx",
                "title": "卫礼贤《中国民间童话》" if title == "Chinesische Volksmärchen" else title,
                "foreignTitle": title,
                "year": year(row.get("year")) or 0,
                "yearText": row.get("year", ""),
                "edition": row.get("全/选/改编", ""),
                "translator": "Richard Wilhelm（卫礼贤）",
                "publisher": row.get("publisher", ""),
                "city": city,
                "country": country,
                "province": "山东",
                "language": row.get("语种", "德语"),
                "note": f"{row.get('全/选/改编', '')} · {row.get('语种', '')}".strip(" ·"),
            }
        )
    return editions


def story_terms(story: dict) -> list[dict]:
    text = f"{story.get('title', '')} {story.get('text', '')}"
    terms = []
    seen = set()
    for category, words in THEME_LEXICON.items():
        for word in words:
            count = len(re.findall(re.escape(word), text))
            if count and word not in seen:
                seen.add(word)
                terms.append({"term": word, "category": category, "count": count})
    return terms


def wilhelm_theme_graph(stories: list[dict]) -> dict:
    term_counts = collections.Counter()
    category_counts = collections.Counter()
    story_links: dict[str, list[dict]] = collections.defaultdict(list)
    pair_counts = collections.Counter()

    for story in stories:
        terms = story_terms(story)
        term_names = [item["term"] for item in terms]
        for item in terms:
            term_counts[item["term"]] += item["count"]
            category_counts[item["category"]] += item["count"]
            story_links[item["term"]].append(
                {
                    "storyId": story["id"],
                    "storyTitle": story["title"],
                    "category": item["category"],
                    "count": item["count"],
                }
            )
        for source, target in combinations(sorted(set(term_names)), 2):
            pair_counts[(source, target)] += 1

    top_terms = [term for term, _ in term_counts.most_common(24)]
    nodes = [{"id": "center", "label": "卫礼贤《中国民间童话》", "type": "故事集", "count": len(stories)}]
    nodes.extend({"id": f"cat-{cat}", "label": cat, "type": "主题类别", "count": count} for cat, count in category_counts.most_common())
    nodes.extend({"id": f"term-{term}", "label": term, "type": "关键词", "count": term_counts[term]} for term in top_terms)
    edges = []
    for cat, count in category_counts.most_common():
        edges.append({"source": "center", "target": f"cat-{cat}", "weight": count, "relation": "主题类别"})
    for term in top_terms:
        category = next((item["category"] for story in stories for item in story_terms(story) if item["term"] == term), "主题")
        edges.append({"source": f"cat-{category}", "target": f"term-{term}", "weight": term_counts[term], "relation": "关键词"})
    for (source, target), count in pair_counts.most_common(34):
        if source in top_terms and target in top_terms:
            edges.append({"source": f"term-{source}", "target": f"term-{target}", "weight": count, "relation": "共现"})

    return {
        "nodes": nodes,
        "edges": edges,
        "terms": [
            {
                "term": term,
                "category": next((link["category"] for link in story_links[term]), ""),
                "count": count,
                "stories": sorted(story_links[term], key=lambda item: item["count"], reverse=True)[:12],
            }
            for term, count in term_counts.most_common(36)
        ],
    }


def wilhelm_single_story_graph(story: dict) -> dict:
    terms = story_terms(story)
    category_counts = collections.Counter(item["category"] for item in terms for _ in range(item["count"]))
    nodes = [
        {"id": "center", "label": story["title"], "type": "单篇译文", "count": 1},
        {"id": "source", "label": story.get("source") or "未记录来源", "type": "故事来源", "count": 1},
        {"id": "class", "label": story.get("category") or "未分类", "type": "卫礼贤分类", "count": 1},
    ]
    edges = [
        {"source": "center", "target": "source", "weight": 1, "relation": "来源"},
        {"source": "center", "target": "class", "weight": 1, "relation": "分类"},
    ]
    for category, count in category_counts.most_common():
        nodes.append({"id": f"cat-{category}", "label": category, "type": "主题类别", "count": count})
        edges.append({"source": "center", "target": f"cat-{category}", "weight": count, "relation": "主题类别"})
    for item in terms:
        nodes.append({"id": f"term-{item['term']}", "label": item["term"], "type": "关键词", "count": item["count"]})
        edges.append({"source": f"cat-{item['category']}", "target": f"term-{item['term']}", "weight": item["count"], "relation": "关键词"})
    for source, target in combinations(sorted({item["term"] for item in terms}), 2):
        edges.append({"source": f"term-{source}", "target": f"term-{target}", "weight": 1, "relation": "共现"})
    return {
        "nodes": nodes,
        "edges": edges,
        "terms": [
            {
                "term": item["term"],
                "category": item["category"],
                "count": item["count"],
                "stories": [{"storyId": story["id"], "storyTitle": story["title"], "category": item["category"], "count": item["count"]}],
            }
            for item in sorted(terms, key=lambda row: row["count"], reverse=True)
        ],
    }


def wilhelm_story_graphs(stories: list[dict]) -> dict[str, dict]:
    return {story["id"]: wilhelm_single_story_graph(story) for story in stories}


def tokenize_meaningful(text: str) -> list[str]:
    tokens = []
    for token in re.findall(r"[\u4e00-\u9fff]{2,}|[A-Za-zÄÖÜäöüß]{3,}", text):
        value = token.strip()
        if value.lower() not in STOP_WORDS and value not in STOP_WORDS:
            tokens.append(value)
    return tokens


def build_stats(collection_items: list[dict], children: list[dict], flows: list[dict]) -> dict:
    return {
        "collectionCount": len(collection_items),
        "childCount": len(children),
        "matchedChildCount": len({child_id for item in collection_items for child_id in item["matchedChildIds"]}),
        "languages": collections.Counter(child["language"] for child in children if child["language"]).most_common(16),
        "countries": collections.Counter(item.get("country", "") for item in collection_items if item.get("country")).most_common(16),
        "translators": collections.Counter(item["editor"] for item in collection_items if item["editor"]).most_common(16),
        "storyNames": collections.Counter(child["canonicalName"] for child in children if child["canonicalName"]).most_common(24),
        "carriers": collections.Counter(child["carrier"] for child in children if child["carrier"]).most_common(10),
        "translationModes": collections.Counter(child["translationMode"] or "未记录" for child in children).most_common(10),
        "editorRoles": collections.Counter(item["editorRole"] for item in collection_items if item["editorRole"]).most_common(12),
        "publishers": collections.Counter(item["publisher"] for item in collection_items if item["publisher"]).most_common(12),
        "provinceCounts": collections.Counter(flow["province"] for flow in flows).most_common(),
        "yearSeries": sorted(collections.Counter(f"{item['year'] // 10 * 10}s" for item in collection_items if item["year"]).items()),
    }


WILHELM_TERM_LEXICON_V2 = {
    "动物": ["龙", "虎", "蛇", "狐", "狐狸", "牛", "马", "鸟", "鱼", "龟", "猴", "犬", "狗", "猫", "兔", "鼠", "鹿", "鹤", "鹰", "虫"],
    "气象": ["天", "云", "风", "雨", "雷", "电", "月", "太阳", "星", "雪", "雾", "霜"],
    "神怪": ["神", "仙", "鬼", "怪", "妖", "魔", "龙王", "女神", "仙女", "观音", "玉帝"],
    "人物身份": ["皇帝", "国王", "公主", "王子", "农夫", "和尚", "道士", "猎人", "妻子", "母亲", "父亲", "兄弟", "老人", "读者"],
    "空间": ["山", "海", "河", "井", "桥", "村", "宫", "天宫", "田", "森林", "庙"],
    "伦理": ["孝", "报恩", "惩罚", "复仇", "智慧", "善良", "忠诚", "贪婪", "仁慈"],
}


def wilhelm_story_terms(story: dict) -> list[dict]:
    text = f"{story.get('title', '')} {story.get('source', '')} {story.get('category', '')} {story.get('text', '')}"
    terms = []
    seen = set()
    for category, words in WILHELM_TERM_LEXICON_V2.items():
        for word in words:
            count = len(re.findall(re.escape(word), text, flags=re.I))
            if count and (category, word) not in seen:
                seen.add((category, word))
                terms.append({"term": word, "category": category, "count": count})
    existing = {item["term"] for item in terms}
    for token, count in collections.Counter(tokenize_meaningful(text)).most_common(18):
        if token not in existing and len(token) >= 2:
            terms.append({"term": token, "category": "关键词", "count": count})
            existing.add(token)
    return sorted(terms, key=lambda item: (item["count"], len(item["term"])), reverse=True)[:42]


def _wilhelm_graph_from_stories(stories: list[dict], graph_id: str = "total", title: str = "卫礼贤《中国民间童话》") -> dict:
    term_counts = collections.Counter()
    category_counts = collections.Counter()
    story_links: dict[str, list[dict]] = collections.defaultdict(list)
    pair_counts = collections.Counter()
    triples = []
    story_nodes = []
    relation_by_category = {
        "动物": "出现于",
        "气象": "关联气象",
        "神怪": "关联神怪",
        "人物身份": "涉及身份",
        "空间": "发生空间",
        "伦理": "表达母题",
    }
    for story in stories:
        story_id = story.get("id") or f"story-{len(story_nodes) + 1}"
        story_title = story.get("title") or "未命名故事"
        terms = wilhelm_story_terms(story)
        story_nodes.append({"id": f"story-{story_id}", "label": story_title, "type": "故事", "count": max(1, len(terms)), "storyId": story_id})
        term_names = [item["term"] for item in terms[:24]]
        for item in terms:
            term_counts[item["term"]] += item["count"]
            category_counts[item["category"]] += item["count"]
            story_links[item["term"]].append({"storyId": story_id, "storyTitle": story_title, "category": item["category"], "count": item["count"]})
            triples.append({"subject": story_title, "predicate": relation_by_category.get(item["category"], "提取关键词"), "object": item["term"], "weight": item["count"], "storyId": story_id})
        for source, target in combinations(sorted(set(term_names)), 2):
            pair_counts[(source, target)] += 1
    top_terms = [term for term, _ in term_counts.most_common(48)]
    nodes = [{"id": "center", "label": title, "type": "故事", "count": len(stories)}]
    if len(stories) <= 16:
        nodes.extend(story_nodes)
    nodes.extend({"id": f"cat-{category}", "label": category, "type": "分类", "count": count} for category, count in category_counts.most_common(8))
    nodes.extend(
        {
            "id": f"term-{term}",
            "label": term,
            "type": next((link["category"] for link in story_links[term]), "关键词"),
            "count": term_counts[term],
            "stories": story_links[term][:12],
        }
        for term in top_terms
    )
    node_ids = {node["id"] for node in nodes}
    edges = []
    if len(stories) <= 16:
        edges.extend({"source": "center", "target": node["id"], "weight": node["count"], "relation": "包含"} for node in story_nodes)
    for category, count in category_counts.most_common(8):
        edges.append({"source": "center", "target": f"cat-{category}", "weight": count, "relation": "主题分类"})
    for term in top_terms:
        category = next((link["category"] for link in story_links[term]), "关键词")
        category_id = f"cat-{category}"
        if category_id in node_ids:
            edges.append({"source": category_id, "target": f"term-{term}", "weight": term_counts[term], "relation": "抽取关键词"})
        if len(stories) <= 16:
            for link in story_links[term][:4]:
                story_id = f"story-{link['storyId']}"
                if story_id in node_ids:
                    edges.append({"source": story_id, "target": f"term-{term}", "weight": link["count"], "relation": "出现于"})
    cooccurrence = []
    for (source, target), count in pair_counts.most_common(120):
        if source in top_terms and target in top_terms:
            cooccurrence.append({"source": source, "target": target, "weight": count})
            edges.append({"source": f"term-{source}", "target": f"term-{target}", "weight": count, "relation": "共现"})
    return {
        "id": graph_id,
        "nodes": nodes,
        "edges": [edge for edge in edges if edge["source"] in node_ids and edge["target"] in node_ids],
        "triples": triples[:260],
        "cooccurrence": cooccurrence,
        "terms": [
            {
                "term": term,
                "category": next((link["category"] for link in story_links[term]), "关键词"),
                "count": count,
                "stories": sorted(story_links[term], key=lambda item: item["count"], reverse=True)[:12],
            }
            for term, count in term_counts.most_common(60)
        ],
    }


def wilhelm_theme_graph(stories: list[dict]) -> dict:
    return _wilhelm_graph_from_stories(stories, "total", "卫礼贤《中国民间童话》")


def wilhelm_single_story_graph(story: dict) -> dict:
    return _wilhelm_graph_from_stories([story], story.get("id") or "single", story.get("title") or "单篇译文")


def main() -> None:
    prefaces = read_prefaces()
    children = read_children()
    collections_data = read_collections(children, prefaces)
    collection_by_title = {norm(item["name"]): item["id"] for item in collections_data}
    for preface in prefaces:
        preface["collectionId"] = collection_by_title.get(norm(preface["collectionTitle"]), "")
    flows = build_flows(collections_data)
    wilhelm_stories = read_wilhelm_stories()
    payload = {
        "collections": collections_data,
        "childStories": children,
        "prefaces": prefaces,
        "wilhelmStories": wilhelm_stories,
        "wilhelmEditions": read_wilhelm_editions(),
        "wilhelmThemeGraph": wilhelm_theme_graph(wilhelm_stories),
        "wilhelmStoryGraphs": wilhelm_story_graphs(wilhelm_stories),
        "flows": flows,
        "stats": build_stats(collections_data, children, flows),
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(
        f"wrote {OUT} collections={len(collections_data)} children={len(children)} "
        f"prefaces={len(prefaces)} wilhelmStories={len(wilhelm_stories)}"
    )


if __name__ == "__main__":
    main()
