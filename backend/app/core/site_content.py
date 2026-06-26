from __future__ import annotations

import json
import os
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

try:
    from backend.app.core import platform_store
except Exception:  # pragma: no cover - site content can run without operation logs.
    platform_store = None  # type: ignore


ROOT = Path(__file__).resolve().parents[3]
APP_DATA_ROOT = Path(os.environ.get("PLATFORM_DATA_ROOT") or os.environ.get("DATA_ROOT") or (ROOT / "data")).resolve()
STORE_PATH = APP_DATA_ROOT / "site_content.json"
ABOUT_ASSET_ROOT = ROOT / "frontend" / "public" / "assets" / "about"
ABOUT_IMAGE_ROOT = ABOUT_ASSET_ROOT / "image"

CONTENT_KINDS = ("team", "committee", "publications", "activities", "dynamics")

SLUG_BY_NAME = {
    "张帆": "zhang-fan",
    "胡文婷": "hu-wenting",
    "童欣": "tong-xin",
    "尹兰曦": "yin-lanxi",
    "朱伟芳": "zhu-weifang",
    "张天资": "zhang-tianzi",
    "张永维": "zhang-yongwei",
    "崔钰": "cui-yu",
    "余晴": "yu-qing",
    "葛桂录": "ge-guilu",
    "高方": "gao-fang",
    "谭渊": "tan-yuan",
    "刘志强": "liu-zhiqiang",
    "陈琦": "chen-qi",
    "周琼": "zhou-qiong",
    "唐珂": "tang-ke",
    "刘启君": "liu-qijun",
}

CATEGORY_BY_NAME = {
    "张帆": "本院学者",
    "胡文婷": "本院学者",
    "童欣": "本院学者",
    "尹兰曦": "本院学者",
    "朱伟芳": "本院学者",
    "张天资": "本院学者",
    "张永维": "本院学者",
    "崔钰": "本院学者",
    "余晴": "本院学者",
    "葛桂录": "特聘专家",
    "高方": "特聘专家",
    "谭渊": "特聘专家",
    "刘志强": "特聘专家",
    "陈琦": "双聘研究员",
    "周琼": "兼职研究员",
    "唐珂": "兼职研究员",
    "刘启君": "兼职研究员",
}

FALLBACK_TEAM = [
    {
        "name": "张帆",
        "slug": "zhang-fan",
        "category": "本院学者",
        "role": "二级教授 / 博士生导师 / 中心主任",
        "organization": "上海外国语大学",
        "focus": ["德语文学", "比较文学", "世界文学"],
        "intro": "长期从事中国话语、世界文学与中国文学海外传播研究。",
        "image": "张帆.jpg",
    },
    {
        "name": "胡文婷",
        "slug": "hu-wenting",
        "category": "本院学者",
        "role": "副教授",
        "organization": "上海外国语大学",
        "focus": ["海外汉学", "译介研究", "比较文学"],
        "intro": "关注中国文学跨语种译介、海外汉学与比较文学研究。",
        "image": "胡文婷.jpg",
    },
    {
        "name": "童欣",
        "slug": "tong-xin",
        "category": "本院学者",
        "role": "助理研究员",
        "organization": "上海外国语大学",
        "focus": ["记忆研究", "中国故事传播", "数字人文"],
        "intro": "围绕文本记忆、跨文化传播与数字人文方法开展研究。",
        "image": "童欣.jpg",
    },
    {
        "name": "尹兰曦",
        "slug": "yin-lanxi",
        "category": "本院学者",
        "role": "博士后",
        "organization": "上海外国语大学",
        "focus": ["莎士比亚戏剧", "修辞学", "批评话语"],
        "intro": "研究莎士比亚戏剧、修辞观嬗变与中国莎士比亚批评话语。",
        "image": "尹兰曦.jpg",
    },
    {
        "name": "朱伟芳",
        "slug": "zhu-weifang",
        "category": "特聘专家",
        "role": "特聘专家",
        "organization": "中国话语与世界文学研究中心",
        "focus": ["世界文学", "跨文化阐释"],
        "intro": "参与中心世界文学与跨文化阐释方向研究。",
        "image": "朱伟芳.jpg",
    },
    {
        "name": "张天资",
        "slug": "zhang-tianzi",
        "category": "特聘专家",
        "role": "特聘专家",
        "organization": "中国话语与世界文学研究中心",
        "focus": ["比较文学", "译介传播"],
        "intro": "关注比较文学视野下的文本传播与话语转化。",
        "image": "张天资.jpg",
    },
    {
        "name": "张永维",
        "slug": "zhang-yongwei",
        "category": "双聘研究员",
        "role": "双聘研究员",
        "organization": "中国话语与世界文学研究中心",
        "focus": ["外国文学", "区域国别研究"],
        "intro": "参与外国文学与区域国别研究方向建设。",
        "image": "张永维.jpg",
    },
    {
        "name": "崔钰",
        "slug": "cui-yu",
        "category": "兼职研究员",
        "role": "兼职研究员",
        "organization": "中国话语与世界文学研究中心",
        "focus": ["文学传播", "文化记忆"],
        "intro": "研究文学传播、文化记忆与中国叙事的跨文化流动。",
        "image": "崔钰.png",
    },
    {
        "name": "余晴",
        "slug": "yu-qing",
        "category": "兼职研究员",
        "role": "兼职研究员",
        "organization": "中国话语与世界文学研究中心",
        "focus": ["世界文学", "中国叙事"],
        "intro": "关注世界文学中的中国叙事与跨语境阐释。",
        "image": "余晴.jpg",
    },
]

DEFAULT_COMMITTEE = [
    {"name": "许钧", "org": "浙江大学", "role": "主任委员"},
    {"name": "查明建", "org": "上海外国语大学", "role": "委员"},
    {"name": "金莉", "org": "北京外国语大学", "role": "委员"},
    {"name": "聂珍钊", "org": "广东外语外贸大学", "role": "委员"},
    {"name": "王克非", "org": "北京外国语大学", "role": "委员"},
    {"name": "杨平", "org": "专家委员", "role": "委员"},
    {"name": "杨金才", "org": "南京大学", "role": "委员"},
    {"name": "彭青龙", "org": "上海交通大学", "role": "委员"},
    {"name": "宋炳辉", "org": "上海外国语大学", "role": "委员"},
    {"name": "梁展", "org": "中国社会科学院", "role": "委员"},
    {"name": "苏晖", "org": "上海外国语大学", "role": "委员"},
    {"name": "张帆", "org": "上海外国语大学", "role": "委员", "image": "张帆.jpg"},
]

DEFAULT_PUBLICATIONS = [
    {
        "title": "中国文学海外译介研究丛书",
        "meta": "上海大学出版社 / 2024",
        "type": "研究丛书",
        "date": "2024",
        "image": "“中国文学海外译介研究丛书”（丛书主编：张帆、孙国亮）由上海大学出版社出版，2024年首推2部。.png",
    },
    {
        "title": "博士后专著和译著",
        "meta": "专著译著 / 3 部",
        "type": "专著译著",
        "date": "2024",
        "image": "研究中心博士后出版专著和译著共计3部。.jpg",
    },
    {
        "title": "中国故事的世界传播与流变",
        "meta": "《国际汉学》特约专栏",
        "type": "学术专栏",
        "date": "2024",
        "image": "研究中心依托教育部重大课题，在CSSCI来源期刊《国际汉学》开设特约专栏“中国故事的世界传播与流变”。.png",
    },
    {
        "title": "中国故事的世界传播",
        "meta": "《中国故事》固定专栏",
        "type": "学术专栏",
        "date": "2024",
        "image": "研究中心在综合性人文期刊《中国故事》开设固定专栏“中国故事的世界传播”。.png",
    },
]

DEFAULT_ACTIVITIES = [
    {
        "title": "主办国内学术会议",
        "type": "学术会议",
        "date": "2024",
        "image": "2024年，研究中心主办3次国内学术会议，参会总人数近150人次。.png",
    },
    {
        "title": "国别区域全球知识前沿讲坛",
        "type": "前沿讲坛",
        "date": "2024",
        "image": "举办“国别区域全球知识前沿讲坛·中国话语与世界文学研究中心系列讲座”。.png",
    },
    {
        "title": "文学经典的跨文化旅行读书会",
        "type": "读书会",
        "date": "2024",
        "image": "开展“思想·诠释·对话：文学经典的跨文化旅行”主题系列读书会。1.png",
    },
    {
        "title": "国内外学术会议论文宣读",
        "type": "国际交流",
        "date": "2024",
        "image": "研究中心成员参加国内外学术会议并宣读论文24篇。.png",
    },
    {
        "title": "成员受邀学术讲座",
        "type": "前沿讲坛",
        "date": "2024",
        "image": "研究中心成员受邀在国内高校及文化机构进行学术讲座11次。.jpg",
    },
]

PERSON_NAMES = {
    "张帆",
    "胡文婷",
    "童欣",
    "尹兰曦",
    "朱伟芳",
    "张天资",
    "张永维",
    "崔钰",
    "余晴",
    "葛桂录",
    "高方",
    "谭渊",
    "刘志强",
    "陈琦",
    "周琼",
    "唐珂",
    "徐林峰",
    "何心怡",
    "刘启君",
    "张丽",
    "唐洁",
    "陈悦",
    "高鸽",
    "陈雨田",
    "徐冠群",
    "段亚男",
    "陈丽竹",
}


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat()


def _ensure_dirs() -> None:
    APP_DATA_ROOT.mkdir(parents=True, exist_ok=True)


def _image_files() -> list[str]:
    if not ABOUT_IMAGE_ROOT.exists():
        return []
    return sorted(
        path.name
        for path in ABOUT_IMAGE_ROOT.iterdir()
        if path.is_file() and path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp"}
    )


def _image_title(filename: str) -> str:
    return re.sub(r"[，,。.\s]+$", "", re.sub(r"\.(png|jpe?g|webp)$", "", filename, flags=re.I))


def _find_image_for_name(name: str) -> str:
    for filename in _image_files():
        if _image_title(filename).startswith(name):
            return filename
    return f"{name}.jpg"


def _clean_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.strip()


def _split_numbered(value: Any) -> list[str]:
    text = _clean_text(value)
    if not text:
        return []
    text = re.sub(r"(\d+[）).、])", r"\n\1", text)
    return [
        re.sub(r"^\d+[）).、]\s*", "", item).strip()
        for item in re.split(r"\n+", text)
        if item.strip()
    ]


def _split_focus(value: Any) -> list[str]:
    if isinstance(value, list):
        return [_clean_text(item) for item in value if _clean_text(item)]
    items = _split_numbered(value)
    if items:
        return items
    return [item.strip() for item in re.split(r"[,，、/|；;]+", _clean_text(value)) if item.strip()]


def _slugify(value: str, fallback: str) -> str:
    if value in SLUG_BY_NAME:
        return SLUG_BY_NAME[value]
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^A-Za-z0-9]+", "-", normalized).strip("-").lower()
    return slug or fallback


def _scholar_workbook_path() -> Path | None:
    assets_root = ROOT / "frontend" / "public" / "assets"
    if not assets_root.exists():
        return None
    for path in assets_root.glob("*.xlsx"):
        if "学者" in path.name:
            return path
    return None


def _parse_team_from_workbook() -> list[dict[str, Any]]:
    workbook_path = _scholar_workbook_path()
    if not workbook_path:
        return []
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        return []
    if not rows:
        return []
    headers = [_clean_text(cell) for cell in rows[0]]
    team: list[dict[str, Any]] = []
    for index, row in enumerate(rows[1:], start=1):
        values = {headers[cell_index]: row[cell_index] if cell_index < len(row) else "" for cell_index in range(len(headers))}
        name = _clean_text(values.get("姓名"))
        if not name:
            continue
        role = _clean_text(values.get("职称/身份")) or "研究人员"
        focus = _split_focus(values.get("研究方向"))
        basic = _split_numbered(values.get("基本信息（教育背景等）"))
        projects = _split_numbered(values.get("课题项目"))
        books = _split_numbered(values.get("著作出版"))
        papers = _split_numbered(values.get("论文发表"))
        honors = _split_numbered(values.get("荣誉奖项"))
        contact_text = _clean_text(values.get("联系方式"))
        contact = []
        if contact_text:
            contact = [{"label": "电子邮件" if "@" in contact_text else "联系方式", "value": contact_text, "href": f"mailto:{contact_text}" if "@" in contact_text else ""}]
        organization = next((item for item in basic if re.search(r"大学|学院|研究中心|研究院", item)), "中国话语与世界文学研究中心")
        team.append(
            {
                "name": name,
                "slug": _slugify(name, f"team-{index}"),
                "category": CATEGORY_BY_NAME.get(name, "兼职研究员" if "兼职" in role else "本院学者"),
                "role": role,
                "organization": organization,
                "focus": focus,
                "intro": basic[0] if basic else (f"研究方向：{'、'.join(focus)}" if focus else "相关简介待补充。"),
                "image": _find_image_for_name(name),
                "detail": {
                    "basic": basic,
                    "direction": focus,
                    "profile": [],
                    "achievements": {"projects": projects, "books": books, "papers": papers, "honors": honors},
                    "contact": contact,
                },
            }
        )
    return team


def _dynamic_type(title: str) -> str:
    if re.search(r"会议|讲座|读书会|驻访", title):
        return "学术活动"
    if re.search(r"丛书|专著|译著|出版|专栏", title):
        return "成果转化"
    if re.search(r"课程|人才培养", title):
        return "人才培养"
    if re.search(r"论文|课题|项目", title):
        return "科研进展"
    if re.search(r"话语传播|新媒体|官方网站", title):
        return "话语传播"
    return "平台动态"


def _dynamic_topic(title: str) -> str:
    if re.search(r"会议|讲座|读书会|课程|人才培养", title):
        return "focus"
    if re.search(r"丛书|专著|译著|论文|课题|项目|专栏", title):
        return "research"
    if re.search(r"话语传播|新媒体|官方网站|社会影响", title):
        return "media"
    return "focus"


def _dynamic_rank(filename: str) -> int:
    title = _image_title(filename)
    priorities = [
        "2024年，研究中心主办",
        "举办“国别区域",
        "开展“思想",
        "中国文学海外译介研究丛书",
        "研究中心博士后出版",
        "研究中心依托教育部",
        "研究中心在综合性",
        "研究中心成员受邀",
        "研究中心成员参加",
        "2024年，研究中心成员开设",
        "话语传播",
        "人才培养",
        "项目",
        "论文",
        "专著",
    ]
    for index, keyword in enumerate(priorities):
        if keyword in title:
            return index
    return 99


def _is_person_image(filename: str) -> bool:
    title = _image_title(filename)
    return any(name in title for name in PERSON_NAMES) and not re.search(r"[。，“”《》：:]", title)


def _default_dynamics() -> list[dict[str, Any]]:
    items = []
    for filename in sorted([name for name in _image_files() if not _is_person_image(name)], key=_dynamic_rank):
        title = _image_title(filename)
        items.append(
            {
                "title": title,
                "type": _dynamic_type(title),
                "topic": _dynamic_topic(title),
                "date": "2024",
                "image": filename,
                "summary": title,
                "content": title,
            }
        )
    return items


def _empty_content() -> dict[str, list[dict[str, Any]]]:
    return {kind: [] for kind in CONTENT_KINDS}


def _build_default_store() -> dict[str, Any]:
    store = {"version": 1, "initialized": True, "next_id": 1, "updated_at": _now_iso(), "content": _empty_content()}

    def add(kind: str, payload: dict[str, Any], index: int) -> None:
        item = _normalize_item(kind, payload, store["next_id"])
        item["order"] = payload.get("order", index)
        item.setdefault("visible", True)
        store["next_id"] += 1
        store["content"][kind].append(item)

    for index, item in enumerate(_parse_team_from_workbook() or FALLBACK_TEAM, start=1):
        add("team", item, index)
    for index, item in enumerate(DEFAULT_COMMITTEE, start=1):
        add("committee", item, index)
    for index, item in enumerate(DEFAULT_PUBLICATIONS, start=1):
        add("publications", item, index)
    for index, item in enumerate(DEFAULT_ACTIVITIES, start=1):
        add("activities", item, index)
    for index, item in enumerate(_default_dynamics(), start=1):
        add("dynamics", item, index)
    return store


def _load_store() -> dict[str, Any]:
    _ensure_dirs()
    if not STORE_PATH.exists():
        store = _build_default_store()
        _save_store(store)
        return store
    try:
        data = json.loads(STORE_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = _build_default_store()
    data.setdefault("version", 1)
    data.setdefault("initialized", True)
    data.setdefault("next_id", 1)
    data.setdefault("updated_at", _now_iso())
    content = data.setdefault("content", {})
    for kind in CONTENT_KINDS:
        content.setdefault(kind, [])
    return data


def _save_store(store: dict[str, Any]) -> None:
    _ensure_dirs()
    store["updated_at"] = _now_iso()
    tmp = STORE_PATH.with_suffix(".tmp")
    tmp.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(STORE_PATH)


def _asset_url(filename: str) -> str:
    if not filename:
        return ""
    if re.match(r"^(https?:|data:|/)", filename):
        return filename
    return f"/assets/about/image/{quote(filename)}"


def _normalize_item(kind: str, payload: dict[str, Any], fallback_id: int | str | None = None) -> dict[str, Any]:
    if kind not in CONTENT_KINDS:
        raise ValueError("Unsupported site content kind.")
    item = {key: value for key, value in (payload or {}).items() if key not in {"created_at", "updated_at"}}
    item["id"] = _clean_text(item.get("id")) or f"{kind}-{fallback_id or int(time.time() * 1000)}"
    item["visible"] = bool(item.get("visible", True))
    try:
        item["order"] = int(item.get("order") or 999)
    except Exception:
        item["order"] = 999
    if "focus" in item:
        item["focus"] = _split_focus(item.get("focus"))
    if kind == "team":
        name = _clean_text(item.get("name"))
        if not name:
            raise ValueError("Name is required.")
        item["name"] = name
        item["slug"] = _clean_text(item.get("slug")) or _slugify(name, item["id"])
        item.setdefault("category", "本院学者")
        item.setdefault("role", "研究人员")
        item.setdefault("organization", "中国话语与世界文学研究中心")
        item.setdefault("focus", [])
        item.setdefault("intro", "")
    elif kind == "committee":
        name = _clean_text(item.get("name"))
        if not name:
            raise ValueError("Name is required.")
        item["name"] = name
        item.setdefault("role", "委员")
        item.setdefault("org", item.get("organization") or "")
    else:
        title = _clean_text(item.get("title"))
        if not title:
            raise ValueError("Title is required.")
        item["title"] = title
        item.setdefault("summary", item.get("content") or title)
        if kind == "dynamics":
            item.setdefault("topic", _dynamic_topic(title))
            item.setdefault("type", _dynamic_type(title))
        if kind == "publications":
            item.setdefault("type", "学术成果")
            item.setdefault("meta", item.get("date") or "")
        if kind == "activities":
            item.setdefault("type", "学术活动")
    return item


def _decorate_item(item: dict[str, Any]) -> dict[str, Any]:
    decorated = json.loads(json.dumps(item, ensure_ascii=False))
    image = _clean_text(decorated.get("image") or decorated.get("imageUrl") or decorated.get("image_url"))
    if image:
        decorated["image"] = image
        decorated["image_url"] = _asset_url(image)
    return decorated


def _sorted_items(items: list[dict[str, Any]], public_only: bool) -> list[dict[str, Any]]:
    source = [item for item in items if not public_only or item.get("visible", True)]
    return [_decorate_item(item) for item in sorted(source, key=lambda value: (int(value.get("order") or 999), str(value.get("name") or value.get("title") or "")))]


def public_content() -> dict[str, Any]:
    store = _load_store()
    return {
        "content": {kind: _sorted_items(store.get("content", {}).get(kind, []), True) for kind in CONTENT_KINDS},
        "updated_at": store.get("updated_at"),
    }


def admin_content() -> dict[str, Any]:
    store = _load_store()
    return {
        "content": {kind: _sorted_items(store.get("content", {}).get(kind, []), False) for kind in CONTENT_KINDS},
        "assets": {"images": _image_files()},
        "updated_at": store.get("updated_at"),
    }


def create_item(kind: str, payload: dict[str, Any], user_id: int | None = None) -> dict[str, Any]:
    store = _load_store()
    item = _normalize_item(kind, payload, store.get("next_id", 1))
    item["created_at"] = _now_iso()
    item["updated_at"] = item["created_at"]
    store["next_id"] = int(store.get("next_id", 1)) + 1
    store.setdefault("content", {}).setdefault(kind, []).append(item)
    _save_store(store)
    _record_operation("site_content_create", f"Created {kind}:{item.get('name') or item.get('title')}", user_id)
    return _decorate_item(item)


def update_item(kind: str, item_id: str, payload: dict[str, Any], user_id: int | None = None) -> dict[str, Any]:
    store = _load_store()
    items = store.setdefault("content", {}).setdefault(kind, [])
    for index, current in enumerate(items):
        if str(current.get("id")) != str(item_id):
            continue
        merged = {**current, **(payload or {}), "id": current.get("id")}
        item = _normalize_item(kind, merged, current.get("id"))
        item["created_at"] = current.get("created_at") or _now_iso()
        item["updated_at"] = _now_iso()
        items[index] = item
        _save_store(store)
        _record_operation("site_content_update", f"Updated {kind}:{item_id}", user_id)
        return _decorate_item(item)
    raise KeyError("Site content item not found.")


def delete_item(kind: str, item_id: str, user_id: int | None = None) -> None:
    store = _load_store()
    items = store.setdefault("content", {}).setdefault(kind, [])
    next_items = [item for item in items if str(item.get("id")) != str(item_id)]
    if len(next_items) == len(items):
        raise KeyError("Site content item not found.")
    store["content"][kind] = next_items
    _save_store(store)
    _record_operation("site_content_delete", f"Deleted {kind}:{item_id}", user_id)


def _record_operation(operation_type: str, content: str, user_id: int | None = None) -> None:
    if not platform_store:
        return
    try:
        platform_store.record_operation(operation_type, content, user_id=user_id)
    except Exception:
        return
