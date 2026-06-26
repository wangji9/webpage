from __future__ import annotations

import base64
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
import threading
from collections import Counter
from pathlib import Path
from typing import Any
from zipfile import ZipFile
from xml.etree import ElementTree as ET


ROOT = Path(__file__).resolve().parents[3]
DATA_DIR = ROOT / "data"
PUBLIC_ASSETS = ROOT / "frontend" / "public" / "assets"
BUILD_STORY_SCRIPT = ROOT / "scripts" / "build_story_data.py"
BUILD_PUBLICATION_MAP_SCRIPT = ROOT / "scripts" / "generate_wilhelm_publication_source_map.py"
MANIFEST_PATH = Path(os.environ.get("DATASET_MANIFEST_PATH") or (ROOT / "backend" / "app" / "dataset_uploads.json")).resolve()
DATASET_LOCK = threading.Lock()

MODULES = [
    {
        "id": "classics",
        "name": "中国典籍海外译介",
        "keywords": ["典籍", "译介", "汉学", "出版", "接受史"],
        "description": "维护中国典籍外译、汉学研究、出版传播与海外接受史相关数据。",
    },
    {
        "id": "shanghai",
        "name": "上海文学海外传播",
        "keywords": ["上海文学", "海外传播", "作家", "译本", "城市文化"],
        "description": "维护上海文学作品、译本、作家、城市文化与海外传播路径数据。",
    },
    {
        "id": "stories",
        "name": "多语种中国故事集",
        "keywords": ["真实数据", "故事集", "子故事", "序跋", "传播路径"],
        "description": "维护故事集总表、子故事、序跋、来源地、出版地与专题地图数据。",
    },
    {
        "id": "world",
        "name": "世界文学的中国叙事",
        "keywords": ["世界文学", "中国叙事", "改编", "翻译本", "跨文化"],
        "description": "维护世界文学中中国形象、改编、翻译本与跨文化叙事数据。",
    },
]

MODULE_NAME_BY_ID = {item["id"]: item["name"] for item in MODULES}
MODULE_ID_BY_NAME = {item["name"]: item["id"] for item in MODULES}


DATASETS = [
    {
        "id": "story-collections",
        "title": "中国故事集总表_知识库.xlsx",
        "filename": "中国故事集总表_知识库.xlsx",
        "aliases": ["中国故事集总表_知识库(1).xlsx"],
        "module": "多语种中国故事集",
        "moduleId": "stories",
        "pages": ["知识库", "知识图谱", "智能问答", "统计图表", "出版地地图"],
        "content": "故事集标题、中文标题、年份、译者/编者、身份、序跋作者、国家、城市、出版社、子故事数量。",
        "requiredColumns": ["故事集标题", "故事集标题（中文）", "年份", "译者/编者", "国家", "城市", "出版社"],
        "rebuild": "story",
    },
    {
        "id": "story-children",
        "title": "中国故事集_子故事（3533篇）.xlsx",
        "filename": "中国故事集_子故事（3533篇）.xlsx",
        "module": "多语种中国故事集",
        "moduleId": "stories",
        "pages": ["知识库嵌套子故事表", "子故事知识图谱", "智能问答召回", "统计图表"],
        "content": "年份、故事集标题、子故事标题、规范故事名、故事民族来源。",
        "requiredColumns": ["年份", "故事集标题", "子故事标题", "规范故事名"],
        "rebuild": "story",
    },
    {
        "id": "story-prefaces",
        "title": "中国故事集_序跋.xlsx",
        "filename": "中国故事集_序跋.xlsx",
        "module": "多语种中国故事集",
        "moduleId": "stories",
        "pages": ["序跋辑录", "序跋主题聚类", "词云", "智能问答"],
        "content": "故事集标题、年份、故事集介绍、序跋作者、序跋类型、序跋文本。",
        "requiredColumns": ["故事集标题", "年份", "故事集介绍", "序跋作者", "序跋类型", "序跋文本"],
        "rebuild": "story",
    },
    {
        "id": "story-source-map",
        "title": "地图_中国故事集_故事来源地.xlsx",
        "filename": "地图_中国故事集_故事来源地.xlsx",
        "module": "多语种中国故事集",
        "moduleId": "stories",
        "pages": ["中国来源省份地图", "传播路径图", "来源地统计"],
        "content": "故事集 title 与 source region，用于匹配来源省份并生成中国地图。",
        "requiredColumns": ["title", "source region"],
        "rebuild": "story",
    },
    {
        "id": "story-publication-source-map",
        "title": "地图_中国故事集_出版地和故事来源地.xlsx",
        "filename": "地图_中国故事集_出版地和故事来源地.xlsx",
        "module": "多语种中国故事集",
        "moduleId": "stories",
        "pages": ["出版地气泡地图", "来源地地图", "传播地图", "可视化图谱"],
        "content": "title、中文标题、年份、国家、城市、出版社、来源区域。",
        "requiredColumns": ["title", "year", "country", "city", "publisher"],
        "rebuild": "story-and-publication-map",
    },
    {
        "id": "wilhelm-text",
        "title": "中国民间童话.xlsx",
        "filename": "中国民间童话.xlsx",
        "module": "多语种中国故事集",
        "moduleId": "stories",
        "group": "卫礼贤专题",
        "pages": ["卫礼贤专题库", "单篇译文知识图谱", "关键词网络", "智能问答"],
        "content": "单篇译文故事名、译文内容、故事来源、卫礼贤分类。",
        "requiredColumns": ["单篇译文故事名", "译文内容"],
        "rebuild": "story",
        "publicAsset": True,
    },
    {
        "id": "wilhelm-map",
        "title": "地图_中国民间童话.xlsx",
        "filename": "地图_中国民间童话.xlsx",
        "module": "多语种中国故事集",
        "moduleId": "stories",
        "group": "卫礼贤专题",
        "pages": ["卫礼贤再版出版地地图", "统计图表", "智能问答"],
        "content": "titel、country、city、publisher、year、全/选/改编、语种。",
        "requiredColumns": ["titel", "country", "city", "publisher", "year"],
        "rebuild": "story",
    },
    {
        "id": "database-info",
        "title": "数据库信息.xlsx",
        "filename": "数据库信息.xlsx",
        "module": "世界文学的中国叙事",
        "moduleId": "world",
        "pages": ["后续扩展知识库", "智能问答补充语料", "外文改编/翻译本表"],
        "content": "外文改编本、外文翻译本等工作表，包含规范故事名、变异故事名、出版时间、作者、国籍、语种等。",
        "requiredColumns": ["规范故事名", "变异故事名", "出版时间", "语种"],
        "rebuild": "cache-only",
    },
    {
        "id": "map-story-kb",
        "title": "地图_中国故事集总表_知识库.xlsx",
        "filename": "地图_中国故事集总表_知识库.xlsx",
        "module": "多语种中国故事集",
        "moduleId": "stories",
        "pages": ["地图数据校验", "出版地/来源地补充"],
        "content": "地图端故事集总表字段，辅助校验出版地与来源地。",
        "requiredColumns": ["title", "year", "country", "city", "publisher"],
        "rebuild": "cache-only",
    },
]

CUSTOM_REBUILD = "cache-only"
CUSTOM_KNOWLEDGE_CACHE: dict[str, tuple[tuple[Any, ...], list[dict[str, Any]]]] = {}
MODULE_DATASET_CACHE: dict[str, tuple[tuple[Any, ...], dict[str, Any]]] = {}
MODULE_DATASET_SUMMARY_CACHE: dict[str, tuple[tuple[Any, ...], dict[str, Any]]] = {}
CHINA_COORDS = [116.4, 39.9]
DEFAULT_MODULE_PAGES = {
    "classics": ["知识库", "知识图谱", "智能问答", "统计图表", "传播地图"],
    "shanghai": ["知识库", "知识图谱", "智能问答", "统计图表", "传播地图"],
    "stories": ["知识库", "知识图谱", "智能问答", "统计图表", "出版地地图", "来源地地图"],
    "world": ["知识库", "知识图谱", "智能问答", "统计图表", "关系图谱"],
}

SUBMODULES = {
    "classics": [
        {
            "id": "classic-translations",
            "name": "典籍译本总库图谱",
            "description": "典籍、译本、语种、国家与版本关系的总库分析。",
            "defaultFocus": "典籍-译本-语种三层知识图谱",
            "tablePlan": ["典籍总表", "译本表", "语种版本表", "再版表"],
            "visualPlan": ["全球典籍译介出版地图", "译本时间流变图", "语种结构环图", "典籍主题词云"],
            "hints": ["典籍", "译本", "语种版本", "再版", "classics", "translation"],
        },
        {
            "id": "classic-sinologists",
            "name": "译者与汉学家关系图谱",
            "description": "译者、汉学家、典籍、机构和身份流变关系分析。",
            "defaultFocus": "译者-典籍-机构网络图",
            "tablePlan": ["译者表", "汉学家表", "译者身份表", "译者-典籍关系表"],
            "visualPlan": ["译者身份演化图", "译者国别分布图", "译者活动地域分布图", "译者序跋关键词云"],
            "hints": ["译者", "汉学家", "身份", "sinologist", "translator"],
        },
        {
            "id": "classic-publishing",
            "name": "出版机构与地域传播图谱",
            "description": "出版机构、出版地、译本和语种通道的传播结构。",
            "defaultFocus": "出版地-语种-出版社传播通道图",
            "tablePlan": ["出版机构表", "出版地表", "版本表", "出版机构-译本关系表"],
            "visualPlan": ["出版机构排行", "出版年代密度图", "语种-出版地矩阵", "全球出版中心热力图"],
            "hints": ["出版", "机构", "地域", "出版社", "publisher", "press", "place"],
        },
        {
            "id": "classic-reception",
            "name": "接受史与研究文献图谱",
            "description": "评论、研究文献、接受事件、引用关系与主题聚类。",
            "defaultFocus": "典籍-接受文本-研究主题图谱",
            "tablePlan": ["评论表", "研究文献表", "接受事件表", "引用关系表"],
            "visualPlan": ["接受史时间线", "研究文献类型分布", "引用频次排行", "研究关键词聚类图"],
            "hints": ["接受", "评论", "研究", "文献", "引用", "reception", "review", "citation"],
        },
    ],
    "shanghai": [
        {
            "id": "shanghai-translations",
            "name": "上海文学译本总库图谱",
            "description": "上海文学作品、译本、语种和国家的总库结构。",
            "defaultFocus": "作品-译本-语种知识图谱",
            "tablePlan": ["上海文学作品表", "译本表", "语种版本表"],
            "visualPlan": ["译本时间流变图", "语种分布图", "作品译介排行", "作品关键词云"],
            "hints": ["上海文学", "译本", "作品", "语种", "literature", "translation"],
        },
        {
            "id": "shanghai-actors",
            "name": "作家—译者—出版社关系图谱",
            "description": "作家、作品、译者、出版社之间的三元传播关系。",
            "defaultFocus": "作家-译者-出版社三元关系图",
            "tablePlan": ["作家表", "译者表", "出版社表", "作品-译者关系表"],
            "visualPlan": ["核心作家排行", "译者参与度图", "出版社参与度图", "出版地与译者地域分布图"],
            "hints": ["作家", "译者", "出版社", "作者", "author", "translator", "publisher"],
        },
        {
            "id": "shanghai-urban-culture",
            "name": "上海城市文化主题图谱",
            "description": "上海意象、城市文化主题、作品和译本关系分析。",
            "defaultFocus": "上海城市意象主题网络图",
            "tablePlan": ["主题词表", "城市意象表", "文本片段表"],
            "visualPlan": ["城市主题频次图", "主题时间演化图", "主题-作品矩阵", "城市文化主题词云"],
            "hints": ["城市", "文化", "意象", "主题", "city", "urban", "theme"],
        },
        {
            "id": "shanghai-reception",
            "name": "海外接受与传播路径图谱",
            "description": "书评、研究文献、媒体报道、奖项与海外传播路径。",
            "defaultFocus": "作品-评论-接受主题图谱",
            "tablePlan": ["书评表", "研究文献表", "媒体报道表", "奖项表"],
            "visualPlan": ["接受史时间线", "评论类型分布", "海外研究热点图", "海外接受地区热力图"],
            "hints": ["接受", "传播", "路径", "书评", "媒体", "奖项", "地图", "source region", "reception", "route"],
        },
    ],
    "stories": [
        {
            "id": "stories-german-collections",
            "name": "百部德译故事集图谱",
            "description": "德译故事集、译者、出版社、年份、出版地和来源地的主体结构。",
            "defaultFocus": "德译故事集主体结构演化图",
            "tablePlan": ["德译故事集表", "再版传播表", "译者表", "出版社表"],
            "visualPlan": ["译者身份流变图", "故事集时间流变图", "出版机构参与度图", "德译故事集主题词云"],
            "hints": ["故事集总表", "出版地", "故事来源地", "德译", "publication", "source map"],
        },
        {
            "id": "stories-prefaces",
            "name": "序跋中文版本图谱",
            "description": "序跋文本、主题聚类、关键词、版本关系和文本可视化。",
            "defaultFocus": "序跋文本主题聚类图",
            "tablePlan": ["序跋文本表", "关键词表", "主题分类表", "版本关系表"],
            "visualPlan": ["序跋词云图", "序跋关键词共现网络", "序跋主题时间流变图", "序跋-版本关系图谱"],
            "hints": ["序跋", "preface", "主题聚类", "词云"],
        },
        {
            "id": "stories-child-knowledge",
            "name": "子故事知识图谱",
            "description": "子故事、母题、故事集、文献类型和主题共现关系。",
            "defaultFocus": "子故事语义-书目结构知识图谱",
            "tablePlan": ["子故事表", "母题表", "故事集-子故事关系表", "文献类型表"],
            "visualPlan": ["母题与文献类型耦合图", "子故事传播频次排行图", "子故事-故事集关系网络", "子故事主题词云"],
            "hints": ["子故事", "母题", "3533", "child"],
        },
        {
            "id": "stories-wilhelm",
            "name": "卫礼贤《中国民间童话》专题图谱",
            "description": "卫礼贤版本、单篇故事、关键词、再版传播和专题关系。",
            "defaultFocus": "卫礼贤《中国民间童话》总知识图谱",
            "tablePlan": ["卫礼贤版本表", "单篇故事表", "关键词表", "再版传播表"],
            "visualPlan": ["关键词共现网络分析图", "单篇译文结构谱系图", "再版传播时间密度图", "版本出版地传播地图"],
            "hints": ["中国民间童话", "卫礼贤", "wilhelm"],
        },
    ],
    "world": [
        {
            "id": "world-narrative-works",
            "name": "中国叙事作品总库图谱",
            "description": "世界文学作品、作者、中国叙事标签和文献类型总库。",
            "defaultFocus": "作者-作品-中国主题知识图谱",
            "tablePlan": ["世界文学作品表", "作者表", "中国叙事标签表"],
            "visualPlan": ["中国叙事作品时间流变图", "国家分布图", "文献类型分布图", "中国叙事关键词云"],
            "hints": ["数据库信息", "叙事", "作品", "work", "narrative"],
        },
        {
            "id": "world-image-themes",
            "name": "中国形象与主题谱系图谱",
            "description": "中国形象、主题标签、文本片段和主题共现谱系。",
            "defaultFocus": "中国形象与主题谱系图",
            "tablePlan": ["中国形象表", "主题标签表", "文本片段表"],
            "visualPlan": ["中国形象类型分布图", "主题演化时间线", "主题共现矩阵", "形象聚类图"],
            "hints": ["形象", "主题", "image", "theme"],
        },
        {
            "id": "world-adaptation-media",
            "name": "改编—翻译—媒介传播图谱",
            "description": "原作、改编、译本、媒介类型和传播事件链路。",
            "defaultFocus": "原作-改编-翻译-媒介链路图",
            "tablePlan": ["改编作品表", "译本表", "媒介类型表", "传播事件表"],
            "visualPlan": ["改编类型分布图", "媒介传播时间线", "翻译版本统计图", "传播路径地图"],
            "hints": ["改编", "媒介", "翻译", "adaptation", "media"],
        },
        {
            "id": "world-author-reception",
            "name": "作家—译者—接受关系图谱",
            "description": "作家、译者、评论、研究文献和接受主题关系。",
            "defaultFocus": "作家-译者-接受关系网络图",
            "tablePlan": ["作家表", "译者表", "评论表", "研究文献表"],
            "visualPlan": ["作家参与度图", "译者参与度图", "接受主题排行图", "接受地理分布图"],
            "hints": ["作家", "译者", "接受", "评论", "研究", "author", "translator", "reception"],
        },
    ],
}

SUBMODULE_DATASET_BINDINGS = {
    "stories:stories-german-collections": {
        "story-collections",
        "story-source-map",
        "story-publication-source-map",
        "map-story-kb",
    },
    "stories:stories-prefaces": {"story-prefaces"},
    "stories:stories-child-knowledge": {"story-children"},
    "stories:stories-wilhelm": {"wilhelm-text", "wilhelm-map"},
    "world:world-narrative-works": {"database-info"},
}

FIELD_ALIASES = {
    "canonicalTitle": [
        "规范名",
        "规范名称",
        "规范标题",
        "题名",
        "标题",
        "名称",
        "书名",
        "作品名",
        "故事集标题（中文）",
        "中文标题",
        "故事集标题",
        "子故事标题",
        "单篇译文故事名",
        "title",
        "name",
        "canonical title",
        "canonical name",
        "work title",
        "book title",
    ],
    "translatedTitle": [
        "译名",
        "译本名",
        "译文名",
        "外文名",
        "英文名",
        "变体名",
        "变异故事名",
        "故事名变体",
        "foreign title",
        "translated title",
        "translation title",
        "variant name",
        "variant",
    ],
    "year": ["年份", "年代", "出版年", "出版年份", "出版时间", "时间", "year", "date"],
    "language": ["语种", "语言", "译入语", "原语种", "language", "lang"],
    "country": ["国家", "地区", "国别", "国籍", "来源国家", "country", "nation", "region"],
    "city": ["城市", "出版地", "地点", "地点城市", "place", "city", "location"],
    "publisher": ["出版社", "出版机构", "机构", "publisher", "press"],
    "translator": ["译者", "翻译者", "译者/编者", "编译者", "编者", "translator", "editor", "compiler"],
    "author": ["作者", "作家", "序跋作者", "创作者", "creator", "author", "writer"],
    "summary": ["摘要", "简介", "说明", "内容", "备注", "注释", "译文内容", "序跋文本", "故事集介绍", "summary", "note", "notes", "text"],
}

SUMMARY_EXCLUDE_FIELDS = {
    "canonicalTitle",
    "translatedTitle",
    "year",
    "language",
    "country",
    "city",
    "publisher",
    "translator",
    "author",
}


def _safe_dataset_id(value: str) -> str:
    cleaned = "".join(ch.lower() if ch.isalnum() else "-" for ch in str(value or "").strip())
    cleaned = "-".join(part for part in cleaned.split("-") if part)
    return cleaned[:60] or f"table-{int(time.time())}"


def _custom_datasets(manifest: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    manifest = manifest if manifest is not None else _read_manifest()
    custom = manifest.get("_customDatasets", [])
    return custom if isinstance(custom, list) else []


def _all_datasets(manifest: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    return [*DATASETS, *_custom_datasets(manifest)]


def dataset_by_id(dataset_id: str) -> dict[str, Any]:
    for item in _all_datasets():
        if item["id"] == dataset_id:
            return item
    raise KeyError("未知数据集。")


def _read_manifest() -> dict[str, Any]:
    if not MANIFEST_PATH.exists():
        return {}
    try:
        return json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _write_manifest(payload: dict[str, Any]) -> None:
    MANIFEST_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _file_meta(filename: str) -> dict[str, Any]:
    path = DATA_DIR / filename
    if not path.exists():
        public_path = PUBLIC_ASSETS / filename
        path = public_path if public_path.exists() else path
    if not path.exists():
        return {"exists": False, "size": 0, "updatedAt": None}
    stat = path.stat()
    return {"exists": True, "size": stat.st_size, "updatedAt": int(stat.st_mtime)}


def _dataset_path(item: dict[str, Any]) -> Path:
    path = DATA_DIR / item["filename"]
    if path.exists():
        return path
    public_path = PUBLIC_ASSETS / item["filename"]
    return public_path if public_path.exists() else path


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def _clip(value: Any, limit: int = 520) -> str:
    text = _clean_text(value)
    return text if len(text) <= limit else f"{text[: limit - 1]}…"


def _normalise_header(value: Any) -> str:
    text = _clean_text(value).lower()
    return "".join(ch for ch in text if ch.isalnum())


def _as_list(value: Any) -> list[str]:
    if isinstance(value, list):
        source = value
    elif isinstance(value, tuple):
        source = list(value)
    else:
        source = str(value or "").replace("，", ",").split(",")
    result: list[str] = []
    for item in source:
        text = _clean_text(item)
        if text and text not in result:
            result.append(text)
    return result


def _unique_headers(headers: list[Any]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for index, header in enumerate(headers):
        text = _clean_text(header) or f"字段{index + 1}"
        count = counts.get(text, 0)
        counts[text] = count + 1
        result.append(text if count == 0 else f"{text}_{count + 1}")
    return result


def _records_from_matrix(raw_rows: list[list[Any]]) -> tuple[list[str], list[dict[str, str]]]:
    if not raw_rows:
        return [], []
    headers = _unique_headers(raw_rows[0])
    records: list[dict[str, str]] = []
    for row in raw_rows[1:]:
        if not any(_clean_text(cell) for cell in row):
            continue
        records.append({header: _clean_text(row[index] if index < len(row) else "") for index, header in enumerate(headers)})
    return headers, records


def _table_records(path: Path, limit: int | None = None) -> dict[str, Any]:
    suffix = path.suffix.lower()
    max_rows = limit if limit is not None else 200000
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            raw_rows = []
            for index, row in enumerate(reader):
                if index > max_rows:
                    break
                raw_rows.append(row)
        headers, records = _records_from_matrix(raw_rows)
        return {"headers": headers, "records": records, "sheet": path.name}

    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        records = data if isinstance(data, list) else data.get("items") or data.get("rows") or data.get("data") or []
        if not isinstance(records, list):
            records = []
        if records and not isinstance(records[0], dict):
            headers, mapped = _records_from_matrix(records[: max_rows + 1])
            return {"headers": headers, "records": mapped, "sheet": path.name}
        headers = list(dict.fromkeys(key for row in records if isinstance(row, dict) for key in row.keys()))
        mapped = [
            {header: _clean_text(row.get(header, "")) for header in headers}
            for row in records[:max_rows]
            if isinstance(row, dict)
        ]
        return {"headers": headers, "records": mapped, "sheet": path.name}

    if suffix in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            raw_rows = []
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index > max_rows:
                    break
                raw_rows.append(["" if value is None else value for value in row])
            headers, records = _records_from_matrix(raw_rows)
            return {"headers": headers, "records": records, "sheet": sheet.title}
        except Exception:
            preview = _xlsx_preview(path, max_rows)
            headers, records = _records_from_matrix([preview.get("headers") or [], *(preview.get("rows") or [])])
            return {"headers": headers, "records": records, "sheet": preview.get("sheet") or "Sheet1"}

    raise ValueError("暂不支持解析该文件类型。")


def _field_lookup(headers: list[str]) -> dict[str, str]:
    normalised = {_normalise_header(header): header for header in headers if _clean_text(header)}
    lookup: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            header = normalised.get(_normalise_header(alias))
            if header:
                lookup[field] = header
                break
    return lookup


def _row_value(row: dict[str, Any], lookup: dict[str, str], field: str) -> str:
    return _clean_text(row.get(lookup.get(field, ""), ""))


def _row_year(value: Any) -> int | str:
    text = _clean_text(value)
    if not text:
        return ""
    matched = re.search(r"(?:16|17|18|19|20)\d{2}", text)
    return int(matched.group(0)) if matched else text


def _first_non_empty(row: dict[str, Any], headers: list[str]) -> str:
    for header in headers:
        value = _clean_text(row.get(header))
        if value:
            return value
    return "未命名条目"


def _row_summary(row: dict[str, Any], headers: list[str], lookup: dict[str, str], key_fields: list[str]) -> str:
    summary = _row_value(row, lookup, "summary")
    if summary:
        return _clip(summary)
    mapped_headers = {lookup[field] for field in SUMMARY_EXCLUDE_FIELDS if lookup.get(field)}
    preferred = [field for field in key_fields if field in row and field not in mapped_headers]
    fallback = [header for header in headers if header not in mapped_headers and header not in preferred]
    parts = []
    for header in [*preferred, *fallback]:
        value = _clean_text(row.get(header))
        if value:
            parts.append(f"{header}: {value}")
        if len(parts) >= 8:
            break
    return _clip("；".join(parts))


def _stable_row_id(dataset_id: str, index: int, row: dict[str, Any]) -> str:
    digest = hashlib.sha1(json.dumps(row, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()[:12]
    return f"uploaded-{dataset_id}-{index + 1}-{digest}"


def _custom_dataset_cache_key(item: dict[str, Any], path: Path, manifest: dict[str, Any]) -> tuple[Any, ...]:
    stat = path.stat()
    upload = manifest.get(item["id"]) or {}
    return (
        item.get("id"),
        item.get("moduleId"),
        item.get("filename"),
        stat.st_mtime_ns,
        stat.st_size,
        tuple(item.get("requiredColumns") or []),
        upload.get("uploadedAt"),
    )


def _dataset_bundle_cache_key(item: dict[str, Any], path: Path, manifest: dict[str, Any], row_limit: int) -> tuple[Any, ...]:
    stat = path.stat()
    upload = manifest.get(item["id"]) or {}
    return (
        item.get("id"),
        item.get("moduleId"),
        item.get("filename"),
        item.get("title"),
        item.get("content"),
        bool(item.get("custom")),
        row_limit,
        stat.st_mtime_ns,
        stat.st_size,
        tuple(_as_list(item.get("pages"))),
        tuple(_as_list(item.get("requiredColumns"))),
        tuple(_as_list(item.get("detectedHeaders"))),
        upload.get("uploadedAt"),
        upload.get("rowCount"),
        upload.get("sheet"),
    )


def _dataset_search_text(item: dict[str, Any]) -> str:
    parts = [
        item.get("id"),
        item.get("title"),
        item.get("filename"),
        item.get("content"),
        item.get("module"),
        item.get("moduleId"),
        " ".join(_as_list(item.get("pages"))),
        " ".join(_as_list(item.get("requiredColumns"))),
        " ".join(_as_list(item.get("detectedHeaders"))),
    ]
    return " ".join(_clean_text(part).lower() for part in parts if _clean_text(part))


def _submodule_for_dataset(item: dict[str, Any], module_id: str) -> str:
    bound = SUBMODULE_DATASET_BINDINGS.get(f"{module_id}:{item.get('id')}")
    if bound is not None or item.get("id") in {
        dataset_id
        for key, dataset_ids in SUBMODULE_DATASET_BINDINGS.items()
        if key.startswith(f"{module_id}:")
        for dataset_id in dataset_ids
    }:
        for key, dataset_ids in SUBMODULE_DATASET_BINDINGS.items():
            if key.startswith(f"{module_id}:") and item.get("id") in dataset_ids:
                return key.split(":", 1)[1]

    text = _dataset_search_text(item)
    submodules = SUBMODULES.get(module_id) or []
    scored: list[tuple[int, str]] = []
    for submodule in submodules:
        score = 0
        for hint in submodule.get("hints") or []:
            hint_text = _clean_text(hint).lower()
            if hint_text and hint_text in text:
                score += 2 if len(hint_text) > 3 else 1
        for token in [submodule.get("name"), submodule.get("description")]:
            for word in re.findall(r"[\u4e00-\u9fff]{2,6}|[A-Za-z][A-Za-z\-]{3,}", _clean_text(token).lower()):
                if word in text:
                    score += 1
        scored.append((score, submodule["id"]))
    matched = max(scored, default=(0, ""))[1] if scored and max(scored)[0] > 0 else ""
    return matched or (submodules[0]["id"] if submodules else "overview")


def _submodule_summary(module_id: str, submodule_id: str, datasets: list[dict[str, Any]]) -> dict[str, Any] | None:
    config = next((item for item in SUBMODULES.get(module_id, []) if item.get("id") == submodule_id), None)
    if not config:
        return None
    sub_datasets = [dataset for dataset in datasets if dataset.get("submoduleId") == submodule_id]
    row_count = sum(dataset.get("stats", {}).get("rowCount", 0) for dataset in sub_datasets)
    return {
        **config,
        "moduleId": module_id,
        "datasetCount": len(sub_datasets),
        "rowCount": row_count,
        "datasets": [
            {
                "id": dataset.get("id"),
                "title": dataset.get("title"),
                "filename": dataset.get("filename"),
                "custom": dataset.get("custom"),
                "rowCount": dataset.get("stats", {}).get("rowCount", 0),
                "columnCount": dataset.get("stats", {}).get("columnCount", 0),
                "previewRowCount": dataset.get("previewRowCount", 0),
            }
            for dataset in sub_datasets
        ],
    }


def _dataset_to_knowledge_items(item: dict[str, Any], manifest: dict[str, Any]) -> list[dict[str, Any]]:
    path = _dataset_path(item)
    if not path.exists():
        return []
    key = _custom_dataset_cache_key(item, path, manifest)
    cached = CUSTOM_KNOWLEDGE_CACHE.get(item["id"])
    if cached and cached[0] == key:
        return [dict(row) for row in cached[1]]

    table = _table_records(path)
    headers = table.get("headers") or []
    records = table.get("records") or []
    lookup = _field_lookup(headers)
    module_id = item.get("moduleId") or MODULE_ID_BY_NAME.get(item.get("module")) or "stories"
    upload = manifest.get(item["id"]) or {}
    key_fields = _as_list(item.get("requiredColumns"))
    used_headers = {header for header in lookup.values() if header}
    source = f"{item.get('filename')}#{table.get('sheet') or 'Sheet1'}"
    rows: list[dict[str, Any]] = []

    for index, row in enumerate(records):
        canonical = _row_value(row, lookup, "canonicalTitle") or _first_non_empty(row, headers)
        translated = _row_value(row, lookup, "translatedTitle") or canonical
        year = _row_year(_row_value(row, lookup, "year"))
        language = _row_value(row, lookup, "language")
        country = _row_value(row, lookup, "country")
        city = _row_value(row, lookup, "city")
        publisher = _row_value(row, lookup, "publisher")
        translator = _row_value(row, lookup, "translator")
        author = _row_value(row, lookup, "author")
        summary = _row_summary(row, headers, lookup, key_fields)
        selected_values = [
            f"{field}: {_clean_text(row.get(field))}"
            for field in key_fields
            if _clean_text(row.get(field)) and field not in used_headers
        ]
        relations = [
            f"所属表格 -> {item.get('title')}",
            f"影响页面 -> {' / '.join(item.get('pages') or [])}",
            f"关键字段 -> {'；'.join(selected_values or key_fields)}",
            f"译者/作者 -> {translator or author}",
            f"出版节点 -> {' / '.join(part for part in [city, country, publisher] if part)}",
        ]
        rows.append({
            "id": _stable_row_id(item["id"], index, row),
            "status": "真实上传",
            "sectionId": module_id,
            "resourceType": item.get("title") or "管理员上传表格",
            "canonicalTitle": canonical,
            "translatedTitle": translated,
            "author": author,
            "translator": translator,
            "language": language,
            "country": country,
            "city": city,
            "publisher": publisher,
            "year": year,
            "uploadedAt": upload.get("uploadedAt") or item.get("createdAt") or "",
            "uploader": upload.get("uploadedBy") or item.get("createdBy") or "管理员",
            "summary": summary,
            "tags": [field for field in key_fields if field in row],
            "evidence": [source],
            "source": source,
            "sourceKind": "uploaded-table",
            "datasetId": item.get("id"),
            "bookName": item.get("title") or item.get("filename"),
            "relations": [line for line in relations if line and not line.endswith("-> ")],
            "searchText": " ".join(_clean_text(value) for value in [canonical, translated, year, language, country, city, publisher, translator, author, summary, *row.values()] if _clean_text(value)),
            "graphNodeIds": [f"dataset:{item.get('id')}", f"dataset-row:{item.get('id')}:{index + 1}"],
            "raw": row,
        })

    CUSTOM_KNOWLEDGE_CACHE[item["id"]] = (key, rows)
    return [dict(row) for row in rows]


def _top_counts(records: list[dict[str, Any]], fields: list[str], limit: int = 8) -> list[list[Any]]:
    counts: Counter[str] = Counter()
    for row in records:
        value = ""
        for field in fields:
            value = _clean_text(row.get(field))
            if value:
                break
        if value:
            counts[value] += 1
    return [[name, count] for name, count in counts.most_common(limit)]


def _first_header(headers: list[str], lookup: dict[str, str], field: str, fallback: list[str] | None = None) -> str:
    if lookup.get(field):
        return lookup[field]
    candidates = fallback or []
    normalised = {_normalise_header(header): header for header in headers}
    for candidate in candidates:
        header = normalised.get(_normalise_header(candidate))
        if header:
            return header
    return ""


def _dataset_columns(headers: list[str], lookup: dict[str, str], required: list[str]) -> list[str]:
    preferred = [
        _first_header(headers, lookup, "canonicalTitle"),
        _first_header(headers, lookup, "translatedTitle"),
        _first_header(headers, lookup, "year"),
        _first_header(headers, lookup, "language"),
        _first_header(headers, lookup, "country"),
        _first_header(headers, lookup, "city"),
        _first_header(headers, lookup, "publisher"),
        _first_header(headers, lookup, "translator"),
        _first_header(headers, lookup, "author"),
        *required,
        *headers,
    ]
    result: list[str] = []
    for header in preferred:
        header = _clean_text(header)
        if header and header in headers and header not in result:
            result.append(header)
        if len(result) >= 10:
            break
    return result or headers[:10]


def _record_to_visual_item(row: dict[str, Any], headers: list[str], lookup: dict[str, str], item: dict[str, Any], index: int) -> dict[str, Any]:
    canonical = _row_value(row, lookup, "canonicalTitle") or _first_non_empty(row, headers)
    translated = _row_value(row, lookup, "translatedTitle") or canonical
    year = _row_year(_row_value(row, lookup, "year"))
    language = _row_value(row, lookup, "language")
    country = _row_value(row, lookup, "country")
    city = _row_value(row, lookup, "city")
    publisher = _row_value(row, lookup, "publisher")
    translator = _row_value(row, lookup, "translator")
    author = _row_value(row, lookup, "author")
    return {
        "id": _stable_row_id(item["id"], index, row),
        "datasetId": item.get("id"),
        "sectionId": item.get("moduleId") or "stories",
        "resourceType": item.get("title") or item.get("filename"),
        "canonicalTitle": canonical,
        "translatedTitle": translated,
        "year": year,
        "language": language,
        "country": country,
        "city": city,
        "publisher": publisher,
        "translator": translator,
        "author": author,
        "summary": _row_summary(row, headers, lookup, _as_list(item.get("requiredColumns"))),
        "source": item.get("filename"),
        "sourceKind": "uploaded-table" if item.get("custom") else "table",
        "raw": row,
    }


def _dataset_flows(records: list[dict[str, Any]], headers: list[str], lookup: dict[str, str], item: dict[str, Any]) -> list[dict[str, Any]]:
    flows = []
    for index, row in enumerate(records):
        city = _row_value(row, lookup, "city")
        country = _row_value(row, lookup, "country")
        year = _row_year(_row_value(row, lookup, "year"))
        source_region = _clean_text(
            row.get(_first_header(headers, lookup, "source", ["来源地", "故事来源", "source region", "source", "省份", "地区"]))
        )
        if not city and not country and not source_region:
            continue
        flows.append({
            "id": f"{item.get('id')}-flow-{index + 1}",
            "sectionId": item.get("moduleId") or "stories",
            "datasetId": item.get("id"),
            "year": year if isinstance(year, int) else 0,
            "from": CHINA_COORDS,
            "to": [10.45, 51.16],
            "fromLabel": source_region or "中国",
            "toLabel": " · ".join(part for part in [city, country] if part) or "海外",
            "city": city,
            "country": country,
            "title": _row_value(row, lookup, "canonicalTitle") or _first_non_empty(row, headers),
        })
    return flows[:300]


def _dataset_graph(records: list[dict[str, Any]], headers: list[str], lookup: dict[str, str], item: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    module_id = item.get("moduleId") or "stories"
    dataset_id = item.get("id")
    nodes: dict[str, dict[str, Any]] = {
        f"dataset:{dataset_id}": {
            "id": f"dataset:{dataset_id}",
            "label": item.get("title") or item.get("filename"),
            "type": "数据表",
            "section": module_id,
            "size": 24,
        }
    }
    edges: list[dict[str, Any]] = []
    entity_fields = [
        ("translator", "译者/编者", "人物"),
        ("author", "作者", "人物"),
        ("publisher", "出版机构", "机构"),
        ("language", "语种", "语种"),
        ("country", "国家/地区", "地点"),
        ("city", "城市", "地点"),
    ]
    for index, row in enumerate(records[:120]):
        title = _row_value(row, lookup, "canonicalTitle") or _first_non_empty(row, headers)
        row_id = f"dataset-row:{dataset_id}:{index + 1}"
        nodes[row_id] = {
            "id": row_id,
            "label": title,
            "type": "表格记录",
            "section": module_id,
            "year": _row_year(_row_value(row, lookup, "year")),
            "lang": _row_value(row, lookup, "language"),
            "size": 8,
        }
        edges.append({"from": f"dataset:{dataset_id}", "to": row_id, "relation": "包含记录"})
        for field, relation, node_type in entity_fields:
            value = _row_value(row, lookup, field)
            if not value:
                continue
            node_id = f"{field}:{hashlib.sha1(value.encode('utf-8')).hexdigest()[:12]}"
            nodes.setdefault(node_id, {
                "id": node_id,
                "label": value,
                "type": node_type,
                "section": module_id,
                "size": 10,
            })
            edges.append({"from": row_id, "to": node_id, "relation": relation})
    return {"nodes": list(nodes.values()), "edges": edges[:420]}


def _dataset_keywords(records: list[dict[str, Any]], headers: list[str], lookup: dict[str, str], limit: int = 42) -> list[dict[str, Any]]:
    skip = {lookup.get(field) for field in SUMMARY_EXCLUDE_FIELDS if lookup.get(field)}
    counts: Counter[str] = Counter()
    for row in records:
        for header in headers:
            if header in skip:
                continue
            text = _clean_text(row.get(header))
            if not text:
                continue
            for token in re.findall(r"[\u4e00-\u9fff]{2,8}|[A-Za-z][A-Za-z\-]{3,}", text):
                if len(token) < 2:
                    continue
                counts[token] += 1
    return [{"term": term, "count": count} for term, count in counts.most_common(limit)]


def _dataset_visual_bundle(item: dict[str, Any], manifest: dict[str, Any], row_limit: int = 500) -> dict[str, Any] | None:
    path = _dataset_path(item)
    if not path.exists():
        return None
    cache_key = _dataset_bundle_cache_key(item, path, manifest, row_limit)
    cached = MODULE_DATASET_CACHE.get(item["id"])
    if cached and cached[0] == cache_key:
        return cached[1]

    table = _table_records(path)
    headers = table.get("headers") or []
    records = table.get("records") or []
    preview_records = records[:row_limit]
    lookup = _field_lookup(headers)
    required = _as_list(item.get("requiredColumns"))
    columns = _dataset_columns(headers, lookup, required)
    rows = [{column: _clean_text(record.get(column)) for column in columns} for record in preview_records]
    visual_items = [_record_to_visual_item(record, headers, lookup, item, index) for index, record in enumerate(preview_records)]
    stats = {
        "rowCount": len(records),
        "columnCount": len(headers),
        "languageTop": _top_counts(records, [_first_header(headers, lookup, "language")]),
        "countryTop": _top_counts(records, [_first_header(headers, lookup, "country")]),
        "publisherTop": _top_counts(records, [_first_header(headers, lookup, "publisher")]),
        "authorTop": _top_counts(records, [_first_header(headers, lookup, "translator"), _first_header(headers, lookup, "author")]),
        "yearTop": _top_counts(records, [_first_header(headers, lookup, "year")]),
    }
    upload = manifest.get(item["id"]) or {}
    module_id = item.get("moduleId") or "stories"
    submodule_id = _submodule_for_dataset(item, module_id)
    package = {
        "id": item.get("id"),
        "title": item.get("title") or item.get("filename"),
        "filename": item.get("filename"),
        "moduleId": module_id,
        "module": MODULE_NAME_BY_ID.get(module_id, item.get("module")),
        "submoduleId": submodule_id,
        "submodule": next((entry.get("name") for entry in SUBMODULES.get(module_id, []) if entry.get("id") == submodule_id), ""),
        "pages": item.get("pages") or [],
        "content": item.get("content") or "",
        "requiredColumns": required,
        "custom": bool(item.get("custom")),
        "exists": True,
        "updatedAt": upload.get("uploadedAt") or _file_meta(item["filename"]).get("updatedAt"),
        "sheet": table.get("sheet") or upload.get("sheet") or "Sheet1",
        "columns": columns,
        "allColumns": headers,
        "rows": rows,
        "previewRowCount": len(rows),
        "rowLimit": row_limit,
        "hasMoreRows": len(records) > len(rows),
        "items": visual_items,
        "stats": stats,
        "graph": _dataset_graph(records, headers, lookup, item),
        "flows": _dataset_flows(records, headers, lookup, item),
        "keywords": _dataset_keywords(records, headers, lookup),
    }
    MODULE_DATASET_CACHE[item["id"]] = (cache_key, package)
    return package


def _dataset_summary_bundle(item: dict[str, Any], manifest: dict[str, Any]) -> dict[str, Any] | None:
    path = _dataset_path(item)
    if not path.exists():
        return None
    module_id = item.get("moduleId") or "stories"
    key = _dataset_bundle_cache_key(item, path, manifest, 0)
    cached = MODULE_DATASET_SUMMARY_CACHE.get(item["id"])
    if cached and cached[0] == key:
        return cached[1]
    table = _table_records(path)
    headers = table.get("headers") or []
    records = table.get("records") or []
    submodule_id = _submodule_for_dataset(item, module_id)
    upload = manifest.get(item["id"]) or {}
    package = {
        "id": item.get("id"),
        "title": item.get("title") or item.get("filename"),
        "filename": item.get("filename"),
        "moduleId": module_id,
        "module": MODULE_NAME_BY_ID.get(module_id, item.get("module")),
        "submoduleId": submodule_id,
        "submodule": next((entry.get("name") for entry in SUBMODULES.get(module_id, []) if entry.get("id") == submodule_id), ""),
        "custom": bool(item.get("custom")),
        "exists": True,
        "updatedAt": upload.get("uploadedAt") or _file_meta(item["filename"]).get("updatedAt"),
        "sheet": table.get("sheet") or upload.get("sheet") or "Sheet1",
        "stats": {
            "rowCount": len(records),
            "columnCount": len(headers),
        },
    }
    MODULE_DATASET_SUMMARY_CACHE[item["id"]] = (key, package)
    return package


def module_dataset_packages(module_id: str | None = None, submodule_id: str | None = None, summary: bool = False) -> dict[str, Any]:
    manifest = _read_manifest()
    datasets = []
    for item in _all_datasets(manifest):
        current_module = item.get("moduleId") or MODULE_ID_BY_NAME.get(item.get("module")) or "stories"
        if module_id and current_module != module_id:
            continue
        current_submodule = _submodule_for_dataset({**item, "moduleId": current_module}, current_module)
        if submodule_id and current_submodule != submodule_id:
            continue
        package = (_dataset_summary_bundle if summary else _dataset_visual_bundle)({**item, "moduleId": current_module}, manifest)
        if package:
            datasets.append(package)
    modules = []
    for module in MODULES:
        if module_id and module["id"] != module_id:
            continue
        module_datasets = [dataset for dataset in datasets if dataset.get("moduleId") == module["id"]]
        submodules = [
            summary_item
            for summary_item in (
                _submodule_summary(module["id"], submodule["id"], module_datasets)
                for submodule in SUBMODULES.get(module["id"], [])
            )
            if summary_item
        ]
        modules.append({
            **module,
            "datasets": module_datasets,
            "submodules": submodules,
            "datasetCount": len(module_datasets),
            "rowCount": sum(dataset.get("stats", {}).get("rowCount", 0) for dataset in module_datasets),
        })
    return {"modules": modules, "datasets": datasets, "summary": summary}


def custom_knowledge_items(module_id: str | None = None) -> list[dict[str, Any]]:
    manifest = _read_manifest()
    items: list[dict[str, Any]] = []
    for item in _all_datasets(manifest):
        current_module = item.get("moduleId") or MODULE_ID_BY_NAME.get(item.get("module")) or "stories"
        if module_id and current_module != module_id:
            continue
        if not item.get("custom") and not (current_module != "stories" and item.get("rebuild") == CUSTOM_REBUILD):
            continue
        try:
            items.extend(_dataset_to_knowledge_items({**item, "moduleId": current_module}, manifest))
        except Exception:
            continue
    return items


def dataset_manifest() -> dict[str, Any]:
    uploads = _read_manifest()
    items = []
    for item in _all_datasets(uploads):
        module_id = item.get("moduleId") or MODULE_ID_BY_NAME.get(item.get("module")) or "stories"
        items.append({
            **item,
            "moduleId": module_id,
            "module": MODULE_NAME_BY_ID.get(module_id, item.get("module") or "其他"),
            **_file_meta(item["filename"]),
            "lastUpload": uploads.get(item["id"]),
        })
    return {"modules": MODULES, "datasets": items}


def _copy_to_runtime_assets(path: Path, item: dict[str, Any]) -> None:
    PUBLIC_ASSETS.mkdir(parents=True, exist_ok=True)
    targets = [PUBLIC_ASSETS / item["filename"]]
    for alias in item.get("aliases", []):
        targets.append(PUBLIC_ASSETS / alias)
    for target in targets:
        shutil.copyfile(path, target)


def _clear_backend_caches() -> None:
    try:
        from backend.app.core import data as data_module
        from backend.app.core import story_visuals
        from backend.app.core import chat_retrieval

        data_module.refresh_data()
        MODULE_DATASET_CACHE.clear()
        try:
            from backend.app.api import routes

            routes.GRAPH_DATA = data_module.GRAPH_DATA
            routes.KNOWLEDGE_ITEMS = data_module.KNOWLEDGE_ITEMS
            routes.MAP_FLOWS = data_module.MAP_FLOWS
        except Exception:
            pass
        story_visuals._story_data_cached.cache_clear()
        story_visuals.workbook_knowledge.cache_clear()
        story_visuals.publication_workbook_rows.cache_clear()
        story_visuals._publication_source_map_cached.cache_clear()
        story_visuals._visual_atlas_cached.cache_clear()
        if hasattr(chat_retrieval, "search_documents"):
            chat_retrieval.search_documents.cache_clear()
    except Exception:
        pass


def _run_script(path: Path) -> str:
    if not path.exists():
        return ""
    completed = subprocess.run(
        [sys.executable, str(path)],
        cwd=str(ROOT),
        text=True,
        capture_output=True,
        check=False,
    )
    output = "\n".join(part for part in [completed.stdout.strip(), completed.stderr.strip()] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"{path.name} 执行失败。")
    return output


def rebuild_for(dataset_id: str) -> dict[str, Any]:
    item = dataset_by_id(dataset_id)
    mode = item.get("rebuild")
    logs: list[str] = []
    if mode in {"story", "story-and-publication-map"}:
        logs.append(_run_script(BUILD_STORY_SCRIPT))
    if mode == "story-and-publication-map":
        logs.append(_run_script(BUILD_PUBLICATION_MAP_SCRIPT))
    _clear_backend_caches()
    return {"ok": True, "logs": [line for line in logs if line]}


def _xlsx_preview(path: Path, limit: int) -> dict[str, Any]:
    try:
      from openpyxl import load_workbook

      workbook = load_workbook(path, read_only=True, data_only=True)
      sheet = workbook[workbook.sheetnames[0]]
      rows = []
      for index, row in enumerate(sheet.iter_rows(values_only=True)):
          if index >= limit + 1:
              break
          rows.append(["" if value is None else str(value) for value in row])
      headers = rows[0] if rows else []
      return {"headers": headers, "rows": rows[1:], "sheet": sheet.title}
    except Exception:
      pass

    namespaces = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as archive:
        shared = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall(".//a:si", namespaces):
                shared.append("".join(node.text or "" for node in item.findall(".//a:t", namespaces)))
        sheet_name = "xl/worksheets/sheet1.xml"
        root = ET.fromstring(archive.read(sheet_name))
        rows = []
        for row in root.findall(".//a:sheetData/a:row", namespaces)[: limit + 1]:
            values = []
            for cell in row.findall("a:c", namespaces):
                value = cell.find("a:v", namespaces)
                text = value.text if value is not None else ""
                if cell.attrib.get("t") == "s" and text.isdigit():
                    text = shared[int(text)] if int(text) < len(shared) else ""
                values.append(text or "")
            rows.append(values)
    return {"headers": rows[0] if rows else [], "rows": rows[1:], "sheet": "Sheet1"}


def dataset_preview(dataset_id: str, limit: int = 12) -> dict[str, Any]:
    item = dataset_by_id(dataset_id)
    path = DATA_DIR / item["filename"]
    if not path.exists():
        path = PUBLIC_ASSETS / item["filename"]
    if not path.exists():
        raise FileNotFoundError("表格文件不存在。")
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter=delimiter))
        return {"dataset": item, "headers": rows[0] if rows else [], "rows": rows[1: limit + 1], "sheet": path.name}
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        records = data if isinstance(data, list) else data.get("items") or data.get("rows") or []
        if not isinstance(records, list):
            records = []
        headers = sorted({key for row in records[:limit] if isinstance(row, dict) for key in row.keys()})
        rows = [[str(row.get(key, "")) for key in headers] for row in records[:limit] if isinstance(row, dict)]
        return {"dataset": item, "headers": headers, "rows": rows, "sheet": path.name}
    if suffix in {".xlsx", ".xls"}:
        return {"dataset": item, **_xlsx_preview(path, limit)}
    raise ValueError("暂不支持预览该文件类型。")


def create_dataset(payload: dict[str, Any], username: str = "admin") -> dict[str, Any]:
    module_id = str(payload.get("moduleId") or "stories")
    if module_id not in MODULE_NAME_BY_ID:
        raise ValueError("请选择有效模块。")
    title = str(payload.get("title") or "").strip()
    if not title:
        raise ValueError("表格名称不能为空。")
    filename = str(payload.get("filename") or title).strip()
    suffix = Path(filename).suffix.lower() or ".xlsx"
    if suffix not in {".xlsx", ".xls", ".csv", ".tsv", ".json"}:
        raise ValueError("仅支持 xlsx、xls、csv、tsv、json 表格文件。")
    if not filename.lower().endswith(suffix):
        filename = f"{filename}{suffix}"
    manifest = _read_manifest()
    existing = {item["id"] for item in _all_datasets(manifest)}
    dataset_id = _safe_dataset_id(payload.get("id") or f"{module_id}-{Path(filename).stem}")
    base_id = dataset_id
    counter = 2
    while dataset_id in existing:
        dataset_id = f"{base_id}-{counter}"
        counter += 1
    item = {
        "id": dataset_id,
        "title": title,
        "filename": filename,
        "module": MODULE_NAME_BY_ID[module_id],
        "moduleId": module_id,
        "pages": _as_list(payload.get("pages")) or DEFAULT_MODULE_PAGES.get(module_id, ["知识库", "知识图谱", "智能问答"]),
        "content": str(payload.get("content") or "管理员新增维护表格。").strip(),
        "requiredColumns": _as_list(payload.get("requiredColumns")),
        "detectedHeaders": _as_list(payload.get("detectedHeaders")),
        "rebuild": CUSTOM_REBUILD,
        "custom": True,
        "createdBy": username,
        "createdAt": int(time.time()),
    }
    custom = _custom_datasets(manifest)
    custom.append(item)
    manifest["_customDatasets"] = custom
    _write_manifest(manifest)
    return {**item, **_file_meta(item["filename"]), "lastUpload": None}


def delete_dataset(dataset_id: str) -> None:
    manifest = _read_manifest()
    custom = _custom_datasets(manifest)
    item = next((entry for entry in custom if entry.get("id") == dataset_id), None)
    if not item:
        raise ValueError("内置核心表格不能删除，只能上传更新或重建。")
    manifest["_customDatasets"] = [entry for entry in custom if entry.get("id") != dataset_id]
    manifest.pop(dataset_id, None)
    _write_manifest(manifest)


def upload_dataset(payload: dict[str, Any], username: str = "admin") -> dict[str, Any]:
    with DATASET_LOCK:
        item = dataset_by_id(str(payload.get("datasetId") or ""))
        filename = str(payload.get("filename") or item["filename"])
        suffix = Path(filename).suffix.lower()
        if suffix not in {".xlsx", ".xls", ".csv", ".tsv", ".json"}:
            raise ValueError("仅支持 xlsx、xls、csv、tsv、json 表格文件。")
        raw = str(payload.get("contentBase64") or "")
        if "," in raw and raw.startswith("data:"):
            raw = raw.split(",", 1)[1]
        try:
            content = base64.b64decode(raw, validate=False)
        except Exception as error:
            raise ValueError("文件内容不是有效 Base64。") from error
        if not content:
            raise ValueError("上传文件为空。")

        DATA_DIR.mkdir(parents=True, exist_ok=True)
        target = DATA_DIR / item["filename"]
        if target.exists():
            stamp = time.strftime("%Y%m%d-%H%M%S")
            backup = DATA_DIR / f"{target.stem}.bak-{stamp}{target.suffix}"
            shutil.copyfile(target, backup)
        target.write_bytes(content)
        _copy_to_runtime_assets(target, item)
        table = _table_records(target)

        uploads = _read_manifest()
        selected_pages = _as_list(payload.get("pages"))
        selected_columns = _as_list(payload.get("requiredColumns"))
        if item.get("custom"):
            custom = _custom_datasets(uploads)
            for entry in custom:
                if entry.get("id") != item["id"]:
                    continue
                if selected_pages:
                    entry["pages"] = selected_pages
                if selected_columns:
                    entry["requiredColumns"] = selected_columns
                entry["detectedHeaders"] = table.get("headers") or []
                entry["rowCount"] = len(table.get("records") or [])
                entry["sheet"] = table.get("sheet") or ""
                item = entry
                break
            uploads["_customDatasets"] = custom
        uploads[item["id"]] = {
            "filename": filename,
            "storedAs": item["filename"],
            "uploadedBy": username,
            "uploadedAt": int(time.time()),
            "size": len(content),
            "headers": table.get("headers") or [],
            "rowCount": len(table.get("records") or []),
            "sheet": table.get("sheet") or "",
        }
        _write_manifest(uploads)
        CUSTOM_KNOWLEDGE_CACHE.pop(item["id"], None)

        rebuild_result = {"ok": True, "logs": []}
        if payload.get("rebuild", True):
            rebuild_result = rebuild_for(item["id"])
        try:
            from backend.app.core.knowledge_indexes import rebuild_index

            rebuild_index(sync_external=False)
        except Exception:
            pass
        return {
            "ok": True,
            "dataset": {**item, **_file_meta(item["filename"]), "lastUpload": uploads[item["id"]]},
            "rebuild": rebuild_result,
            "headers": table.get("headers") or [],
            "rowCount": len(table.get("records") or []),
        }
