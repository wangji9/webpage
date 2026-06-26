from __future__ import annotations

import json
import re
import hashlib
import math
from collections import Counter, defaultdict
from functools import lru_cache
from itertools import combinations
from pathlib import Path
from typing import Any, Optional
from xml.etree import ElementTree as ET
import zipfile

from backend.app.core.llm_client import chat_completion

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    import jieba
    import jieba.analyse
    import jieba.posseg as pseg
except Exception:  # pragma: no cover - optional NLP dependency
    jieba = None
    pseg = None


ROOT = Path(__file__).resolve().parents[3]
STORY_DATA_PATH = ROOT / "frontend" / "src" / "data" / "storyCollections.json"
WILHELM_GRAPH_CACHE_PATH = ROOT / "backend" / "app" / "wilhelm_graph_cache.json"
WILHELM_KEYWORD_CACHE_PATH = ROOT / "backend" / "app" / "wilhelm_keyword_category_cache.json"
WILHELM_KEYWORD_NETWORK_CACHE_PATH = ROOT / "backend" / "app" / "wilhelm_keyword_network_cache.json"
EXTERNAL_KNOWLEDGE_WORKBOOKS = [
    "地图_中国故事集总表_知识库.xlsx",
    "地图_中国民间童话.xlsx",
    "数据库信息.xlsx",
    "中国故事集_序跋.xlsx",
    "中国故事集_子故事（3533篇）.xlsx",
    "中国故事集总表_知识库.xlsx",
]
PUBLICATION_MAP_WORKBOOK = "地图_中国故事集_出版地和故事来源地.xlsx"
PUBLICATION_SOURCE_MAP_JSON_PATH = ROOT / "frontend" / "src" / "data" / "wilhelmPublicationSourceMap.json"

STAGES = [
    {"id": "early", "title": "1910s-1930s", "note": "早期译介", "range": [1910, 1940]},
    {"id": "middle", "title": "1950s-1970s", "note": "学术整理", "range": [1950, 1980]},
    {"id": "late", "title": "1980s-2000s", "note": "文学与出版扩展", "range": [1980, 2010]},
    {"id": "recent", "title": "2010s-2020s", "note": "多主体参与", "range": [2010, 2030]},
]

CITY_COORDS = {
    "Berlin": [13.405, 52.52],
    "Jena": [11.5892, 50.9271],
    "München": [11.582, 48.1351],
    "Leipzig": [12.3731, 51.3397],
    "Frankfurt am Main": [8.6821, 50.1109],
    "Stuttgart": [9.1829, 48.7758],
    "Basel": [7.5886, 47.5596],
    "Sankt Augustin": [7.1902, 50.7754],
    "Esslingen": [9.3103, 48.7428],
    "Norderstedt": [9.9791, 53.7088],
    "Bickenbach": [8.6106, 49.7595],
    "Hamburg": [9.9937, 53.5511],
    "Köln": [6.9603, 50.9375],
    "Düsseldorf": [6.7735, 51.2277],
    "Freiburg": [7.8421, 47.999],
    "Eisennach": [10.3157, 50.9795],
    "Eisenach": [10.3157, 50.9795],
    "Kassel": [9.4797, 51.3127],
    "Zürich": [8.5417, 47.3769],
    "Prag": [14.4378, 50.0755],
    "Wien": [16.3738, 48.2082],
    "Peking": [116.4, 39.9],
    "Beijing": [116.4, 39.9],
    "北京": [116.4, 39.9],
    "Shanghai": [121.47, 31.23],
    "上海": [121.47, 31.23],
    "Bayreuth": [11.5783, 49.9456],
    "Meerbusch": [6.6897, 51.2529],
    "Augsburg": [10.8978, 48.3705],
    "Bielefeld": [8.5325, 52.0302],
    "Schiedlberg": [14.0546, 48.111],
    "Kreuzlingen": [9.175, 47.65],
}

PROVINCE_COORDS = {
    "北京": [116.4, 39.9],
    "天津": [117.2, 39.12],
    "河北": [114.5, 38.04],
    "山西": [112.55, 37.87],
    "内蒙古": [111.67, 40.82],
    "辽宁": [123.43, 41.8],
    "吉林": [125.32, 43.9],
    "黑龙江": [126.64, 45.76],
    "上海": [121.47, 31.23],
    "江苏": [118.8, 32.1],
    "浙江": [120.2, 30.3],
    "安徽": [117.27, 31.86],
    "福建": [119.3, 26.08],
    "江西": [115.86, 28.68],
    "山东": [117.0, 36.7],
    "河南": [113.62, 34.75],
    "湖北": [114.3, 30.6],
    "湖南": [112.98, 28.2],
    "广东": [113.27, 23.13],
    "广西": [108.32, 22.82],
    "海南": [110.35, 20.02],
    "重庆": [106.55, 29.56],
    "四川": [104.06, 30.67],
    "贵州": [106.63, 26.65],
    "云南": [102.71, 25.04],
    "西藏": [91.13, 29.65],
    "陕西": [108.94, 34.34],
    "甘肃": [103.82, 36.06],
    "青海": [101.78, 36.62],
    "宁夏": [106.27, 38.47],
    "新疆": [87.62, 43.82],
    "台湾": [121.0, 23.7],
    "香港": [114.17, 22.32],
    "澳门": [113.55, 22.2],
}

CITY_ZH = {
    "Berlin": "柏林",
    "Jena": "耶拿",
    "München": "慕尼黑",
    "Leipzig": "莱比锡",
    "Frankfurt am Main": "法兰克福",
    "Stuttgart": "斯图加特",
    "Basel": "巴塞尔",
    "Sankt Augustin": "圣奥古斯丁",
    "Esslingen": "埃斯林根",
    "Norderstedt": "诺德施泰特",
    "Bickenbach": "比肯巴赫",
    "Hamburg": "汉堡",
    "Köln": "科隆",
    "Düsseldorf": "杜塞尔多夫",
    "Freiburg": "弗赖堡",
    "Eisennach": "艾森纳赫",
    "Eisenach": "艾森纳赫",
    "Kassel": "卡塞尔",
    "Zürich": "苏黎世",
    "Prag": "布拉格",
    "Wien": "维也纳",
    "Peking": "北京",
    "Beijing": "北京",
    "北京": "北京",
    "Shanghai": "上海",
    "上海": "上海",
    "Bayreuth": "拜罗伊特",
    "Meerbusch": "梅尔布施",
    "Augsburg": "奥格斯堡",
    "Bielefeld": "比勒费尔德",
    "Schiedlberg": "席德尔贝格",
    "Kreuzlingen": "克罗伊茨林根",
}

PREFACE_CLUSTERS = [
    {"id": "translate", "title": "翻译与改写", "terms": ["翻译", "改写", "选编", "转述", "忠实", "传播"]},
    {"id": "image", "title": "中国形象", "terms": ["中国", "东方", "古老", "神秘", "智慧", "异域"]},
    {"id": "folk", "title": "民间叙事", "terms": ["民间故事", "童话", "传说", "寓言", "神怪", "动物故事"]},
    {"id": "reader", "title": "读者与教育", "terms": ["儿童", "教育", "读者", "道德", "学习", "后记"]},
    {"id": "mediation", "title": "文化中介", "terms": ["汉学", "译者", "出版社", "交流", "理解", "文化"]},
]

STOP_WORDS = {
    "the", "and", "und", "der", "die", "das", "aus", "von", "mit", "für", "eine", "einer", "oder", "aber",
    "中国", "故事", "民间", "童话", "文本", "版本", "作者", "一个", "一些", "我们", "他们", "人们", "什么",
    "这里", "其中", "因此", "以及", "同时", "后来", "比如", "例如", "所谓", "进行", "可以", "没有", "未记录",
    "不同", "出现", "作品", "方式", "形式", "一方面", "另一方面", "一部分", "一部", "一种", "一点",
    "几个", "许多", "很多", "多少", "数量", "第一", "第二", "之一", "两个", "三个", "各个", "某些", "这种", "那种",
    "几乎", "仍然", "于是", "这个", "那个", "已经", "能够", "应该", "表示", "说明", "认为", "适用", "会遇",
    "至今", "凡是", "凡", "意思", "谁知道", "它们", "他们", "人们", "国人", "中国人", "喜爱", "喜欢", "彦语", "谚语", "成语越多", "格言式",
    "chinesische", "märchen", "sagen", "volksmärchen", "verlag", "wilhelm",
    "的", "了", "和", "与", "及", "或", "也", "都", "在", "对", "为", "以", "其", "这", "那", "中", "上", "下",
    "把", "被", "将", "会", "从", "到", "更", "等", "并", "而", "之", "者", "所", "使", "让", "于", "乃", "又",
    "不是", "作为", "通过", "关于", "由于", "如果", "只是", "这样", "这些", "那些", "时候", "方面", "意义",
}
PREFACE_LOW_INFO_TERMS = {
    "部分", "大量", "内容", "类型", "工作", "时期", "地方", "地区", "国家", "人民", "人类", "社会", "生活", "时代",
    "范围", "全部", "整体", "一般", "普通", "基本", "主要", "重要", "相关", "直接", "间接", "可能", "现实", "状况",
    "无法", "不能", "不可", "常常", "往往", "总是", "未能", "显然", "尤其", "比较", "特殊", "具体", "抽象",
    "情况", "问题", "结果", "过程", "目的", "原因", "基础", "关系", "方面", "意义", "价值", "特点", "特征", "位置",
    "道路", "力量", "精神", "意识", "意志", "活动", "事件", "文章", "本册", "本书", "资料", "记录", "发展", "产生",
    "形成", "存在", "发现", "使用", "采用", "提供", "作出", "成为", "属于", "具有", "包括", "包含", "显示", "指出",
}
STOP_WORDS.update(PREFACE_LOW_INFO_TERMS)

MEANINGFUL_TERMS = [
    "翻译", "改写", "选编", "转述", "忠实", "传播", "出版", "接受", "采录", "整理", "讲述", "叙述",
    "中国形象", "东方", "古老", "神秘", "智慧", "异域", "文化", "文学", "民族", "宗教", "伦理",
    "民间故事", "童话故事", "动物故事", "民间文学", "口头文学", "民族文学", "故事集",
    "寓言", "传说", "神话", "神怪", "读者", "儿童", "教育", "道德", "少数民族",
    "汉学", "译者", "编者", "出版社", "交流", "理解", "想象", "知识", "来源", "口头", "传统",
]
PREFACE_PROJECT_TERMS = [
    "德译中国故事集", "序跋", "译者", "卫礼贤", "中国民间童话", "民间故事", "中国故事", "民间文学",
    "文化传播", "跨文化传播", "翻译策略", "出版机构", "译介活动", "故事母题", "伦理观念", "神话传说",
    "地域文化", "中国形象", "德国汉学", "汉学研究", "文化中介", "翻译出版", "文学翻译", "民俗学",
    "民间叙事", "口头传统", "民族文学", "儿童文学", "童话故事", "动物故事", "宗教伦理", "道德教化",
    "少数民族", "故事集", "出版社", "出版史", "接受史", "改写策略", "编译策略", "跨文化交流",
]

VERB_HINTS = ["翻译", "传播", "出版", "改写", "转述", "采录", "整理", "讲述", "叙述", "接受", "理解", "交流", "影响", "建构", "呈现", "保留"]
ADJ_HINTS = ["古老", "神秘", "重要", "美妙", "深远", "丰富", "奇幻", "客观", "忠实", "广泛", "独特", "典型"]

PREFACE_ALLOWED_POS = {"n", "nr", "ns", "nt", "nz", "vn"}
PREFACE_BLOCKED_POS_PREFIXES = ("r", "d", "p", "c", "u", "y")
PREFACE_FUNCTION_STOPS = {
    "的", "了", "着", "过", "是", "在", "有", "和", "与", "及", "或", "而", "并", "也", "都", "就", "才", "乃", "其", "之", "者", "所",
    "于", "从", "到", "对", "为", "以", "因", "由", "但", "被", "把", "将", "让", "使", "给", "这", "那", "他", "她", "它", "们",
}
PREFACE_NOISE_PARTS = {
    "本书", "书中", "文本", "版本", "作者", "未记录", "这些", "那些", "他们", "我们", "一方面", "另一方面",
    "也就是", "所谓", "因此", "但是", "然后", "因为", "所以", "此外", "并且", "或者", "时候", "方面",
    "不同", "出现", "作品", "方式", "形式", "一部分", "一个", "一种", "一些", "几个", "许多", "很多",
    "多少", "数量", "第一", "第二", "之一", "各个", "某些", "什么", "几乎", "仍然", "于是", "这个", "那个",
    "已经", "能够", "应该", "表示", "说明", "认为", "至今", "凡是", "凡", "意思", "谁知道", "它们", "人们", "国人", "人喜爱", "人喜爱彦",
    "爱彦语", "喜欢彦", "言式表", "式表达", "随处", "凡意思", "几乎随", "人们几乎",
}
PREFACE_BAD_FRAGMENTS = {
    "的中国", "中国的", "数民族", "故事的", "故事中", "一部书", "什么意", "凡意思", "几乎随", "人们几乎",
    "语越多", "成语越多", "格言式表", "言式表", "式表达", "们几乎", "人喜爱彦", "国人喜爱", "爱彦语", "至今仍",
}
PREFACE_QUANTIFIER_RE = re.compile(r"^[一二三四五六七八九十百千万两几数多许若某各每半]+(个|部|种|些|点|位|篇|本|条|件|次|处|方面|部分|类|层|批|群|段)$")
PREFACE_TOPIC_GROUPS = [
    {"id": "translation", "title": "翻译出版", "seeds": ["翻译", "改写", "选编", "转述", "译者", "编者", "出版", "传播", "接受", "介绍", "出版社", "德译"]},
    {"id": "genre", "title": "民间叙事", "seeds": ["民间故事", "童话故事", "动物故事", "寓言", "传说", "神话", "神怪", "故事集", "口头文学", "民间文学"]},
    {"id": "china-image", "title": "中国形象", "seeds": ["中国形象", "东方", "古老", "神秘", "智慧", "风俗", "民族", "少数民族", "蒙古族", "维吾尔族"]},
    {"id": "ethics-religion", "title": "宗教伦理", "seeds": ["宗教", "信仰", "伦理", "道德", "善恶", "孝道", "教化", "神仙", "佛教", "道教"]},
    {"id": "reader-education", "title": "读者教育", "seeds": ["读者", "儿童", "教育", "学习", "阅读", "青年", "家庭", "知识"]},
    {"id": "source-method", "title": "来源方法", "seeds": ["来源", "采录", "整理", "材料", "版本", "注释", "分类", "研究", "题材"]},
    {"id": "culture", "title": "文化中介", "seeds": ["文化", "文学", "汉学", "交流", "理解", "德国", "欧洲", "世界", "民族文学"]},
    {"id": "orality", "title": "口头传统", "seeds": ["口头", "传统", "讲述", "叙述", "讲述者", "语言", "民歌", "歌谣"]},
]
PREFACE_DOMAIN_TERMS = sorted(
    set(MEANINGFUL_TERMS + PREFACE_PROJECT_TERMS + [seed for group in PREFACE_TOPIC_GROUPS for seed in group["seeds"]]),
    key=len,
    reverse=True,
)
PREFACE_DOMAIN_TERM_SET = set(PREFACE_DOMAIN_TERMS)
_PREFACE_JIEBA_READY = False


def ensure_preface_jieba_words() -> None:
    global _PREFACE_JIEBA_READY
    if not jieba or _PREFACE_JIEBA_READY:
        return
    for term in PREFACE_DOMAIN_TERMS:
        try:
            jieba.add_word(term, freq=2400, tag="n")
        except Exception:
            pass
    _PREFACE_JIEBA_READY = True


def _story_data_cache_key() -> tuple[int, int]:
    stat = STORY_DATA_PATH.stat()
    return stat.st_mtime_ns, stat.st_size


@lru_cache(maxsize=8)
def _story_data_cached(mtime_ns: int, size: int) -> dict[str, Any]:
    with STORY_DATA_PATH.open("r", encoding="utf-8") as file:
        return json.load(file)


def story_data() -> dict[str, Any]:
    return _story_data_cached(*_story_data_cache_key())


def _xlsx_col_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", str(cell_ref or ""))
    if not match:
        return 0
    value = 0
    for char in match.group(1):
        value = value * 26 + (ord(char) - 64)
    return max(0, value - 1)


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    values: list[str] = []
    for item in root.findall("a:si", ns):
        values.append("".join(node.text or "" for node in item.findall(".//a:t", ns)))
    return values


def _xlsx_sheet_targets(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    ns_main = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    ns_rel = {"a": "http://schemas.openxmlformats.org/package/2006/relationships"}
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    target_by_id = {
        rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
        for rel in rels.findall("a:Relationship", ns_rel)
    }
    sheets: list[tuple[str, str]] = []
    for sheet in workbook.findall("a:sheets/a:sheet", ns_main):
        name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
        target = target_by_id.get(rel_id, "")
        if name and target:
            sheets.append((name, f"xl/{target.lstrip('/')}"))
    return sheets


def _xlsx_sheet_rows(path: Path, sheet_name: str) -> list[list[str]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _xlsx_shared_strings(archive)
        sheets = dict(_xlsx_sheet_targets(archive))
        target = sheets.get(sheet_name)
        if not target:
            return []
        root = ET.fromstring(archive.read(target))
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rows: list[list[str]] = []
    for row in root.findall(".//a:sheetData/a:row", ns):
        values: list[str] = []
        for cell in row.findall("a:c", ns):
            index = _xlsx_col_index(cell.attrib.get("r", ""))
            while len(values) <= index:
                values.append("")
            cell_type = cell.attrib.get("t", "")
            if cell_type == "inlineStr":
                text = "".join(node.text or "" for node in cell.findall(".//a:t", ns))
            else:
                raw = cell.findtext("a:v", default="", namespaces=ns)
                if cell_type == "s" and raw.isdigit():
                    shared_index = int(raw)
                    text = shared_strings[shared_index] if 0 <= shared_index < len(shared_strings) else ""
                else:
                    text = raw or ""
            values[index] = text.strip()
        if any(value != "" for value in values):
            rows.append(values)
    return rows


def _workbook_sheet_names(path: Path) -> list[str]:
    if load_workbook is not None:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            return list(workbook.sheetnames)
        except Exception:
            pass
    try:
        with zipfile.ZipFile(path) as archive:
            return [name for name, _ in _xlsx_sheet_targets(archive)]
    except Exception:
        return []


def _read_workbook_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    rows: list[list[Any]] = []
    if load_workbook is not None:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                return []
            rows = [
                ["" if cell is None else str(cell).strip() for cell in row]
                for row in workbook[sheet_name].iter_rows(values_only=True)
            ]
        except Exception:
            rows = []
    if not rows:
        try:
            rows = _xlsx_sheet_rows(path, sheet_name)
        except Exception:
            rows = []
    if not rows:
        return []
    header = [str(cell or "").strip() for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        cells = ["" if cell is None else str(cell).strip() for cell in row]
        if not any(cells):
            continue
        payload = {header[index] or f"col_{index}": cells[index] for index in range(min(len(header), len(cells)))}
        payload["_workbook"] = path.name
        payload["_sheet"] = sheet_name
        records.append(payload)
    return records


@lru_cache(maxsize=1)
def workbook_knowledge() -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for filename in EXTERNAL_KNOWLEDGE_WORKBOOKS:
        for path in list(ROOT.rglob(filename))[:1]:
            for sheet_name in _workbook_sheet_names(path)[:8]:
                records.extend(_read_workbook_sheet(path, sheet_name))
    return records


@lru_cache(maxsize=1)
def publication_workbook_rows() -> list[dict[str, Any]]:
    matches = list(ROOT.rglob(PUBLICATION_MAP_WORKBOOK))
    if not matches:
        return []
    path = matches[0]
    rows: list[dict[str, Any]] = []
    for sheet_name in _workbook_sheet_names(path)[:4]:
        rows.extend(_read_workbook_sheet(path, sheet_name))
    return rows


def _publication_source_cache_key() -> tuple[int, int]:
    if not PUBLICATION_SOURCE_MAP_JSON_PATH.exists():
        return 0, 0
    stat = PUBLICATION_SOURCE_MAP_JSON_PATH.stat()
    return stat.st_mtime_ns, stat.st_size


@lru_cache(maxsize=4)
def _publication_source_map_cached(mtime_ns: int, size: int) -> dict[str, Any]:
    if not mtime_ns or not size:
        return {}
    try:
        return json.loads(PUBLICATION_SOURCE_MAP_JSON_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def publication_source_map_data() -> dict[str, Any]:
    return _publication_source_map_cached(*_publication_source_cache_key())


def _record_year(item: dict[str, Any]) -> int:
    match = re.search(r"\d{4}", str(item.get("year") or item.get("yearText") or ""))
    return int(match.group(0)) if match else 0


def _record_title(item: dict[str, Any]) -> str:
    parts = [
        str(item.get("title") or item.get("title(Chinese)") or item.get("chineseTitle") or "").strip(),
        str(item.get("foreignTitle") or item.get("name") or "").strip(),
    ]
    return " / ".join(part for part in parts if part) or "未记录"


def short(text: Any, limit: int = 16) -> str:
    value = str(text or "未记录")
    return value if len(value) <= limit else value[: limit - 1] + "…"


def classify_role(role: str = "", editor: str = "") -> str:
    text = f"{role} {editor}"
    if "传教士" in text:
        return "传教士"
    if "民俗" in text:
        return "民俗学家"
    if any(mark in text for mark in ["汉学", "藏学", "教授", "学者"]):
        return "汉学家"
    if any(mark in text for mark in ["出版", "出版社", "出版商"]):
        return "出版机构"
    if any(mark in text for mark in ["作家", "改写", "记者", "文学"]):
        return "作家/改写者"
    if any(mark in text for mark in ["中国", "中德", "北外", "教师", "译者"]):
        return "中国译者/中德合作"
    return "学术编选者"


def stage_for(year: Any) -> dict[str, Any]:
    value = int(year or 0)
    for stage in STAGES:
        if value >= stage["range"][0] and value < stage["range"][1]:
            return stage
    return STAGES[0]


def clean_cities(value: str = "") -> list[str]:
    text = str(value or "").replace("Zurich", "Zürich")
    matched = [city for city in CITY_COORDS if city in text]
    if matched:
        return matched
    return [item.strip() for item in re.split(r"/|,|，| und | u\.a\.|;|；", text) if item.strip()]


def city_coords(value: Any, default: Optional[list[float]] = None) -> list[float]:
    coords = [CITY_COORDS[city] for city in clean_cities(str(value or "")) if city in CITY_COORDS]
    if coords:
        lon = sum(float(item[0]) for item in coords) / len(coords)
        lat = sum(float(item[1]) for item in coords) / len(coords)
        return [round(lon, 4), round(lat, 4)]
    return list(default or CITY_COORDS["Berlin"])


def token_pos(token: str) -> str:
    if any(hint in token for hint in VERB_HINTS):
        return "动词"
    if any(hint in token for hint in ADJ_HINTS):
        return "形容词"
    return "名词"


def clean_preface_text(text: Any) -> str:
    value = str(text or "")
    value = re.sub(r"https?://\S+|www\.\S+", " ", value, flags=re.I)
    value = re.sub(r"[\u0000-\u001f]+", " ", value)
    value = re.sub(r"[^\u4e00-\u9fffA-Za-zÄÖÜäöüß0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_preface_token(term: Any) -> str:
    punctuation = "\uFF0C\u3002\u3001\uFF1B\uFF1A:,.!?\uFF01\uFF1F()\uFF08\uFF09[]\u3010\u3011\u300A\u300B\u2018\u2019\u201C\u201D\"'“”‘’"
    text = re.sub(r"\s+", "", str(term or "")).strip(punctuation)
    return text


PREFACE_ALLOWED_STORY_TERMS = {
    "中国故事", "民间故事", "童话故事", "动物故事", "神话故事", "寓言故事", "志怪故事", "机智故事", "故事集",
}
PREFACE_BROKEN_STORY_PREFIXES = {
    "间", "些", "话", "个", "则", "类", "种", "篇", "部", "本", "段", "中", "的", "其", "这", "那", "些",
}


def is_broken_preface_fragment(value: str) -> bool:
    if value in PREFACE_BAD_FRAGMENTS:
        return True
    if value in PREFACE_ALLOWED_STORY_TERMS or value in PREFACE_DOMAIN_TERM_SET:
        return False
    if value.endswith("故事") and (len(value) <= 4 or value[:1] in PREFACE_BROKEN_STORY_PREFIXES):
        return True
    if value.endswith(("童话", "民间", "民族", "文学", "传统")) and len(value) <= 3:
        return True
    return False


def domain_term_score(term: str) -> float:
    value = normalize_preface_token(term)
    if value in PREFACE_DOMAIN_TERM_SET:
        return 1.0
    if any(value in domain or domain in value for domain in PREFACE_DOMAIN_TERMS if len(value) >= 2):
        return 0.55
    return 0.0


def preface_pos_score(term: str) -> float:
    if normalize_preface_token(term) in PREFACE_DOMAIN_TERM_SET:
        return 1.0
    return 0.92 if token_pos(term) == "名词" else 0.82


def valid_token(token: str) -> bool:
    value = normalize_preface_token(token)
    if len(value) < 2 or len(value) > 12:
        return False
    if value.lower() in STOP_WORDS or value in STOP_WORDS:
        return False
    if value in PREFACE_LOW_INFO_TERMS:
        return False
    if is_broken_preface_fragment(value) or PREFACE_QUANTIFIER_RE.fullmatch(value):
        return False
    if value in PREFACE_FUNCTION_STOPS or any(part in value for part in PREFACE_NOISE_PARTS):
        return False
    if re.fullmatch(r"\d+|[a-zA-ZÄÖÜäöüß]{1,3}", value):
        return False
    if any(mark in value for mark in ["因此", "但是", "然后", "因为", "所以", "此外", "并且", "或者", "的中国"]):
        return False
    if value.endswith("数民族") and value != "少数民族":
        return False
    if re.search(r"[的了着过是有和与及或而并也都其之者所于从到对为以把被将让使给]", value):
        if value not in MEANINGFUL_TERMS and not any(value in group["seeds"] for group in PREFACE_TOPIC_GROUPS):
            return False
    if any(value.startswith(stop) or value.endswith(stop) for stop in PREFACE_FUNCTION_STOPS):
        return False
    return bool(re.search(r"[\u4e00-\u9fffA-Za-zÄÖÜäöüß]", value))


def is_allowed_preface_pos(pos: str) -> bool:
    value = str(pos or "").lower()
    if not value or any(value.startswith(prefix) for prefix in PREFACE_BLOCKED_POS_PREFIXES):
        return False
    return value in PREFACE_ALLOWED_POS or any(value.startswith(prefix) for prefix in ("nr", "ns", "nt", "nz"))


def longest_domain_terms_in_text(source: str) -> list[str]:
    matches: list[tuple[int, int, str]] = []
    occupied: list[tuple[int, int]] = []
    for term in PREFACE_DOMAIN_TERMS:
        clean = normalize_preface_token(term)
        if not valid_token(clean):
            continue
        for match in re.finditer(re.escape(clean), source, flags=re.I):
            start, end = match.span()
            if any(not (end <= used_start or start >= used_end) for used_start, used_end in occupied):
                continue
            occupied.append((start, end))
            matches.append((start, end, clean))
    matches.sort(key=lambda item: item[0])
    return [term for _, _, term in matches]


def merge_domain_phrases(tokens: list[str]) -> list[str]:
    merged: list[str] = []
    index = 0
    while index < len(tokens):
        best_term = ""
        best_size = 0
        max_size = min(6, len(tokens) - index)
        for size in range(max_size, 1, -1):
            candidate = normalize_preface_token("".join(tokens[index : index + size]))
            if candidate in PREFACE_DOMAIN_TERM_SET and valid_token(candidate):
                best_term = candidate
                best_size = size
                break
        if best_term:
            merged.append(best_term)
            index += best_size
        else:
            merged.append(tokens[index])
            index += 1
    return merged


def dedupe_domain_fallback_tokens(tokens: list[str], fallback_terms: list[str]) -> list[str]:
    result = list(tokens)
    remaining = Counter(term for term in tokens if term in PREFACE_DOMAIN_TERM_SET)
    for term in fallback_terms:
        if remaining.get(term, 0) > 0:
            remaining[term] -= 1
        elif valid_token(term):
            result.append(term)
    return result


def preface_topic_for(term: str, pair_counts: Counter | None = None) -> str:
    value = normalize_preface_token(term)
    best_id = ""
    best_score = 0.0
    for group in PREFACE_TOPIC_GROUPS:
        score = 0.0
        for seed in group["seeds"]:
            if value == seed:
                score += 9.0
            elif seed in value or value in seed:
                score += 4.0
            if pair_counts:
                key = "|".join(sorted([value, seed]))
                score += min(3.0, float(pair_counts.get(key, 0)))
        if score > best_score:
            best_id = group["id"]
            best_score = score
    if best_id:
        return best_id
    if token_pos(value) == "动词":
        return "translation"
    if any(mark in value for mark in ["故事", "童话", "传说", "神话", "寓言", "神怪"]):
        return "genre"
    if any(mark in value for mark in ["中国", "民族", "东方", "蒙古", "维吾尔", "满族"]):
        return "china-image"
    if any(mark in value for mark in ["材料", "版本", "来源", "分类", "研究", "题材"]):
        return "source-method"
    return "culture"


def tokenize_text(text: str) -> list[str]:
    return list(_tokenize_text_cached(str(text or "")))


@lru_cache(maxsize=512)
def _tokenize_text_cached(source: str) -> tuple[str, ...]:
    source = clean_preface_text(source)
    if not source:
        return tuple()
    ensure_preface_jieba_words()
    domain_fallback = longest_domain_terms_in_text(source)
    tokens: list[str] = []
    if jieba:
        segmented: list[str] = []
        for word, flag in pseg.cut(source) if pseg else [(word, "n") for word in jieba.cut(source)]:
            clean = normalize_preface_token(word)
            pos = str(flag or "").lower()
            if not valid_token(clean):
                continue
            if clean in PREFACE_DOMAIN_TERM_SET or is_allowed_preface_pos(pos):
                segmented.append(clean)
        tokens = merge_domain_phrases(segmented)
        tokens = [token for token in tokens if valid_token(token)]
        tokens = dedupe_domain_fallback_tokens(tokens, domain_fallback)
        return tuple(tokens)
    for word in re.findall(r"[A-Za-zÄÖÜäöüß]{3,}", source):
        clean = normalize_preface_token(word)
        if valid_token(clean):
            tokens.append(clean)
    tokens.extend(term for term in domain_fallback if valid_token(term))
    return tuple(dict.fromkeys(tokens))


def top_terms_for_text(text: str, limit: int = 28) -> list[str]:
    return [item["text"] for item in preface_keyword_weights_from_texts([text], limit)]


def _normalized_preface_counts(counts: Counter) -> Counter:
    normalized: Counter = Counter()
    for term, count in counts.items():
        clean = normalize_preface_token(term)
        if valid_token(clean):
            normalized[clean] += count
    return normalized


def _normalize_score_map(scores: dict[str, float] | Counter) -> dict[str, float]:
    if not scores:
        return {}
    max_score = max(float(value) for value in scores.values()) or 1.0
    return {term: float(value) / max_score for term, value in scores.items()}


def _preface_domain_bonus(term: str) -> float:
    return 1.0 + (domain_term_score(term) * 0.18) + (preface_pos_score(term) * 0.08)


def _select_preface_weighted_terms(
    scores: dict[str, float] | Counter,
    total_counts: Counter,
    limit: int,
    tfidf_scores: dict[str, float] | Counter | None = None,
    textrank_scores: dict[str, float] | Counter | None = None,
    doc_counts: Counter | None = None,
    min_count: int = 0,
) -> list[dict[str, Any]]:
    cleaned_scores: dict[str, float] = {}
    cleaned_counts = _normalized_preface_counts(total_counts)
    for term, score in scores.items():
        clean = normalize_preface_token(term)
        if valid_token(clean):
            cleaned_scores[clean] = max(cleaned_scores.get(clean, 0.0), float(score))
    ranked = sorted(
        cleaned_scores.items(),
        key=lambda item: (item[1], cleaned_counts.get(item[0], 0) * min(len(item[0]), 6), len(item[0])),
        reverse=True,
    )
    selected: list[tuple[str, float]] = []
    for term, score in ranked:
        if any(term == existing for existing, _ in selected):
            continue
        if any(term in existing and (len(term) <= 2 or domain_term_score(existing) >= domain_term_score(term)) and score <= existing_score * 1.12 for existing, existing_score in selected):
            continue
        selected = [
            (existing, existing_score)
            for existing, existing_score in selected
            if not (existing in term and (len(existing) <= 2 or domain_term_score(term) >= domain_term_score(existing)) and existing_score <= score * 1.12)
        ]
        selected.append((term, score))
        if len(selected) >= limit:
            break
    if min_count > 0:
        selected_terms = {term for term, _ in selected}
        frequent_terms = sorted(
            (term for term, count in cleaned_counts.items() if count >= min_count and term not in selected_terms),
            key=lambda term: (cleaned_counts[term], cleaned_scores.get(term, 0.0), len(term)),
            reverse=True,
        )
        for term in frequent_terms:
            selected.append((term, cleaned_scores.get(term, float(cleaned_counts[term]))))
            selected_terms.add(term)
    max_selected = max((score for _, score in selected), default=1.0) or 1.0
    tfidf_norm = _normalize_score_map(tfidf_scores or {})
    textrank_norm = _normalize_score_map(textrank_scores or {})
    return [
        {
            "text": term,
            "value": round((score / max_selected) * 100, 3),
            "weight": round(score, 6),
            "count": int(cleaned_counts.get(term, 0)),
            "docCount": int((doc_counts or Counter()).get(term, 0)),
            "tfidf": round(tfidf_norm.get(term, 0.0), 6),
            "textrank": round(textrank_norm.get(term, 0.0), 6),
            "pos": token_pos(term),
            "topic": preface_topic_for(term),
        }
        for term, score in selected
    ]


def _preface_textrank_scores(token_docs: list[list[str]], window: int = 4, iterations: int = 30) -> Counter:
    graph: dict[str, Counter] = defaultdict(Counter)
    for tokens in token_docs:
        sequence = [normalize_preface_token(token) for token in tokens if valid_token(token)]
        for index, source in enumerate(sequence):
            for distance, target in enumerate(sequence[index + 1 : index + window], start=1):
                if source == target:
                    continue
                weight = 1.0 / distance
                graph[source][target] += weight
                graph[target][source] += weight
    terms = set(graph)
    for neighbors in graph.values():
        terms.update(neighbors)
    if not terms:
        return Counter()
    scores = {term: 1.0 for term in terms}
    damping = 0.85
    for _ in range(iterations):
        next_scores = {term: 1.0 - damping for term in terms}
        for source, neighbors in graph.items():
            total_weight = sum(neighbors.values()) or 1.0
            contribution = scores.get(source, 1.0) / total_weight
            for target, weight in neighbors.items():
                next_scores[target] += damping * contribution * weight
        scores = next_scores
    return Counter(scores)


def preface_keyword_weights_from_texts(texts: list[str], limit: int = 120, min_count: int = 0) -> list[dict[str, Any]]:
    token_docs: list[list[str]] = []
    doc_counts: list[Counter] = []
    total_counts: Counter = Counter()
    for text in texts:
        tokens = [normalize_preface_token(token) for token in tokenize_text(text) if valid_token(token)]
        if not tokens:
            continue
        counts = Counter(tokens)
        token_docs.append(tokens)
        doc_counts.append(counts)
        total_counts.update(counts)
    if not doc_counts:
        return []
    document_frequency: Counter = Counter()
    for counts in doc_counts:
        document_frequency.update(counts.keys())
    document_count = len(doc_counts)
    tfidf_scores: Counter = Counter()
    doc_coverage: Counter = Counter()
    for counts in doc_counts:
        max_count = max(counts.values()) or 1
        for term, count in counts.items():
            idf = math.log((1 + document_count) / (1 + document_frequency[term])) + 1.0
            tf = math.log1p(count) / math.log1p(max_count)
            tfidf_scores[term] += tf * idf
            doc_coverage[term] += 1
    textrank_scores = _preface_textrank_scores(token_docs)
    tfidf_norm = _normalize_score_map(tfidf_scores)
    textrank_norm = _normalize_score_map(textrank_scores)
    fused_scores: dict[str, float] = {}
    for term in set(total_counts) | set(tfidf_scores) | set(textrank_scores):
        if not valid_token(term):
            continue
        fused = (
            0.45 * tfidf_norm.get(term, 0.0)
            + 0.30 * textrank_norm.get(term, 0.0)
            + 0.15 * domain_term_score(term)
            + 0.10 * preface_pos_score(term)
        )
        fused_scores[term] = fused
    return _select_preface_weighted_terms(fused_scores, total_counts, limit, tfidf_scores, textrank_scores, doc_coverage, min_count=min_count)


def preface_keyword_weights_from_entries(entries: list[dict[str, Any]], limit: int = 120, min_count: int = 0) -> list[dict[str, Any]]:
    return preface_keyword_weights_from_texts([str(entry.get("text") or "") for entry in entries], limit, min_count=min_count)


def ranked_preface_words(counts: Counter, limit: int = 120) -> list[dict[str, Any]]:
    normalized = _normalized_preface_counts(counts)
    scores = {term: count * min(len(term), 6) * _preface_domain_bonus(term) for term, count in normalized.items()}
    return _select_preface_weighted_terms(scores, normalized, limit, scores, {}, Counter({term: 1 for term in normalized}))


def preface_entries_from_collections(collections: list[dict[str, Any]]) -> list[dict[str, Any]]:
    entries = []
    for item in collections:
        text = " ".join(
            str(part or "")
            for part in [item.get("prefaceText"), item.get("prefaceIntro"), item.get("name"), item.get("foreignTitle"), item.get("prefaceAuthor")]
        )
        if text.strip():
            entries.append({"id": item.get("id"), "label": item.get("name"), "year": int(item.get("year") or 0), "text": text})
    return entries


def topic_cluster_from_entries(entries: list[dict[str, Any]]) -> dict[str, Any]:
    term_counts: Counter = Counter()
    term_weights: Counter = Counter()
    term_years: dict[str, set[int]] = defaultdict(set)
    term_docs: dict[str, set[str]] = defaultdict(set)
    pair_counts: Counter = Counter()
    for entry in entries:
        keywords = preface_keyword_weights_from_texts([entry.get("text", "")], 56)
        terms: list[str] = []
        term_seen: set[str] = set()
        for keyword in keywords:
            term = normalize_preface_token(keyword.get("text", ""))
            if not valid_token(term) or term in term_seen:
                continue
            term_seen.add(term)
            terms.append(term)
            term_counts[term] += max(1, int(round(float(keyword.get("count") or 1))))
            term_weights[term] += max(float(keyword.get("value") or 0), float(keyword.get("weight") or 0), 1.0)
            if entry.get("year"):
                term_years[term].add(int(entry["year"]))
            term_docs[term].add(str(entry.get("id") or entry.get("label") or "entry"))
        ranked_terms = terms[:34]
        for index, source in enumerate(ranked_terms):
            for offset, target in enumerate(ranked_terms[index + 1 : index + 8], start=1):
                source_topic = preface_topic_for(source)
                target_topic = preface_topic_for(target)
                if source_topic != target_topic and offset > 3:
                    continue
                source_weight = max(0.3, min(1.0, term_weights.get(source, 1.0) / 100))
                target_weight = max(0.3, min(1.0, term_weights.get(target, 1.0) / 100))
                weight = (1.7 if source_topic == target_topic else 0.48) * ((source_weight + target_weight) / 2)
                pair_counts["|".join(sorted([source, target]))] += weight

    global_keywords = preface_keyword_weights_from_entries(entries, 110)
    top_terms: list[str] = []
    for keyword in global_keywords:
        term = normalize_preface_token(keyword.get("text", ""))
        if not valid_token(term):
            continue
        if term not in term_counts:
            term_counts[term] += max(1, int(round(float(keyword.get("count") or 1))))
            term_weights[term] += max(float(keyword.get("value") or 0), float(keyword.get("weight") or 0), 1.0)
        if any(term in longer and (len(term) <= 2 or domain_term_score(longer) >= domain_term_score(term)) for longer in top_terms):
            continue
        top_terms.append(term)
        if len(top_terms) >= 92:
            break
    if len(top_terms) < 60:
        ranked_terms = sorted(
            term_weights,
            key=lambda term: (term_weights[term], len(term_docs.get(term, set())), term_counts[term] * min(len(term), 5)),
            reverse=True,
        )
        for term in ranked_terms:
            if not valid_token(term) or term in top_terms:
                continue
            if any(term in longer and (len(term) <= 2 or domain_term_score(longer) >= domain_term_score(term)) for longer in top_terms):
                continue
            top_terms.append(term)
            if len(top_terms) >= 92:
                break

    labels = {term: preface_topic_for(term, pair_counts) for term in top_terms}
    clusters = []
    group_map = {group["id"]: group for group in PREFACE_TOPIC_GROUPS}
    for group in PREFACE_TOPIC_GROUPS:
        members = [term for term in top_terms if labels[term] == group["id"]]
        if not members:
            continue
        title_terms = sorted(members, key=lambda term: (term_counts[term], len(term_docs.get(term, set()))), reverse=True)[:3]
        clusters.append(
            {
                "id": group["id"],
                "title": group["title"] if len(members) >= 3 else " / ".join(title_terms[:2]),
                "terms": members[:14],
                "size": round(sum(term_weights[member] for member in members), 3),
            }
        )
    clusters = sorted(clusters, key=lambda item: item["size"], reverse=True)[:9]
    live_cluster_ids = {cluster["id"] for cluster in clusters}
    group_order = {cluster["id"]: index for index, cluster in enumerate(clusters)}

    nodes = [
        {
            "id": term,
            "label": term,
            "cluster": labels[term],
            "count": max(1, int(round(term_counts[term]))),
            "weight": round(term_weights[term], 3),
            "years": sorted(term_years.get(term, set())),
            "pos": token_pos(term),
            "docCount": len(term_docs.get(term, set())),
            "topic": group_map.get(labels[term], {}).get("title", "文化中介"),
        }
        for term in top_terms
        if labels[term] in live_cluster_ids
    ]
    nodes.sort(key=lambda node: (group_order.get(node["cluster"], 99), -node["count"], node["label"]))
    node_ids = {node["id"] for node in nodes}
    edge_candidates = []
    for key, weight in sorted(pair_counts.items(), key=lambda item: item[1], reverse=True):
        source, target = key.split("|")
        if source in node_ids and target in node_ids:
            same_cluster = labels.get(source) == labels.get(target)
            if not same_cluster and weight < 2.5:
                continue
            edge_years = sorted(set(term_years.get(source, set())) | set(term_years.get(target, set())))
            edge_candidates.append({"source": source, "target": target, "weight": round(weight, 3), "years": edge_years})
    degree: Counter = Counter()
    edges = []
    for edge in edge_candidates:
        if degree[edge["source"]] >= 5 or degree[edge["target"]] >= 5:
            continue
        edges.append(edge)
        degree[edge["source"]] += 1
        degree[edge["target"]] += 1
        if len(edges) >= 180:
            break
    return {
        "type": "prefaceCluster",
        "title": "序跋文本主题聚类图",
        "subtitle": "展示序跋文本中的主题群、关键词关系与文档覆盖。",
        "clusters": clusters,
        "nodes": nodes,
        "edges": edges,
        "years": sorted({int(entry.get("year") or 0) for entry in entries if entry.get("year")}),
    }


def preface_source(collections: list[dict[str, Any]]) -> str:
    return "\n".join(
        " ".join(
            str(part or "")
            for part in [item.get("name"), item.get("foreignTitle"), item.get("prefaceAuthor"), item.get("editorRole"), item.get("publisher"), item.get("prefaceText"), item.get("prefaceIntro")]
        )
        for item in collections
    )


def child_themes(child: dict[str, Any]) -> list[str]:
    text = " ".join(str(child.get(key) or "") for key in ["canonicalName", "variantName", "storyType", "notes", "ethnicity", "bookName"])
    terms = top_terms_for_text(text, 8)
    dynamic = [term for term in terms if valid_token(term) and len(term) <= 8]
    if not dynamic:
        title = str(child.get("canonicalName") or child.get("variantName") or "").strip()
        if title:
            dynamic = [short(title, 6)]
    return list(dict.fromkeys(dynamic))[:3]


def infer_child_story_type(child: dict[str, Any], themes: list[str] | None = None) -> str:
    explicit = str(child.get("storyType") or "").strip()
    if explicit:
        return explicit
    text = " ".join(str(child.get(key) or "") for key in ["canonicalName", "variantName", "notes", "bookName"])
    text += " " + " ".join(themes or [])
    if any(mark in text for mark in ["鬼", "尸", "妖", "怪", "狐", "狐仙", "精"]):
        return "志怪故事"
    if any(mark in text for mark in ["龙", "雷", "神", "仙", "天", "月", "太阳", "盘古", "女娲", "嫦娥", "织女"]):
        return "神话传说"
    if any(mark in text for mark in ["佛", "观音", "菩萨", "和尚", "道士", "老子", "庄子", "寺", "庙"]):
        return "宗教传说"
    if any(mark in text for mark in ["虎", "蛇", "牛", "马", "鸟", "鱼", "猴", "兔", "犬", "狗", "猫", "羊", "虫", "蚌", "龟"]):
        return "动物寓言"
    if any(mark in text for mark in ["公主", "娘娘", "妻", "女", "嫁", "婚", "牛郎", "织女", "母", "父", "子"]):
        return "婚恋家庭"
    if any(mark in text for mark in ["王", "皇帝", "将", "关羽", "官", "战", "兵", "臣", "唐", "秦", "汉"]):
        return "历史传说"
    if any(mark in text for mark in ["智", "计", "骗", "谎", "审", "案", "盗", "疑", "辨", "救"]):
        return "机智故事"
    if any(mark in text for mark in ["孝", "忠", "义", "善", "恶", "报恩", "惩罚", "节", "廉", "德"]):
        return "伦理寓言"
    return child.get("ethnicity") or child.get("carrier") or "民间传说"


def identity_process(collections: list[dict[str, Any]]) -> dict[str, Any]:
    roles = ["传教士", "汉学家", "民俗学家", "作家/改写者", "学术编选者", "出版机构", "中国译者/中德合作"]
    columns = []
    for stage in STAGES:
        rows = [item for item in collections if stage_for(item.get("year"))["id"] == stage["id"]]
        top = Counter(classify_role(item.get("editorRole", ""), item.get("editor", "")) for item in rows).most_common(3)
        columns.append({**stage, "roles": [{"name": name, "count": count} for name, count in top]})
    years = [int(item.get("year") or 0) for item in collections if int(item.get("year") or 0)]
    first_decade = (min(years) // 10 * 10) if years else 1910
    last_decade = (max(years) // 10 * 10) if years else 2020
    decade_values = list(range(first_decade, last_decade + 10, 10))
    decades = []
    for decade in decade_values:
        rows = [item for item in collections if decade <= int(item.get("year") or 0) < decade + 10]
        counts = Counter(classify_role(item.get("editorRole", ""), item.get("editor", "")) for item in rows)
        total = sum(counts.values())
        top_role, top_count = counts.most_common(1)[0] if counts else ("未记录", 0)
        shares = [counts.get(role, 0) / total for role in roles if total and counts.get(role, 0)]
        diversity = (-sum(share * math.log(share) for share in shares) / math.log(len(roles))) if shares and len(roles) > 1 else 0
        concentration = sum(share * share for share in shares) if shares else 0
        decades.append(
            {
                "id": f"{decade}s",
                "label": f"{decade}s",
                "decade": decade,
                "range": [decade, decade + 9],
                "total": total,
                "topRole": top_role,
                "topCount": top_count,
                "activeRoles": len(shares),
                "diversity": round(diversity, 4),
                "concentration": round(concentration, 4),
                "roles": [
                    {
                        "role": role,
                        "count": int(counts.get(role, 0)),
                        "share": round((counts.get(role, 0) / total) if total else 0, 4),
                    }
                    for role in roles
                ],
            }
        )
    role_totals = Counter()
    role_peaks: dict[str, dict[str, Any]] = {}
    for decade in decades:
        for row in decade["roles"]:
            role_totals[row["role"]] += row["count"]
            peak = role_peaks.get(row["role"], {"count": -1})
            if row["count"] > peak["count"]:
                role_peaks[row["role"]] = {"decade": decade["id"], "count": row["count"]}
    transitions = []
    for index in range(len(decades) - 1):
        current = decades[index]
        next_decade = decades[index + 1]
        for role in roles:
            source_count = next((item["count"] for item in current["roles"] if item["role"] == role), 0)
            target_count = next((item["count"] for item in next_decade["roles"] if item["role"] == role), 0)
            if source_count or target_count:
                transitions.append(
                    {
                        "sourceDecade": current["id"],
                        "targetDecade": next_decade["id"],
                        "sourceRole": role,
                        "targetRole": role,
                        "weight": int(max(source_count, target_count)),
                        "delta": int(target_count - source_count),
                    }
                )
    return {
        "type": "identityProcess",
        "title": "德译中国故事集译介主体结构演化图",
        "subtitle": "按十年尺度统计译介主体身份，以热度矩阵、占比曲线和相邻年代流动展示主体结构变迁。",
        "columns": columns,
        "roles": [
            {
                "name": role,
                "total": int(role_totals.get(role, 0)),
                "share": round((role_totals.get(role, 0) / sum(role_totals.values())) if sum(role_totals.values()) else 0, 4),
                "peakDecade": role_peaks.get(role, {}).get("decade", ""),
                "peakCount": int(role_peaks.get(role, {}).get("count", 0)),
            }
            for role in roles
        ],
        "decades": decades,
        "transitions": transitions,
        "steps": ["整理译者/编者信息", "标注身份类别", "按年代统计变化", "分析传播主体转移"],
    }


def identity_river(collections: list[dict[str, Any]]) -> dict[str, Any]:
    roles = ["传教士", "汉学家", "民俗学家", "作家/改写者", "学术编选者", "出版机构", "中国译者/中德合作"]
    years = [int(item.get("year") or 0) for item in collections if int(item.get("year") or 0)]
    first_decade = (min(years) // 10 * 10) if years else 1910
    last_decade = (max(years) // 10 * 10) if years else 2020
    decades = [
        {
            "id": f"{decade}s",
            "title": f"{decade}s",
            "label": f"{decade}s",
            "decade": decade,
            "note": f"{decade}-{decade + 9}",
            "range": [decade, decade + 10],
        }
        for decade in range(first_decade, last_decade + 10, 10)
    ]
    series = []
    for role in roles:
        values = []
        for decade in decades:
            count = sum(
                1
                for item in collections
                if decade["range"][0] <= int(item.get("year") or 0) < decade["range"][1]
                and classify_role(item.get("editorRole", ""), item.get("editor", "")) == role
            )
            values.append(count)
        series.append({"role": role, "values": values})
    return {
        "type": "identityRiver",
        "title": "译者身份流变图：时间河流",
        "subtitle": "横轴按十年尺度展开；线带越宽，表示该身份在该年代参与度越高。",
        "stages": decades,
        "decades": decades,
        "series": series,
    }


def publication_map(collections: list[dict[str, Any]], title: str = "德译中国故事集出版地图") -> dict[str, Any]:
    cities: dict[str, dict[str, Any]] = {}
    china_cities = {"Peking", "Beijing", "北京", "Shanghai", "上海"}
    for item in collections:
        for city in clean_cities(item.get("city") or item.get("publisher", "")):
            if city not in CITY_COORDS:
                continue
            year = _record_year(item)
            title_text = _record_title(item)
            publisher_text = str(item.get("publisher") or "").strip()
            record = cities.setdefault(
                city,
                {
                    "id": city,
                    "city": city,
                    "label": CITY_ZH.get(city, city),
                    "coords": CITY_COORDS[city],
                    "count": 0,
                    "years": [],
                    "yearCounts": {},
                    "works": [],
                    "publishers": set(),
                    "country": "中国" if city in china_cities else item.get("country") or "德国/德语区",
                },
            )
            record["count"] += 1
            if year:
                record["years"].append(year)
                record["yearCounts"][year] = int(record["yearCounts"].get(year, 0)) + 1
            if publisher_text:
                record["publishers"].add(publisher_text)
            record["works"].append({"title": title_text, "year": year, "publisher": publisher_text})
    points = []
    for item in sorted(cities.values(), key=lambda row: row["count"], reverse=True):
        seen_works: set[tuple[str, int, str]] = set()
        works = []
        for work in sorted(item["works"], key=lambda row: ((row.get("year") or 9999), str(row.get("title") or ""))):
            key = (str(work.get("title") or ""), int(work.get("year") or 0), str(work.get("publisher") or ""))
            if key in seen_works:
                continue
            seen_works.add(key)
            works.append({"title": key[0] or "未记录", "year": key[1], "publisher": key[2]})
        years = sorted({int(year) for year in item["years"] if year})
        points.append({
            **item,
            "years": years,
            "yearCounts": {str(year): int(item["yearCounts"].get(year, 0)) for year in years},
            "publishers": sorted(item["publishers"])[:8],
            "works": works,
        })
    return {
        "type": "publicationMap",
        "title": title,
        "subtitle": "圆点越大表示出版城市越活跃；真实地图显示德国、德语区与中国出版节点。",
        "geo": {
            "world": "/api/basemap/boundary",
            "china": "/api/basemap/province",
            "countries": ["Germany", "China", "Switzerland", "Austria", "Czechia"],
        },
        "points": points,
    }


def source_map(flows: list[dict[str, Any]], records: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    counts = Counter(flow.get("province") or "未记录" for flow in flows)
    records_by_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records or []:
        if record.get("id"):
            records_by_id[str(record.get("id"))].append(record)
    points = []
    for province, count in counts.most_common():
        if province == "未记录":
            continue
        first = next((flow for flow in flows if flow.get("province") == province), {})
        province_flows = [flow for flow in flows if flow.get("province") == province]
        works: list[dict[str, Any]] = []
        for flow in province_flows:
            matched_records = records_by_id.get(str(flow.get("id"))) or [flow]
            for record in matched_records:
                year = _record_year(record)
                works.append(
                    {
                        "title": _record_title(record),
                        "year": year,
                        "city": str(record.get("city") or flow.get("toLabel") or "").split("·")[0].strip(),
                        "country": record.get("country") or flow.get("country") or "",
                        "publisher": record.get("publisher") or "",
                    }
                )
        year_counts = Counter(int(work.get("year") or 0) for work in works if int(work.get("year") or 0))
        points.append(
            {
                "id": province,
                "province": province,
                "label": province,
                "count": count,
                "coords": first.get("from") or [116.4, 39.9],
                "examples": [flow.get("title") for flow in flows if flow.get("province") == province][:5],
                "works": works[:60],
                "years": sorted(year_counts),
                "yearCounts": {str(year): total for year, total in sorted(year_counts.items())},
            }
        )
    return {
        "type": "sourceMap",
        "title": "德译中国故事集取材来源地图",
        "subtitle": "以中国地图呈现故事来源、民族来源和地域叙事分布。",
        "geo": {"china": "/api/basemap/province"},
        "points": points,
    }


def preface_cluster(collections: list[dict[str, Any]]) -> dict[str, Any]:
    return topic_cluster_from_entries(preface_entries_from_collections(collections))


def word_cloud(collections: list[dict[str, Any]]) -> dict[str, Any]:
    entries = preface_entries_from_collections(collections)
    words = preface_keyword_weights_from_entries(entries, 1000, min_count=4) if entries else preface_keyword_weights_from_texts([preface_source(collections)], 1000, min_count=4)
    return {
        "type": "wordCloud",
        "title": "序跋文本词云图",
        "subtitle": "展示序跋文本中不同主题群的高频关键词。",
        "words": words,
    }


def preface_tokens(text: str) -> list[str]:
    return tokenize_text(text)


def preface_visuals(prefaces: dict[str, Any] | None = None) -> dict[str, Any]:
    entries = []
    for collection_id, preface in (prefaces or {}).items():
        text = str((preface or {}).get("text") or "")
        words = preface_keyword_weights_from_texts([text], 500)
        if words:
            entries.append(
                {
                    "id": collection_id,
                    "label": (preface or {}).get("sourceTitle") or (preface or {}).get("filename") or collection_id,
                    "year": int((preface or {}).get("year") or 0),
                    "text": text,
                    "words": words,
                }
            )
    word_clouds = [{"id": "all", "label": "总词云", "words": preface_keyword_weights_from_entries(entries, 1000, min_count=4)}]
    word_clouds.extend(entries)
    return {
        "cluster": topic_cluster_from_entries(entries),
        "wordClouds": [item for item in word_clouds if item["words"]],
    }


def child_cooccurrence(collections: list[dict[str, Any]], children: list[dict[str, Any]]) -> dict[str, Any]:
    child_by_id = {item["id"]: item for item in children}
    node_counts: Counter = Counter()
    node_years: dict[str, set[int]] = defaultdict(set)
    pair_counts: Counter = Counter()
    pair_years: dict[str, set[int]] = defaultdict(set)
    decade_theme_counts: dict[str, Counter] = defaultdict(Counter)
    theme_type_counts: dict[str, Counter] = defaultdict(Counter)
    theme_language_counts: dict[str, Counter] = defaultdict(Counter)
    theme_examples: dict[str, list[str]] = defaultdict(list)
    for collection in collections:
        collection_context = " ".join(
            str(collection.get(key) or "")
            for key in ["name", "foreignTitle", "prefaceText", "prefaceIntro", "editorRole", "publisher"]
        )
        collection_terms = set(top_terms_for_text(collection_context, 10)[:4])
        themes = set(term for term in collection_terms if valid_token(term) and len(term) <= 8)
        collection_year = int(collection.get("year") or 0)
        decade = f"{collection_year // 10 * 10}s" if collection_year else "未记录"
        for child_id in collection.get("matchedChildIds", []):
            child = child_by_id.get(child_id)
            if child:
                child_theme_values = child_themes(child)
                themes.update(child_theme_values)
                story_type = infer_child_story_type(child, child_theme_values)
                language = child.get("language") or collection.get("language") or "未标注语种"
                title = child.get("canonicalName") or child.get("variantName") or collection.get("name")
                for theme in child_theme_values:
                    theme_type_counts[theme][story_type] += 1
                    theme_language_counts[theme][language] += 1
                    if title and len(theme_examples[theme]) < 8:
                        theme_examples[theme].append(title)
        ordered = sorted(themes)
        node_counts.update(ordered)
        for theme in ordered:
            if collection_year:
                node_years[theme].add(collection_year)
            decade_theme_counts[decade][theme] += 1
        for index, source in enumerate(ordered):
            for target in ordered[index + 1 :]:
                key = "|".join(sorted([source, target]))
                pair_counts[key] += 1
                if collection_year:
                    pair_years[key].add(collection_year)
    nodes = [
        {
            "id": key,
            "label": key,
            "count": value,
            "years": sorted(node_years.get(key, set())),
            "types": [{"name": name, "count": count} for name, count in theme_type_counts[key].most_common(5)],
            "languages": [{"name": name, "count": count} for name, count in theme_language_counts[key].most_common(5)],
            "examples": theme_examples.get(key, [])[:6],
        }
        for key, value in node_counts.most_common(28)
    ]
    node_ids = {node["id"] for node in nodes}
    edges = []
    for key, value in pair_counts.most_common(72):
        source, target = key.split("|")
        if source in node_ids and target in node_ids:
            edges.append({"source": source, "target": target, "weight": value, "years": sorted(pair_years.get(key, set()))})
    decades = sorted(decade_theme_counts, key=lambda value: (value == "未记录", value))
    top_theme_ids = [node["id"] for node in nodes[:14]]
    timeline = [
        {
            "period": decade,
            "total": sum(decade_theme_counts[decade].values()),
            "themes": [{"theme": theme, "count": decade_theme_counts[decade][theme]} for theme in top_theme_ids if decade_theme_counts[decade][theme]],
        }
        for decade in decades
    ]
    type_nodes = Counter()
    type_edges = []
    for theme in top_theme_ids[:12]:
        for name, count in theme_type_counts[theme].most_common(4):
            type_nodes[name] += count
            type_edges.append({"source": theme, "target": name, "weight": count})
    burst_rows = []
    for theme in top_theme_ids:
        counts = [decade_theme_counts[decade][theme] for decade in decades]
        total = sum(counts)
        if not total:
            continue
        max_count = max(counts)
        peak = decades[counts.index(max_count)]
        burst_rows.append({"theme": theme, "peak": peak, "score": round(max_count / total, 3), "total": total})
    burst_rows.sort(key=lambda row: (row["score"], row["total"]), reverse=True)

    selected_collections: list[dict[str, Any]] = []
    seen_collection_ids: set[str] = set()
    for collection in [*(collections[:1]), *sorted(collections, key=lambda item: len(item.get("matchedChildIds", [])), reverse=True)]:
        if not collection.get("matchedChildIds"):
            continue
        collection_id = str(collection.get("id") or collection.get("name") or len(selected_collections))
        if collection_id in seen_collection_ids:
            continue
        selected_collections.append(collection)
        seen_collection_ids.add(collection_id)
        if len(selected_collections) >= 16:
            break

    structure_nodes: dict[str, dict[str, Any]] = {}
    structure_edge_weights: dict[tuple[str, str, str], float] = defaultdict(float)
    structure_edge_examples: dict[tuple[str, str, str], list[str]] = defaultdict(list)

    def graph_id(prefix: str, value: Any) -> str:
        raw = str(value or "未记录").strip() or "未记录"
        digest = hashlib.md5(raw.encode("utf-8")).hexdigest()[:10]
        return f"{prefix}-{digest}"

    def add_node(node_id: str, label: Any, node_type: str, cluster: str, count: float = 1, **extra: Any) -> None:
        label_text = str(label or "未记录").strip() or "未记录"
        current = structure_nodes.get(node_id)
        if current:
            current["count"] = current.get("count", 0) + count
            current.setdefault("examples", [])
            for example in extra.get("examples", []) or []:
                if example and example not in current["examples"] and len(current["examples"]) < 8:
                    current["examples"].append(example)
            return
        structure_nodes[node_id] = {
            "id": node_id,
            "label": label_text,
            "type": node_type,
            "cluster": cluster,
            "count": count,
            **extra,
        }

    def add_edge(source: str, target: str, relation: str, weight: float = 1, example: str = "") -> None:
        if not source or not target or source == target:
            return
        key = (source, target, relation)
        structure_edge_weights[key] += weight
        if example and example not in structure_edge_examples[key] and len(structure_edge_examples[key]) < 6:
            structure_edge_examples[key].append(example)

    def publisher_label(collection: dict[str, Any]) -> str:
        publisher = str(collection.get("publisher") or "未记录出版社").strip()
        if ":" in publisher:
            publisher = publisher.split(":", 1)[1].strip()
        return short(re.sub(r"\s+", " ", publisher), 28)

    added_children: set[str] = set()
    type_theme_edges: Counter = Counter()
    for collection_index, collection in enumerate(selected_collections):
        collection_raw_id = str(collection.get("id") or collection.get("name") or f"collection-{collection_index}")
        collection_id = f"collection-{collection_raw_id}"
        collection_title = collection.get("chineseTitle") or collection.get("name") or collection.get("foreignTitle") or collection_raw_id
        matched_ids = [child_id for child_id in collection.get("matchedChildIds", []) if child_id in child_by_id]
        year = int(collection.get("year") or 0)
        decade = f"{year // 10 * 10}s" if year else "未记录年代"
        role = classify_role(str(collection.get("editorRole") or ""), str(collection.get("editor") or ""))
        city = (clean_cities(str(collection.get("city") or collection.get("publisher") or "")) or ["未记录出版地"])[0]
        publisher = publisher_label(collection)
        editor = collection.get("editor") or "未记录译者/编者"
        region = collection.get("sourceRegion") or collection.get("sourceProvince") or collection.get("country") or "未记录来源"

        add_node(
            collection_id,
            collection_title,
            "故事集",
            "书目层",
            count=max(1, len(matched_ids)),
            year=year,
            summary=f"{collection.get('foreignTitle') or collection.get('name') or ''}；匹配子故事 {len(matched_ids)} 条",
            examples=[collection.get("foreignTitle") or ""],
        )
        decade_id = graph_id("decade", decade)
        role_id = graph_id("role", role)
        editor_id = graph_id("editor", editor)
        publisher_id = graph_id("publisher", publisher)
        city_id = graph_id("city", city)
        region_id = graph_id("region", region)
        add_node(decade_id, decade, "年代", "时间层", count=1, summary=f"出版年份：{year or '未记录'}")
        add_node(role_id, role, "译者身份", "主体层", count=1, summary=str(collection.get("editorRole") or ""))
        add_node(editor_id, short(editor, 24), "译者/编者", "主体层", count=1, summary=str(collection.get("editorRole") or ""))
        add_node(publisher_id, publisher, "出版机构", "出版层", count=1, summary=str(collection.get("publisher") or ""))
        add_node(city_id, city, "出版地", "出版层", count=1, summary=str(collection.get("country") or ""))
        add_node(region_id, short(region, 24), "来源区域", "来源层", count=1)
        add_edge(collection_id, decade_id, "出版年代", 1.6)
        add_edge(collection_id, role_id, "编译身份", 1.4)
        add_edge(collection_id, editor_id, "译者/编者", 1.3)
        add_edge(collection_id, publisher_id, "出版机构", 1.2)
        add_edge(collection_id, city_id, "出版地", 1.1)
        add_edge(collection_id, region_id, "来源区域", 1.0)

        child_limit = len(matched_ids) if collection_index == 0 else (10 if collection_index < 4 else 6)
        for child_id in matched_ids[:child_limit]:
            if len(added_children) >= 90 and child_id not in added_children:
                continue
            child = child_by_id[child_id]
            child_title = child.get("canonicalName") or child.get("variantName") or child_id
            child_node_id = f"child-{child_id}"
            child_theme_values = child_themes(child)
            inferred_type = infer_child_story_type(child, child_theme_values)
            language = child.get("language") or collection.get("language") or "未标注语种"
            type_id = graph_id("type", inferred_type)
            language_id = graph_id("language", language)
            add_node(
                child_node_id,
                child_title,
                "子故事",
                "文本层",
                count=1 + len(child_theme_values) * 0.35,
                year=int(child.get("year") or year or 0),
                summary=f"{child.get('variantName') or ''}；{inferred_type}",
                examples=[collection_title],
            )
            added_children.add(child_id)
            add_node(type_id, inferred_type, "推断类型", "类型层", count=1, examples=[child_title])
            add_node(language_id, language, "语种", "出版层", count=1)
            add_edge(collection_id, child_node_id, "包含子故事", 1.8, child_title)
            add_edge(child_node_id, type_id, "类型归属", 1.5, child_title)
            add_edge(child_node_id, language_id, "译介语种", 0.9, child_title)
            add_edge(collection_id, type_id, "类型覆盖", 0.55, child_title)
            for theme in child_theme_values:
                theme_id = graph_id("theme", theme)
                add_node(theme_id, theme, "抽取主题", "主题层", count=1.2, examples=[child_title])
                add_edge(child_node_id, theme_id, "抽取主题", 1.4, child_title)
                add_edge(type_id, theme_id, "类型主题分布", 0.85, child_title)
                type_theme_edges[(type_id, theme_id)] += 1
            for source, target in combinations(child_theme_values[:4], 2):
                add_edge(graph_id("theme", source), graph_id("theme", target), "主题共现", 0.65, child_title)

    for (type_id, theme_id), count in type_theme_edges.most_common(80):
        add_edge(type_id, theme_id, "高频类型主题", max(0.7, min(4.0, count / 2)))

    structure_edges = [
        {"source": source, "target": target, "relation": relation, "weight": round(weight, 3), "examples": structure_edge_examples.get((source, target, relation), [])}
        for (source, target, relation), weight in structure_edge_weights.items()
        if source in structure_nodes and target in structure_nodes
    ]
    structure_edges.sort(key=lambda edge: edge["weight"], reverse=True)
    structure_node_values = list(structure_nodes.values())
    structure_node_values.sort(key=lambda node: (node.get("type") != "故事集", -float(node.get("count") or 0), str(node.get("label") or "")))

    return {
        "type": "childCooccurrence",
        "title": "德译中国故事集主题共现图",
        "subtitle": "基于子故事题名、类型、民族来源、语种与书目字段抽取主题，展示动态共现网络与子故事书目结构图谱。",
        "nodes": nodes,
        "edges": edges,
        "timeline": timeline,
        "types": [{"id": name, "label": name, "count": count} for name, count in type_nodes.most_common(14)],
        "typeEdges": type_edges[:48],
        "bursts": burst_rows[:12],
        "years": sorted({int(item.get("year") or 0) for item in collections if item.get("year")}),
        "structureGraph": {
            "title": "子故事语义—书目结构知识图谱",
            "subtitle": "将数据表格中的故事集、子故事、抽取主题、推断类型、年份、译者身份、出版机构与来源区域组织为多关系动态图谱。",
            "nodes": structure_node_values[:120],
            "edges": structure_edges[:220],
            "triples": [
                {
                    "subject": structure_nodes[edge["source"]]["label"],
                    "predicate": edge["relation"],
                    "object": structure_nodes[edge["target"]]["label"],
                    "weight": edge["weight"],
                }
                for edge in structure_edges[:220]
            ],
            "stats": {
                "collections": len(selected_collections),
                "children": len(added_children),
                "nodes": len(structure_node_values),
                "edges": len(structure_edges),
                "triples": len(structure_edges),
            },
        },
    }


def wilhelm_rows(collections: list[dict[str, Any]]) -> list[dict[str, Any]]:
    pattern = re.compile(r"卫礼贤|Richard Wilhelm|Wilhelm|Chinesische Volksmärchen", re.I)
    rows = [item for item in collections if pattern.search(" ".join(str(item.get(key) or "") for key in ["name", "foreignTitle", "editor", "prefaceAuthor", "publisher"]))]
    return rows or [item for item in collections if 1910 <= int(item.get("year") or 0) <= 1920]


def normalize_province(value: Any) -> str:
    text = re.sub(r"省|市|自治区|特别行政区|壮族|回族|维吾尔", "", str(value or "北京")).strip()
    return text if text in PROVINCE_COORDS else "北京"


def wilhelm_visuals(records: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    data = story_data()
    rows = records or wilhelm_rows(data["collections"])
    normalized: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        city = row.get("city") or row.get("publisher") or "Berlin"
        province = normalize_province(row.get("province") or row.get("sourceProvince") or "北京")
        year_match = re.search(r"\d{4}", str(row.get("year") or row.get("yearText") or ""))
        year = int(year_match.group(0)) if year_match else int(row.get("year") or 0)
        normalized.append(
            {
                **row,
                "id": row.get("id") or f"wilhelm-backend-{index}",
                "title": row.get("title") or row.get("name") or "卫礼贤《中国民间童话》",
                "year": year,
                "city": city,
                "country": row.get("country") or ("中国" if city in {"Peking", "Beijing", "北京", "Shanghai", "上海"} else "德国"),
                "publisher": row.get("publisher") or "未记录",
                "province": province,
            }
        )
    flows = []
    for row in normalized:
        province = normalize_province(row.get("province"))
        target_coords = city_coords(row.get("city") or row.get("publisher") or "Berlin")
        flows.append(
            {
                "id": row["id"],
                "title": row.get("title") or row.get("name") or "卫礼贤《中国民间童话》",
                "sectionId": "stories",
                "resourceType": "卫礼贤《中国民间童话》专题",
                "language": row.get("language") or "德语",
                "year": row.get("year") or 0,
                "from": PROVINCE_COORDS[province],
                "to": target_coords,
                "fromLabel": province,
                "toLabel": f"{row.get('city') or '德国'} · {row.get('country') or '德国'}",
                "province": province,
                "country": row.get("country") or "德国",
                "weight": 0.9,
            }
        )
    return {
        "publicationMap": publication_map(normalized, "卫礼贤《中国民间童话》再版出版地图"),
        "flows": flows,
    }


WILHELM_TERM_LEXICON = {
    "动物形象": ["龙", "龙王", "虎", "蛇", "白蛇", "狐", "狐狸", "牛", "马", "鸟", "乌鸦", "鱼", "龟", "猴", "犬", "狗", "猫", "兔", "鼠", "鹿", "鹤", "鹰", "羊", "虫"],
    "自然气象": ["天", "云", "风", "雨", "雷", "电", "月", "太阳", "星", "雪", "雾", "霜", "海潮", "月宫", "广寒宫"],
    "神怪信仰": ["神", "仙", "鬼", "怪", "妖", "魔", "女神", "仙女", "观音", "玉帝", "术士", "道士", "法术", "天宫", "地府"],
    "人物身份": ["皇帝", "国王", "公主", "王子", "农夫", "和尚", "猎人", "妻子", "母亲", "父亲", "兄弟", "老人", "书生", "姑娘", "少年", "官员"],
    "空间场景": ["山", "海", "河", "井", "桥", "村", "宫", "田", "森林", "庙", "池塘", "龙宫", "月宫", "天庭"],
    "伦理母题": ["孝", "报恩", "惩罚", "复仇", "智慧", "善良", "忠诚", "贪婪", "仁慈", "欺骗", "禁忌", "婚姻", "考验"],
    "行为事件": ["变形", "求婚", "逃亡", "救助", "惩罚", "复仇", "偷盗", "斗争", "献祭", "成仙", "显灵", "报恩"],
    "核心意象": ["月亮", "珍珠", "宝物", "长生不老药", "葫芦", "莲花", "钟", "镜子", "绳子", "火", "水", "树"],
}

WILHELM_RELATIONS = [
    ("动物形象", "出现于"),
    ("自然气象", "关联气象"),
    ("神怪信仰", "关联神怪"),
    ("人物身份", "涉及身份"),
    ("空间场景", "发生空间"),
    ("伦理母题", "表达母题"),
    ("行为事件", "驱动情节"),
    ("核心意象", "象征意象"),
]

WILHELM_KEYWORD_STOP = {
    "本书", "本书中", "在本书中", "选取", "选取的", "被选取", "读者", "儿童", "童话", "艺术童话", "书中",
    "中国", "民间", "故事", "本文", "这里", "那里", "一个", "一些", "后来", "时候", "因此", "以及", "同时",
    "出现", "内容", "文本", "译文", "意义上", "作品", "讲述", "叙述", "分类", "关键词", "未分类",
    "在中国", "中选取", "选取在", "本书选取", "选取的故事", "本书的", "中的", "的话", "他们", "人们", "自己",
    "请求", "因为", "但是", "这个", "那个", "得到", "进行", "具有", "成为", "使用", "需要", "可以", "关系",
    "文学", "民间文学", "卫礼贤", "中国民间童话",
}

WILHELM_ALLOWED_CATEGORIES = {"动物形象", "自然气象", "神怪信仰", "人物身份", "空间场景", "伦理母题", "行为事件", "核心意象"}

NLP_ALLOWED_POS = {"n", "nr", "ns", "nt", "nz", "vn", "v", "vd", "a", "an", "i", "l"}
CN_FUNCTION_STOP = {
    "的", "了", "着", "过", "是", "在", "有", "和", "与", "及", "或", "而", "并", "也", "都", "就", "才", "乃", "其", "之", "者", "所",
    "于", "从", "到", "对", "为", "以", "因", "由", "但", "被", "把", "将", "让", "使", "给", "这", "那", "他", "她", "它", "们", "一个",
    "一些", "这个", "那个", "这些", "那些", "自己", "人们", "他们", "她们", "里面", "这里", "那里", "后来", "于是", "因此", "所以",
    "可以", "没有", "已经", "正在", "进行", "出现", "内容", "文本", "译文", "本书", "书中", "读者", "故事", "童话", "民间", "中国",
}
NOISE_PHRASE_PARTS = {
    "本书", "书中", "在本书", "选取", "被选取", "读者", "关键词", "未分类", "意义上", "请求", "因为", "但是", "需要", "可以", "具有",
    "进行", "出现", "内容", "文本", "译文", "中国民间童话", "卫礼贤", "艺术童话", "童话故事", "民间故事",
    "意义", "艺术", "体裁", "动物", "民间传说", "颇受欢迎", "叙述", "讲述", "来源", "题名", "分类", "素材", "说道", "领域",
}
ONE_CHAR_TERMS = {"龙", "虎", "蛇", "牛", "马", "鸟", "狐", "猴", "鬼", "神", "仙", "山", "海", "河", "月", "日", "星", "风", "雨", "云", "水", "火", "树", "村", "宫", "庙", "桥", "父", "母", "妻", "夫", "王"}


def _story_full_text(story: dict[str, Any]) -> str:
    return " ".join(str(part or "") for part in [story.get("title"), story.get("source"), story.get("category"), story.get("text")])



def wilhelm_story_terms(story: dict[str, Any]) -> list[dict[str, Any]]:
    terms: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in candidate_terms_for_wilhelm([story], 54):
        token = normalize_wilhelm_keyword(item["term"])
        category = item.get("categoryHint") or category_for_wilhelm_term(token)
        if token and token not in seen and category in WILHELM_ALLOWED_CATEGORIES:
            terms.append({"term": token, "category": category, "count": item["count"]})
            seen.add(token)
    return sorted(terms, key=lambda item: (item["count"], len(item["term"])), reverse=True)[:42]


def normalize_wilhelm_keyword(term: Any) -> str:
    punctuation = "\uFF0C\u3002\u3001\uFF1B\uFF1A:,.!?\uFF01\uFF1F()\uFF08\uFF09[]\u3010\u3011\u300A\u300B\u2018\u2019\u201C\u201D"
    text = re.sub(r"\s+", "", str(term or "")).strip(punctuation)
    while len(text) > 1 and text[:1] in CN_FUNCTION_STOP:
        text = text[1:]
    while len(text) > 2 and text[-1:] in CN_FUNCTION_STOP:
        text = text[:-1]
    return text


def meaningful_wilhelm_keyword(term: Any) -> bool:
    text = normalize_wilhelm_keyword(term)
    if len(text) > 8:
        return False
    if len(text) < 2 and text not in ONE_CHAR_TERMS:
        return False
    if text in WILHELM_KEYWORD_STOP or text.lower() in WILHELM_KEYWORD_STOP or text in CN_FUNCTION_STOP:
        return False
    if re.fullmatch(r"\d+|[A-Za-z]{1,3}", text):
        return False
    if any(stop in text for stop in NOISE_PHRASE_PARTS):
        return False
    if any(text.startswith(stop) or text.endswith(stop) for stop in CN_FUNCTION_STOP if len(stop) == 1):
        return False
    if any(noise in text for noise in ["一个", "一些", "这个", "那个", "这些", "那些", "他们", "她们", "它们", "自己", "正在", "已经"]):
        return False
    return bool(re.search(r"[\u4e00-\u9fffA-Za-z]", text))


def candidate_terms_for_wilhelm(stories: list[dict[str, Any]], limit: int = 220) -> list[dict[str, Any]]:
    counts: Counter[str] = Counter()
    category_hint: dict[str, str] = {}
    for story in stories:
        text = _story_full_text(story)
        if jieba and pseg:
            for word, flag in pseg.cut(text):
                token = normalize_wilhelm_keyword(word)
                pos = str(flag or "").lower()
                if not meaningful_wilhelm_keyword(token):
                    continue
                if not any(pos.startswith(prefix) for prefix in NLP_ALLOWED_POS):
                    continue
                counts[token] += 1
                category_hint.setdefault(token, category_for_wilhelm_term(token))
            if getattr(jieba, "analyse", None):
                for extractor in (jieba.analyse.textrank, jieba.analyse.extract_tags):
                    try:
                        for token, weight in extractor(text, topK=max(40, limit), withWeight=True, allowPOS=tuple(NLP_ALLOWED_POS)):
                            token = normalize_wilhelm_keyword(token)
                            if not meaningful_wilhelm_keyword(token):
                                continue
                            counts[token] += max(1, round(float(weight) * 6))
                            category_hint.setdefault(token, category_for_wilhelm_term(token))
                    except Exception:
                        continue
        else:
            for token, count in Counter(tokenize_text(text)).most_common(max(60, limit)):
                token = normalize_wilhelm_keyword(token)
                if meaningful_wilhelm_keyword(token):
                    counts[token] += count
                    category_hint.setdefault(token, category_for_wilhelm_term(token))
    ranked = sorted(counts.items(), key=lambda item: (item[1] * min(6, len(item[0])), item[1]), reverse=True)
    return [
        {"term": term, "count": count, "categoryHint": category_hint.get(term, category_for_wilhelm_term(term))}
        for term, count in ranked[:limit]
        if meaningful_wilhelm_keyword(term) and category_hint.get(term, category_for_wilhelm_term(term)) in WILHELM_ALLOWED_CATEGORIES
    ]


def _load_wilhelm_keyword_network_cache() -> dict[str, Any]:
    if not WILHELM_KEYWORD_NETWORK_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(WILHELM_KEYWORD_NETWORK_CACHE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_wilhelm_keyword_network_cache(cache: dict[str, Any]) -> None:
    WILHELM_KEYWORD_NETWORK_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def wilhelm_keyword_cache_key(stories: list[dict[str, Any]]) -> str:
    digest = hashlib.sha256()
    digest.update(b"wilhelm-keywords-v6-jieba-pos")
    for story in stories:
        digest.update(str(story.get("id") or "").encode("utf-8"))
        digest.update(str(story.get("title") or "").encode("utf-8"))
        digest.update(str(story.get("text") or "")[:3000].encode("utf-8"))
    return digest.hexdigest()


def llm_extract_wilhelm_keywords(stories: list[dict[str, Any]], model: str = "") -> list[dict[str, Any]]:
    candidates = candidate_terms_for_wilhelm(stories)
    if not candidates:
        return []
    prompt = (
        "请只输出 JSON，不要输出解释。你要为卫礼贤《中国民间童话》全部译文筛选关键词并分类。\n"
        "只保留具有故事学、民俗学、叙事母题、角色形象、空间意象或传播研究意义的关键词。\n"
        "必须剔除泛词和无意义词，例如：本书、本书中、选取、读者、童话、艺术童话、在中国、意义上、关键词、未分类。\n"
        "类别只能从这些值选择：动物形象、自然气象、神怪信仰、人物身份、空间场景、伦理母题、行为事件、核心意象。\n"
        "输出格式：{\"keywords\":[{\"term\":\"龙\",\"category\":\"动物形象\"}]}\n"
        "候选词频与算法提示如下：\n"
        f"{json.dumps(candidates, ensure_ascii=False)}"
    )
    answer = chat_completion(
        [{"role": "system", "content": "你是民间文学关键词抽取与分类助手，只返回合法 JSON。"}, {"role": "user", "content": prompt}],
        model=model,
        temperature=0.15,
        timeout=120,
    )
    payload = _extract_json_object(answer)
    raw_keywords = payload.get("keywords") or payload.get("terms") or []
    keywords: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in raw_keywords:
        term = normalize_wilhelm_keyword(item.get("term") if isinstance(item, dict) else item)
        if not meaningful_wilhelm_keyword(term) or term in seen:
            continue
        category = str(item.get("category") if isinstance(item, dict) else fallback_keyword_category(term))
        if category not in WILHELM_ALLOWED_CATEGORIES:
            category = fallback_keyword_category(term)
        keywords.append({"term": term, "category": category})
        seen.add(term)
        if len(keywords) >= 90:
            break
    if not keywords:
        raise RuntimeError("大模型未返回有效关键词。")
    return keywords


def fallback_wilhelm_keywords(stories: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts: Counter[str] = Counter()
    categories: dict[str, str] = {}
    for story in stories:
        for item in wilhelm_story_terms(story):
            term = normalize_wilhelm_keyword(item.get("term"))
            category = item.get("category") or fallback_keyword_category(term)
            if meaningful_wilhelm_keyword(term) and category in WILHELM_ALLOWED_CATEGORIES:
                counts[term] += int(item.get("count") or 1)
                categories.setdefault(term, category)
    if not counts:
        for item in candidate_terms_for_wilhelm(stories, 90):
            term = normalize_wilhelm_keyword(item["term"])
            category = item.get("categoryHint") or fallback_keyword_category(term)
            if meaningful_wilhelm_keyword(term) and category in WILHELM_ALLOWED_CATEGORIES:
                counts[term] += int(item.get("count") or 1)
                categories.setdefault(term, category)
    ranked = sorted(counts, key=lambda term: (counts[term], len(term)), reverse=True)
    return [{"term": term, "category": categories.get(term) or fallback_keyword_category(term)} for term in ranked[:90]]


def build_wilhelm_keyword_graph(stories: list[dict[str, Any]], keywords: list[dict[str, Any]], graph_id: str = "total", title: str = "卫礼贤《中国民间童话》") -> dict[str, Any]:
    allowed = [(item["term"], item.get("category") or fallback_keyword_category(item["term"])) for item in keywords if meaningful_wilhelm_keyword(item.get("term"))]
    term_counts: Counter[str] = Counter()
    category_counts: Counter[str] = Counter()
    story_links: dict[str, list[dict[str, Any]]] = defaultdict(list)
    pair_counts: Counter[tuple[str, str]] = Counter()
    triples: list[dict[str, Any]] = []
    for story in stories:
        story_id = story.get("id") or f"story-{len(story_links) + 1}"
        story_title = story.get("title") or "未命名故事"
        text = " ".join(str(part or "") for part in [story_title, story.get("source"), story.get("category"), story.get("text")])
        present: list[str] = []
        for term, category in allowed:
            count = len(re.findall(re.escape(term), text, flags=re.I))
            if not count:
                continue
            term_counts[term] += count
            category_counts[category] += count
            present.append(term)
            story_links[term].append({"storyId": story_id, "storyTitle": story_title, "category": category, "count": count})
            triples.append({"subject": term, "predicate": category, "object": story_title, "weight": count, "storyId": story_id})
        for source, target in combinations(sorted(set(present)), 2):
            pair_counts[(source, target)] += 1

    top_terms = [term for term, _ in term_counts.most_common(70)]
    nodes = [
        {
            "id": term,
            "label": term,
            "type": next((category for word, category in allowed if word == term), fallback_keyword_category(term)),
            "count": term_counts[term],
            "stories": sorted(story_links[term], key=lambda row: row["count"], reverse=True)[:16],
        }
        for term in top_terms
    ]
    node_ids = {node["id"] for node in nodes}
    cooccurrence = [
        {"source": source, "target": target, "weight": count}
        for (source, target), count in pair_counts.most_common(160)
        if source in node_ids and target in node_ids
    ]
    return {
        "id": graph_id,
        "title": title,
        "nodes": nodes,
        "edges": [{"source": edge["source"], "target": edge["target"], "relation": "共现", "weight": edge["weight"]} for edge in cooccurrence],
        "triples": triples[:260],
        "cooccurrence": cooccurrence,
        "terms": [
            {
                "term": term,
                "category": next((category for word, category in allowed if word == term), fallback_keyword_category(term)),
                "count": count,
                "stories": sorted(story_links[term], key=lambda row: row["count"], reverse=True)[:16],
            }
            for term, count in term_counts.most_common(90)
        ],
    }


def wilhelm_keyword_network(stories: list[dict[str, Any]] | None = None, force: bool = False, model: str = "", method: str = "algorithm") -> dict[str, Any]:
    data = story_data()
    rows = stories or data.get("wilhelmStories", [])
    cache_key = wilhelm_keyword_cache_key(rows)
    cache = _load_wilhelm_keyword_network_cache()
    cached = cache.get(cache_key)
    used_fallback = False
    use_llm = method == "llm"
    if cached and not force and cached.get("keywords"):
        keywords = cached.get("keywords") or []
        source = cached.get("source") or "algorithm"
    else:
        if use_llm:
            try:
                keywords = llm_extract_wilhelm_keywords(rows, model=model)
                source = "llm"
            except Exception:
                keywords = fallback_wilhelm_keywords(rows)
                source = "algorithm"
                used_fallback = True
        else:
            keywords = fallback_wilhelm_keywords(rows)
            source = "algorithm"
        if keywords:
            cache[cache_key] = {"keywords": keywords, "fallback": used_fallback, "source": source}
            _save_wilhelm_keyword_network_cache(cache)
    total = build_wilhelm_keyword_graph(rows, keywords, "total", "卫礼贤《中国民间童话》")
    by_story = {
        (story.get("id") or f"story-{index + 1}"): build_wilhelm_keyword_graph([story], keywords, story.get("id") or f"story-{index + 1}", story.get("title") or "单篇译文")
        for index, story in enumerate(rows)
    }
    return {
        "cached": bool(cached and not force and cached.get("keywords")),
        "fallback": used_fallback,
        "source": source,
        "keywords": keywords,
        "total": total,
        "byStory": by_story,
    }


def build_wilhelm_graph(stories: list[dict[str, Any]], graph_id: str = "total", title: str = "卫礼贤《中国民间童话》") -> dict[str, Any]:
    term_counts: Counter = Counter()
    category_counts: Counter = Counter()
    story_links: dict[str, list[dict[str, Any]]] = defaultdict(list)
    pair_counts: Counter = Counter()
    triples: list[dict[str, Any]] = []
    story_nodes: list[dict[str, Any]] = []

    for story in stories:
        story_id = story.get("id") or f"story-{len(story_nodes) + 1}"
        story_title = story.get("title") or "未命名故事"
        terms = wilhelm_story_terms(story)
        story_nodes.append({"id": f"story-{story_id}", "label": story_title, "type": "故事", "count": max(1, len(terms)), "storyId": story_id})
        term_names = [item["term"] for item in terms[:24]]
        for item in terms:
            term_counts[item["term"]] += item["count"]
            category_counts[item["category"]] += item["count"]
            story_links[item["term"]].append({"storyId": story_id, "storyTitle": story_title, "category": item["category"], "count": item["count"]})
            relation = next((rel for category, rel in WILHELM_RELATIONS if category == item["category"]), "提取关键词")
            triples.append({"subject": story_title, "predicate": relation, "object": item["term"], "weight": item["count"], "storyId": story_id})
        for source, target in combinations(sorted(set(term_names)), 2):
            pair_counts[(source, target)] += 1

    top_terms = [term for term, _ in term_counts.most_common(48)]
    nodes: list[dict[str, Any]] = [{"id": "center", "label": title, "type": "故事", "count": len(stories)}]
    if len(stories) <= 16:
        nodes.extend(story_nodes)
    nodes.extend({"id": f"cat-{category}", "label": category, "type": "分类", "count": count} for category, count in category_counts.most_common(8))
    nodes.extend(
        {
            "id": f"term-{term}",
            "label": term,
            "type": next((link["category"] for link in story_links[term]), "关键词"),
            "count": term_counts[term],
            "stories": story_links[term][:12],
        }
        for term in top_terms
    )
    node_ids = {node["id"] for node in nodes}
    edges: list[dict[str, Any]] = []
    if len(stories) <= 16:
        for node in story_nodes:
            edges.append({"source": "center", "target": node["id"], "weight": node["count"], "relation": "包含"})
    for category, count in category_counts.most_common(8):
        edges.append({"source": "center", "target": f"cat-{category}", "weight": count, "relation": "主题分类"})
    for term in top_terms:
        category = next((link["category"] for link in story_links[term]), "关键词")
        category_id = f"cat-{category}"
        if category_id in node_ids:
            edges.append({"source": category_id, "target": f"term-{term}", "weight": term_counts[term], "relation": "抽取关键词"})
        if len(stories) <= 16:
            for link in story_links[term][:4]:
                story_id = f"story-{link['storyId']}"
                if story_id in node_ids:
                    edges.append({"source": story_id, "target": f"term-{term}", "weight": link["count"], "relation": "出现于"})
    cooccurrence = []
    for (source, target), count in pair_counts.most_common(120):
        if source in top_terms and target in top_terms:
            cooccurrence.append({"source": source, "target": target, "weight": count})
            edges.append({"source": f"term-{source}", "target": f"term-{target}", "weight": count, "relation": "共现"})
    return {
        "id": graph_id,
        "nodes": nodes,
        "edges": [edge for edge in edges if edge["source"] in node_ids and edge["target"] in node_ids],
        "triples": triples[:260],
        "cooccurrence": cooccurrence,
        "terms": [
            {
                "term": term,
                "category": next((link["category"] for link in story_links[term]), "关键词"),
                "count": count,
                "stories": sorted(story_links[term], key=lambda item: item["count"], reverse=True)[:12],
            }
            for term, count in term_counts.most_common(60)
        ],
    }


def _wilhelm_story_structure(stories: list[dict[str, Any]]) -> dict[str, Any]:
    categories: dict[str, dict[str, Any]] = {}
    for story in stories:
        key = str(story.get("category") or "未分类").strip() or "未分类"
        row = categories.setdefault(
            key,
            {
                "category": key,
                "count": 0,
                "length": 0,
                "sources": set(),
                "examples": [],
            },
        )
        row["count"] += 1
        row["length"] += len(str(story.get("text") or ""))
        source = str(story.get("source") or "").strip()
        if source:
            row["sources"].add(source)
        if len(row["examples"]) < 4:
            row["examples"].append(story.get("title") or "未命名译文")

    rows = []
    for row in categories.values():
        count = max(1, int(row["count"]))
        sources = sorted(row["sources"])
        rows.append(
            {
                "category": row["category"],
                "count": row["count"],
                "avgLength": round(row["length"] / count),
                "sourceCount": len(sources),
                "sources": sources,
                "examples": row["examples"],
            }
        )
    rows.sort(key=lambda item: (-item["count"], item["category"]))

    links = []
    for source_index, source in enumerate(rows):
        source_sources = set(source.get("sources") or [])
        for target_index, target in enumerate(rows[source_index + 1 :], start=source_index + 1):
            shared = sorted(source_sources.intersection(target.get("sources") or []))
            if shared:
                links.append(
                    {
                        "sourceIndex": source_index,
                        "targetIndex": target_index,
                        "weight": len(shared),
                        "sharedSources": shared[:6],
                    }
                )
    links.sort(key=lambda item: item["weight"], reverse=True)
    return {
        "title": "单篇译文结构谱系",
        "categories": rows[:16],
        "links": links[:24],
        "method": "后端按卫礼贤分类、译文长度与来源重合度计算",
    }


def _wilhelm_time_density(records: list[dict[str, Any]]) -> dict[str, Any]:
    periods: dict[str, dict[str, Any]] = {}
    for record in records:
        year_match = re.search(r"\d{4}", str(record.get("year") or record.get("yearText") or ""))
        year = int(year_match.group(0)) if year_match else 0
        period = f"{year // 10 * 10}s" if year else "未记录"
        row = periods.setdefault(
            period,
            {
                "period": period,
                "count": 0,
                "years": [],
                "publishers": set(),
                "cities": set(),
                "languages": set(),
                "examples": [],
            },
        )
        row["count"] += 1
        if year:
            row["years"].append(year)
        for key, bucket in [("publisher", "publishers"), ("city", "cities"), ("language", "languages")]:
            value = str(record.get(key) or "").strip()
            if value:
                row[bucket].add(value)
        if len(row["examples"]) < 4:
            row["examples"].append(record.get("title") or record.get("name") or "卫礼贤《中国民间童话》")

    def sort_key(item: dict[str, Any]) -> tuple[int, str]:
        match = re.search(r"\d{4}", item["period"])
        return (int(match.group(0)) if match else 9999, item["period"])

    rows = []
    for row in periods.values():
        years = sorted(row["years"])
        publishers = sorted(row["publishers"])
        cities = sorted(row["cities"])
        languages = sorted(row["languages"])
        rows.append(
            {
                "period": row["period"],
                "count": row["count"],
                "years": years,
                "minYear": years[0] if years else None,
                "maxYear": years[-1] if years else None,
                "publishers": publishers,
                "publisherCount": len(publishers),
                "cities": cities,
                "cityCount": len(cities),
                "languages": languages,
                "languageCount": len(languages),
                "examples": row["examples"],
            }
        )
    rows.sort(key=sort_key)
    return {
        "title": "再版传播时间密度",
        "periods": rows,
        "method": "后端按再版年代、出版城市、出版社与语种聚合计算",
    }


def wilhelm_story_analysis(stories: list[dict[str, Any]] | None = None, records: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    data = story_data()
    rows = stories or data.get("wilhelmStories", [])
    edition_rows = records or data.get("wilhelmEditions", []) or wilhelm_rows(data.get("collections", []))
    cached_by_story = data.get("wilhelmStoryGraphs", {})
    cached_total = data.get("wilhelmThemeGraph")
    total_graph = cached_total if isinstance(cached_total, dict) and cached_total.get("nodes") else build_wilhelm_graph(rows, "total", "卫礼贤《中国民间童话》")
    return {
        "total": total_graph,
        "byStory": cached_by_story if isinstance(cached_by_story, dict) else {},
        "structure": _wilhelm_story_structure(rows),
        "timeDensity": _wilhelm_time_density(edition_rows),
        "recordCount": len(edition_rows),
        "storyCount": len(rows),
    }


def _load_wilhelm_graph_cache() -> dict[str, Any]:
    if not WILHELM_GRAPH_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(WILHELM_GRAPH_CACHE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_wilhelm_graph_cache(cache: dict[str, Any]) -> None:
    WILHELM_GRAPH_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def _extract_json_object(text: str) -> dict[str, Any]:
    raw = str(text or "").strip()
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", raw, flags=re.S)
    if fenced:
        raw = fenced.group(1)
    else:
        start = raw.find("{")
        end = raw.rfind("}")
        if start >= 0 and end > start:
            raw = raw[start:end + 1]
    return json.loads(raw)


def _normalize_llm_graph(payload: dict[str, Any], title: str) -> dict[str, Any]:
    nodes = []
    seen = set()
    for index, node in enumerate(payload.get("nodes") or []):
        label = str(node.get("label") or node.get("name") or node.get("id") or "").strip()
        if not label:
            continue
        node_id = str(node.get("id") or f"node-{index + 1}")
        if node_id in seen:
            node_id = f"{node_id}-{index + 1}"
        seen.add(node_id)
        nodes.append(
            {
                "id": node_id,
                "label": label,
                "type": str(node.get("type") or "实体"),
                "count": int(node.get("count") or node.get("weight") or 1),
                "summary": str(node.get("summary") or ""),
            }
        )
    label_to_id = {node["label"]: node["id"] for node in nodes}
    edges = []
    triples = []
    for index, edge in enumerate(payload.get("edges") or payload.get("triples") or []):
        source_label = str(edge.get("source") or edge.get("subject") or "").strip()
        target_label = str(edge.get("target") or edge.get("object") or "").strip()
        relation = str(edge.get("relation") or edge.get("predicate") or "关联").strip()
        source = edge.get("sourceId") or label_to_id.get(source_label) or source_label
        target = edge.get("targetId") or label_to_id.get(target_label) or target_label
        if source not in seen and source_label:
            source = label_to_id.get(source_label, source)
        if target not in seen and target_label:
            target = label_to_id.get(target_label, target)
        if source in seen and target in seen:
            weight = int(edge.get("weight") or edge.get("count") or 1)
            edges.append({"source": source, "target": target, "relation": relation, "weight": weight})
            triples.append({"subject": source_label or source, "predicate": relation, "object": target_label or target, "weight": weight})
    if not nodes:
        nodes = [{"id": "center", "label": title, "type": "故事", "count": 1, "summary": ""}]
    return {"nodes": nodes[:80], "edges": edges[:140], "triples": triples[:180], "title": title, "source": "llm"}


def wilhelm_llm_knowledge_graph(scope_id: str, title: str, text: str, force: bool = False, model: str = "", method: str = "algorithm") -> dict[str, Any]:
    cache = _load_wilhelm_graph_cache()
    cache_key = scope_id or title or "total"
    if not force and cache_key in cache:
        return {"cached": True, "graph": cache[cache_key]}

    source_text = str(text or "")[:9000]
    if method != "llm":
        fallback_story = {"id": cache_key, "title": title, "text": source_text, "category": "算法抽取"}
        graph = build_wilhelm_graph([fallback_story], cache_key, title)
        graph["source"] = "algorithm"
        graph["notice"] = "已使用后端自然语言算法抽取知识图谱；可选择大模型抽取覆盖。"
        cache[cache_key] = graph
        _save_wilhelm_graph_cache(cache)
        return {"cached": False, "fallback": False, "algorithm": True, "graph": graph}

    system_prompt = (
        "你是中国民间文学与知识图谱构建助手。请只输出 JSON，不要输出解释文字。"
        "根据给定译文抽取知识图谱，必须包含 nodes 和 edges。"
        "nodes 每项字段：id,label,type,count,summary。type 可用：故事、人物、动物、神怪、空间、母题、事件、物象、观念。"
        "edges 每项字段：source,target,relation,weight，其中 source/target 必须使用节点 id。"
        "关系应体现三元组，例如：龙 - 关联气象 - 风雨；公主 - 遭遇 - 变形；故事 - 包含母题 - 报恩。"
    )
    user_prompt = (
        f"图谱标题：{title}\n"
        f"请抽取 16-32 个节点、24-56 条关系，突出故事情节、角色、动物意象、空间、母题和象征。\n"
        f"译文文本：\n{source_text}"
    )
    try:
        answer = chat_completion(
            [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
            model=model,
            temperature=0.2,
            timeout=120,
        )
        graph = _normalize_llm_graph(_extract_json_object(answer), title)
        cache[cache_key] = graph
        _save_wilhelm_graph_cache(cache)
        return {"cached": False, "fallback": False, "graph": graph}
    except Exception as error:
        fallback_story = {"id": cache_key, "title": title, "text": source_text, "category": "临时抽取"}
        graph = build_wilhelm_graph([fallback_story], cache_key, title)
        graph["source"] = "algorithm-fallback"
        graph["notice"] = f"大模型未返回可用图谱文本，已临时使用本地抽取结果：{error}"
        graph["notice"] = "大模型暂未返回可用知识图谱，已自动切换为后端自然语言算法抽取并保存。"
        cache[cache_key] = graph
        _save_wilhelm_graph_cache(cache)
        return {"cached": False, "fallback": True, "graph": graph}


def _load_keyword_category_cache() -> dict[str, Any]:
    if not WILHELM_KEYWORD_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(WILHELM_KEYWORD_CACHE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_keyword_category_cache(cache: dict[str, Any]) -> None:
    WILHELM_KEYWORD_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")




def fallback_keyword_category(term: str) -> str:
    return category_for_wilhelm_term(term)


def category_for_wilhelm_term(term: str) -> str:
    text = normalize_wilhelm_keyword(term)
    groups = [
        ("\u52a8\u7269\u5f62\u8c61", ["\u9f99", "\u864e", "\u86c7", "\u725b", "\u9a6c", "\u9e1f", "\u72d0", "\u72d0\u72f8", "\u7334", "\u9c7c", "\u9e7f", "\u72ac", "\u72d7", "\u732b", "\u9e21", "\u732a", "\u7f8a", "\u868c", "\u9f9f", "\u9e70", "\u517d", "\u51e4\u51f0"]),
        ("\u4eba\u7269\u8eab\u4efd", ["\u7687\u5e1d", "\u56fd\u738b", "\u738b\u5b50", "\u516c\u4e3b", "\u519c\u592b", "\u730e\u4eba", "\u59bb\u5b50", "\u6bcd\u4eb2", "\u7236\u4eb2", "\u7236\u6bcd", "\u5144\u5f1f", "\u54e5\u54e5", "\u5f1f\u5f1f", "\u59d0\u59d0", "\u59b9\u59b9", "\u5973\u5a7f", "\u5973\u513f", "\u513f\u5b50", "\u8001\u4eba", "\u59d1\u5a18", "\u4e2b\u9b1f", "\u5c11\u5e74", "\u4e66\u751f", "\u5b98\u5458", "\u9053\u58eb", "\u548c\u5c1a", "\u9a6c\u500c", "\u4f8d\u5973", "\u8fdb\u58eb", "\u5cb3\u7236\u6bcd"]),
        ("\u795e\u602a\u4fe1\u4ef0", ["\u795e", "\u4ed9", "\u9b3c", "\u602a", "\u9b54", "\u5996", "\u9f99\u738b", "\u5929\u5bab", "\u5730\u5e9c", "\u89c2\u97f3", "\u7389\u5e1d", "\u6708\u5bab", "\u6cd5\u672f", "\u795e\u7075", "\u795e\u4ed9", "\u7cbe\u602a"]),
        ("\u7a7a\u95f4\u573a\u666f", ["\u5c71", "\u6d77", "\u6cb3", "\u4e95", "\u6865", "\u6751", "\u5bab", "\u5e99", "\u7530", "\u68ee\u6797", "\u5e9c", "\u90bb\u6751", "\u9f99\u5bab", "\u6708\u5bab", "\u5929\u5ead", "\u5bb6", "\u57ce"]),
        ("\u81ea\u7136\u6c14\u8c61", ["\u5929", "\u98ce", "\u96e8", "\u4e91", "\u96f7", "\u7535", "\u6708", "\u65e5", "\u661f", "\u592a\u9633", "\u6c34", "\u706b", "\u96ea", "\u971c", "\u6625", "\u590f", "\u79cb", "\u51ac"]),
        ("\u4f26\u7406\u6bcd\u9898", ["\u5b5d", "\u62a5\u6069", "\u60e9\u7f5a", "\u590d\u4ec7", "\u667a\u6167", "\u5584\u826f", "\u5fe0\u8bda", "\u8d2a\u5a6a", "\u4ec1\u6148", "\u5a5a\u59fb", "\u8003\u9a8c", "\u7981\u5fcc", "\u6b3a\u9a97", "\u6551\u52a9"]),
        ("\u884c\u4e3a\u4e8b\u4ef6", ["\u53d8\u5f62", "\u6c42\u5a5a", "\u9003\u4ea1", "\u6551\u52a9", "\u60e9\u7f5a", "\u590d\u4ec7", "\u5077\u76d7", "\u6597\u4e89", "\u732e\u796d", "\u6210\u4ed9", "\u663e\u7075", "\u62a5\u6069", "\u5bfb\u627e", "\u4f24\u5bb3", "\u62ef\u6551", "\u964d\u5bb3", "\u79cd\u51fa"]),
    ]
    for category, words in groups:
        if any(word and word in text for word in words):
            return category
    for category, terms in WILHELM_TERM_LEXICON.items():
        if text in terms:
            return category
    return "\u6838\u5fc3\u610f\u8c61"


def wilhelm_keyword_categories(terms: Optional[list[Any]] = None, force: bool = False, model: str = "") -> dict[str, Any]:
    labels = []
    for item in terms or []:
        label = str(item.get("term") if isinstance(item, dict) else item or "").strip()
        if label and label not in labels:
            labels.append(label)
    labels = labels[:80]
    cache_key = hashlib.sha256("|".join(sorted(labels)).encode("utf-8")).hexdigest()
    cache = _load_keyword_category_cache()
    if not force and cache_key in cache:
        return {"cached": True, "categories": cache[cache_key]}
    if not labels:
        return {"cached": False, "categories": {}}

    prompt = (
        "请只输出 JSON，不要输出解释。对下列中国民间故事关键词进行学术化分类。\n"
        "类别只能从这些值选择：动物形象、自然气象、神怪信仰、人物身份、空间场景、伦理母题、行为事件、核心意象。\n"
        "输出格式：{\"categories\":{\"关键词\":\"类别\"}}\n"
        f"关键词：{json.dumps(labels, ensure_ascii=False)}"
    )
    fallback = {label: fallback_keyword_category(label) for label in labels}
    try:
        answer = chat_completion(
            [{"role": "system", "content": "你是民间文学关键词分类助手，只返回合法 JSON。"}, {"role": "user", "content": prompt}],
            model=model,
            temperature=0.1,
            timeout=90,
        )
        parsed = _extract_json_object(answer)
        categories = parsed.get("categories") or parsed
        used_fallback = False
    except Exception:
        categories = fallback
        used_fallback = True
    normalized = {label: str(categories.get(label) or fallback_keyword_category(label)) for label in labels}
    allowed = {"动物形象", "自然气象", "神怪信仰", "人物身份", "空间场景", "伦理母题", "行为事件", "核心意象"}
    normalized = {label: (cat if cat in allowed else fallback_keyword_category(label)) for label, cat in normalized.items()}
    cache[cache_key] = normalized
    _save_keyword_category_cache(cache)
    return {"cached": False, "fallback": used_fallback, "categories": normalized}


def bucket_year(year: Any) -> str:
    value = int(year or 0)
    if not value:
        return "未知"
    if value < 1900:
        return "1900前"
    if value < 1920:
        return "1900-1919"
    if value < 1940:
        return "1920-1939"
    if value < 1960:
        return "1940-1959"
    if value < 1980:
        return "1960-1979"
    if value < 2000:
        return "1980-1999"
    if value < 2020:
        return "2000-2019"
    return "2020后"


def stats_visual(items: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    rows = items or []
    order = ["1900前", "1900-1919", "1920-1939", "1940-1959", "1960-1979", "1980-1999", "2000-2019", "2020后", "未知"]
    year_counts = Counter(bucket_year(row.get("year")) for row in rows)

    def top(key: str, limit: int) -> list[list[Any]]:
        counts = Counter(str(row.get(key) or "未知") for row in rows)
        return [[name, count] for name, count in counts.most_common(limit)]

    return {
        "yearBuckets": [[name, year_counts.get(name, 0)] for name in order],
        "languageTop": top("language", 7),
        "countryTop": top("country", 7),
        "authorTop": top("translator", 8),
        "storyTop": top("canonicalName", 12),
        "carrierTop": top("carrier", 5),
        "modeTop": top("translationMode", 5),
    }


def _normalize_atlas_mode(mode: str = "all") -> str:
    value = str(mode or "all").lower()
    return value if value in {"all", "collections", "prefaces", "children"} else "all"


@lru_cache(maxsize=12)
def _visual_atlas_cached(mode: str, mtime_ns: int, size: int, map_mtime_ns: int, map_size: int) -> dict[str, Any]:
    data = _story_data_cached(mtime_ns, size)
    collections = data["collections"]
    charts: dict[str, Any] = {}

    if mode in {"all", "collections"}:
        map_data = _publication_source_map_cached(map_mtime_ns, map_size)
        map_records = map_data.get("records") or []
        map_flows = map_data.get("flows") or []
        flows = map_flows or data["flows"]
        publication_rows = map_records or publication_workbook_rows() or collections
        publication_chart = publication_map(publication_rows)
        if map_records:
            publication_chart["subtitle"] = "圆点越大表示出版城市越活跃；当前地图优先使用《地图_中国故事集_出版地和故事来源地.xlsx》生成的出版地数据。"
            publication_chart["dataSource"] = map_data.get("sourceWorkbook") or PUBLICATION_MAP_WORKBOOK
        elif publication_workbook_rows():
            publication_chart["subtitle"] = "圆点越大表示出版城市越活跃；当前地图优先使用《地图_中国故事集_出版地和故事来源地.xlsx》。"
            publication_chart["dataSource"] = PUBLICATION_MAP_WORKBOOK
        wilhelm = wilhelm_rows(collections)
        charts.update(
            {
                "identityProcess": identity_process(collections),
                "identityRiver": identity_river(collections),
                "publicationMap": publication_chart,
                "sourceMap": source_map(flows, map_records),
                "wilhelmPublicationMap": publication_map(wilhelm, "卫礼贤《中国民间童话》再版出版地图"),
            }
        )

    if mode in {"all", "prefaces"}:
        charts.update(
            {
                "prefaceCluster": preface_cluster(collections),
                "wordCloud": word_cloud(collections),
            }
        )

    if mode in {"all", "children"}:
        charts["childCooccurrence"] = child_cooccurrence(collections, data["childStories"])

    return {
        "stats": data.get("stats", {}),
        "charts": charts,
    }


def visual_atlas(mode: str = "all") -> dict[str, Any]:
    return _visual_atlas_cached(_normalize_atlas_mode(mode), *_story_data_cache_key(), *_publication_source_cache_key())


def collection_graph(collection_id: str) -> dict[str, Any]:
    data = story_data()
    collections = data["collections"]
    children = data["childStories"]
    child_by_id = {item["id"]: item for item in children}
    collection = next((item for item in collections if item["id"] == collection_id), collections[0])
    flow = next((item for item in data["flows"] if item["id"] == collection["id"]), None)
    selected_children = [child_by_id[child_id] for child_id in collection.get("matchedChildIds", []) if child_id in child_by_id]
    nodes: dict[str, dict[str, Any]] = {}
    edges: list[dict[str, Any]] = []

    def node(node_id: str, **payload: Any) -> None:
        current = nodes.get(node_id)
        nodes[node_id] = {**(current or {}), "id": node_id, **payload}

    center_id = f"collection-{collection['id']}"
    node(center_id, label=collection.get("chineseTitle") or collection["name"], type="故事集", section="stories", year=collection.get("year"), lang="德语", x=0.48, y=0.22, size=27)
    editor_id = f"editor-{short(collection.get('editor'), 50)}"
    node(editor_id, label=collection.get("editor") or "未记录译者", type="译者/编者", section="stories", year=collection.get("year"), lang="德语", x=0.22, y=0.18, size=18)
    edges.append({"from": editor_id, "to": center_id, "relation": "编译", "note": collection["name"]})
    role_id = f"role-{short(collection.get('editorRole'), 50)}"
    node(role_id, label=collection.get("editorRole") or "未记录身份", type="译者身份", section="stories", year=collection.get("year"), lang="中文", x=0.1, y=0.26, size=15)
    edges.append({"from": role_id, "to": editor_id, "relation": "身份", "note": "译者身份"})
    publisher_id = f"publisher-{short(collection.get('publisher'), 60)}"
    node(publisher_id, label=collection.get("publisher") or "未记录出版社", type="出版社", section="stories", year=collection.get("year"), lang="德语", x=0.76, y=0.18, size=17)
    edges.append({"from": center_id, "to": publisher_id, "relation": "出版", "note": collection.get("publisher")})
    if collection.get("prefaceAuthor") and collection["prefaceAuthor"] != "/":
        preface_id = f"preface-{short(collection['prefaceAuthor'], 60)}"
        node(preface_id, label=collection["prefaceAuthor"], type="序跋作者", section="stories", year=collection.get("year"), lang="中文", x=0.74, y=0.32, size=15)
        edges.append({"from": preface_id, "to": center_id, "relation": "阐释", "note": "序跋作者如何塑造故事阐释"})
    if flow and flow.get("province"):
        province_id = f"province-{flow['province']}"
        node(province_id, label=flow["province"], type="取材来源", section="stories", year=collection.get("year"), lang="中文", x=0.5, y=0.38, size=15)
        edges.append({"from": province_id, "to": center_id, "relation": "来源", "note": "故事取材来源"})

    motif_counts = Counter((child.get("canonicalName") or child.get("variantName") or "未记录") for child in selected_children)
    motifs = {name: f"motif-{short(name, 60)}" for name, _ in motif_counts.most_common(28)}
    for index, (name, count) in enumerate(motif_counts.most_common(28)):
        node(motifs[name], label=short(name, 12), type="故事母题", section="stories", year=collection.get("year"), lang="中文", x=0.1 + (index % 7) * 0.11, y=0.48 + (index // 7) * 0.07, size=10 + min(10, count))
    for index, child in enumerate(selected_children[:80]):
        child_id = f"child-{child['id']}"
        node(child_id, label=short(child.get("variantName") or child.get("canonicalName"), 10), type="子故事", section="stories", year=child.get("year") or collection.get("year"), lang=child.get("language"), x=0.28 + (index % 9) * 0.068, y=min(0.94, 0.58 + (index // 9) * 0.041), size=7)
        edges.append({"from": center_id, "to": child_id, "relation": "收录", "note": child.get("bookName")})
        motif_id = motifs.get(child.get("canonicalName") or child.get("variantName") or "未记录")
        if motif_id:
            edges.append({"from": motif_id, "to": child_id, "relation": "母题关联", "note": child.get("variantName")})
    return {
        "collection": collection,
        "children": selected_children,
        "flow": flow,
        "graph": {"nodes": list(nodes.values()), "edges": edges},
    }
